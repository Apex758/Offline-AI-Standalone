# Enforce UTF-8 encoding for all std streams and environment
import os
import sys
import uuid
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

# Set environment variable for all subprocesses
os.environ["PYTHONIOENCODING"] = "utf-8"

# Reconfigure stdout/stderr if possible (Python 3.7+)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, Request, File, UploadFile, Form
from fastapi.responses import JSONResponse, Response, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from image_service import get_image_service
import base64
from pydantic import BaseModel
from typing import List, Optional
import subprocess
import io
import logging
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import threading
import time
import json
import asyncio
import signal
import atexit
from config import (
    get_model_path, get_selected_model, set_selected_model,
    LLAMA_CLI_PATH, LLAMA_PARAMS, CORS_ORIGINS, MODELS_DIR,
    MODEL_PATH, MODEL_VERBOSE, MODEL_N_CTX, MODEL_MAX_TOKENS, MODEL_TEMPERATURE,
    IMAGE_MODELS_DIR, get_selected_diffusion_model, set_selected_diffusion_model,
    get_diffusion_model_path, scan_diffusion_models, get_image_model_info, get_image_model_path,
    get_tier_config, set_tier_config, get_model_tier, compute_effective_tier,
    LAMA_MODEL_PATH,
)
from pathlib import Path
from datetime import datetime
from routes import milestones, achievements, insights
from apscheduler.schedulers.asyncio import AsyncIOScheduler

_scheduler: AsyncIOScheduler | None = None
import student_service
from llama_inference import LlamaInference
from process_pool import submit_task, shutdown_executor
from llama_inference import run_llama_inference
from curriculum_matcher import CurriculumMatcher
from chat_memory import get_chat_memory
from metrics_service import get_metrics_collector
sys.stdout.reconfigure(encoding='utf-8')

# ── Ring-buffer log handler (captures recent logs for support tickets) ──
class _RingBufferHandler(logging.Handler):
    """Keeps the last N log records in memory for support diagnostics."""
    def __init__(self, capacity: int = 200):
        super().__init__()
        from collections import deque
        self._buffer: "deque[dict]" = deque(maxlen=capacity)

    def emit(self, record: logging.LogRecord):
        try:
            self._buffer.append({
                "ts": self.format(record).split(" ")[0] if False else
                      datetime.utcnow().isoformat(timespec="milliseconds"),
                "level": record.levelname,
                "logger": record.name,
                "msg": record.getMessage()[:500],  # cap long messages
            })
        except Exception:
            pass

    def recent(self, limit: int = 100) -> list:
        items = list(self._buffer)
        return items[-limit:]

_log_ring = _RingBufferHandler(capacity=200)
_log_ring.setLevel(logging.DEBUG)
logging.root.addHandler(_log_ring)

# Set up logging
logger = logging.getLogger(__name__)

# Global set to track all active subprocess processes
active_processes = set()
process_lock = threading.Lock()

def register_process(process):
    """Register a subprocess for cleanup tracking"""
    with process_lock:
        active_processes.add(process)

def unregister_process(process):
    """Unregister a subprocess after it's been cleaned up"""
    with process_lock:
        active_processes.discard(process)

def cleanup_process(process):
    """Safely cleanup a subprocess and all its children"""
    if not process or process.poll() is not None:
        unregister_process(process)
        return
    
    pid = process.pid
    
    try:
        process.terminate()
        
        try:
            process.wait(timeout=2)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait(timeout=1)
        
        if os.name == 'nt':
            try:
                subprocess.run(
                    ['taskkill', '/F', '/T', '/PID', str(pid)],
                    capture_output=True,
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
            except Exception as e:
                logger.error(f"Error killing process tree: {e}")
    
    except Exception as e:
        logger.error(f"Error cleaning up process PID {pid}: {e}")
    finally:
        unregister_process(process)

def cleanup_all_processes():
    """Cleanup all tracked processes"""
    with process_lock:
        processes_to_clean = list(active_processes)
    
    for process in processes_to_clean:
        cleanup_process(process)

# Register cleanup function to run on exit
atexit.register(cleanup_all_processes)

# Handle termination signals
def signal_handler(signum, frame):
    cleanup_all_processes()
    sys.exit(0)

if hasattr(signal, 'SIGTERM'):
    signal.signal(signal.SIGTERM, signal_handler)
if hasattr(signal, 'SIGINT'):
    signal.signal(signal.SIGINT, signal_handler)


def get_data_directory():
    """Get user-writable data directory"""
    if os.name == 'nt':  # Windows
        app_data = os.environ.get('APPDATA', os.path.expanduser('~'))
        data_dir = Path(app_data) / 'OECS Learning Hub' / 'data'
    else:  # macOS/Linux
        data_dir = Path.home() / '.olh_ai_education' / 'data'
    
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir

_json_cache = {}

def load_json_data(filename: str):
    """Load data from a JSON file with in-memory cache"""
    if filename in _json_cache:
        return _json_cache[filename]

    filepath = get_data_directory() / filename

    if not filepath.exists():
        with open(filepath, 'w') as f:
            json.dump([], f)
        _json_cache[filename] = []
        return []

    try:
        with open(filepath, 'r') as f:
            data = json.load(f)
        _json_cache[filename] = data
        return data
    except json.JSONDecodeError:
        _json_cache[filename] = []
        return []

def save_json_data(filename: str, data):
    """Save data to a JSON file and update cache"""
    filepath = get_data_directory() / filename

    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)
    _json_cache[filename] = data
        
# (Create app only once, after defining lifespan)

# Path for chat history storage
CHAT_HISTORY_FILE = os.path.join(os.path.dirname(__file__), "chat_history.json")

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app):
    global _scheduler
    # Start APScheduler
    _scheduler = AsyncIOScheduler()
    _scheduler.start()
    # Re-apply any saved insights schedule
    try:
        from routes.insights import _load_schedule, _apply_schedule
        _apply_schedule(_load_schedule())
    except Exception as e:
        logger.error(f"Failed to restore insights schedule: {e}")

    # Startup logic
    logger.info("Checking for required files...")
    if not os.path.exists(get_model_path()):
        logger.warning(f"Model file not found at {get_model_path()}")
    else:
        logger.info(f"Model file found: {get_model_path()}")
    
    if not os.path.exists(LLAMA_CLI_PATH):
        logger.warning(f"llama-cli.exe not found at {LLAMA_CLI_PATH}")
    else:
        logger.info(f"llama-cli.exe found: {LLAMA_CLI_PATH}")
    
    # Initialize SQLite chat memory and migrate old JSON data if present
    memory = get_chat_memory()
    memory.import_from_json(CHAT_HISTORY_FILE)
    logger.info(f"Chat memory initialized (SQLite): {memory.db_path}")
    
    # NOTE: Llama model is now loaded LAZILY via /api/model/preload endpoint
    # to speed up startup time. Model loads on first tab that needs AI.

    # Initialize CurriculumMatcher singleton
    global curriculum_matcher
    try:
        curriculum_matcher = CurriculumMatcher()
        logger.info("CurriculumMatcher initialized successfully")
    except Exception as e:
        curriculum_matcher = None
        logger.error(f"Failed to initialize CurriculumMatcher: {e}")

    # Initialize MilestoneDB
    from milestones.milestone_db import get_milestone_db
    try:
        milestone_db = get_milestone_db()
        logger.info("Milestone database initialized")
    except Exception as e:
        logger.error(f"Failed to initialize milestone database: {e}")

    # Initialize Student DB
    try:
        student_service.init_db()
        logger.info("Student database initialized")
    except Exception as e:
        logger.error(f"Failed to initialize student database: {e}")

    # Initialize Achievement DB
    try:
        import achievement_service
        achievement_service.init_db()
        logger.info("Achievement database initialized")
    except Exception as e:
        logger.error(f"Failed to initialize achievement database: {e}")

    # Initialize Image Service
    try:
        from image_service import get_image_service
        image_service = get_image_service()
        logger.info("Image service initialized")

        # Start IOPaint on startup
        image_service.start_iopaint()
    except Exception as e:
        logger.error(f"Failed to initialize image service: {e}")

    logger.info("Server ready!")
    yield

    # Shutdown logic
    try:
        from inference_factory import get_inference_instance
        inference = get_inference_instance()
        inference.cleanup()
        logger.info("Inference backend cleaned up")
    except Exception as e:
        logger.error(f"Error cleaning up inference backend: {e}")
        
    try:
        from image_service import get_image_service
        image_service = get_image_service()
        image_service.cleanup()
        logger.info("Image service cleaned up")
    except Exception as e:
        logger.error(f"Error cleaning up image service: {e}")
        
    cleanup_all_processes()
    shutdown_executor()
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)

app = FastAPI(lifespan=lifespan)

# Add CORS middleware FIRST, before routers
# allow_origins=["*"] is needed so phones on the local network (e.g.
# http://192.168.x.x:8000) can hit the API for Photo Transfer.
# This is safe because the server is only reachable on the LAN.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Catch-all exception handler so error responses go through CORSMiddleware
@app.exception_handler(Exception)
async def _global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled error on {request.method} {request.url.path}: {exc}", exc_info=True)
    return JSONResponse(status_code=500, content={"detail": str(exc)})

# Include milestone routes
app.include_router(milestones.router)

# Include achievement routes
app.include_router(achievements.router)

# Include educator insights routes
app.include_router(insights.router)

# Include scene-based image generation routes
from scene_api_endpoints import router as scene_router
app.include_router(scene_router)

# Include photo transfer routes (phone-to-PC via local WiFi)
from routes.photo_transfer import router as photo_transfer_router
app.include_router(photo_transfer_router)

# Include school year calendar routes
from routes.school_year import router as school_year_router
app.include_router(school_year_router)


# ── Serve the phone PWA page at /phone ──────────────────────────────────────
@app.get("/phone", response_class=HTMLResponse)
async def serve_phone_page():
    """Serve the phone camera/upload page (teachers scan QR to reach this)."""
    phone_html = Path(BASE_DIR) / "static" / "phone.html"
    if phone_html.exists():
        return HTMLResponse(phone_html.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>Phone page not found</h1>", status_code=404)

# Suppress noisy polling endpoints from uvicorn access logs
class _QuietPollFilter(logging.Filter):
    """Hide repetitive polling requests from the access log."""
    _quiet_paths = {"/api/metrics/live-stats", "/api/logs/recent"}
    def filter(self, record: logging.LogRecord) -> bool:
        msg = record.getMessage()
        return not any(p in msg for p in self._quiet_paths)

logging.getLogger("uvicorn.access").addFilter(_QuietPollFilter())

# CORS middleware is registered above, right after app creation


class LoginRequest(BaseModel):
    username: str
    password: str

class ChatMessage(BaseModel):
    message: str
    conversation_history: Optional[List[dict]] = []

class Message(BaseModel):
    id: str
    role: str
    content: str
    timestamp: str
    attachments: Optional[List[dict]] = None

class ChatHistory(BaseModel):
    id: str
    title: str
    timestamp: str
    messages: List[Message]

class TitleGenerateRequest(BaseModel):
    user_message: str
    assistant_message: str

class TitleGenerateResponse(BaseModel):
    title: str
    generated: bool
    fallback: bool
    generationTime: float

class AutocompleteRequest(BaseModel):
    text: str
    max_tokens: int = 20

class OrganizeNoteRequest(BaseModel):
    content: str

class SmartSearchRequest(BaseModel):
    query: str

class StudentCreate(BaseModel):
    full_name: str
    date_of_birth: Optional[str] = None
    class_name: Optional[str] = None
    grade_level: Optional[str] = None
    gender: Optional[str] = None
    contact_info: Optional[str] = None

class QuizGradeCreate(BaseModel):
    student_id: str
    quiz_title: str
    subject: str
    score: float
    total_points: float
    percentage: float
    letter_grade: str
    answers: Optional[dict] = {}

class WorksheetGradeCreate(BaseModel):
    student_id: str
    worksheet_title: str
    subject: str
    score: float
    total_points: float
    percentage: float
    letter_grade: str
    answers: Optional[dict] = {}

class AttendanceRecord(BaseModel):
    student_id: str
    class_name: str
    date: str
    status: str
    engagement_level: str
    notes: Optional[str] = ''

class AttendanceSave(BaseModel):
    records: list[AttendanceRecord]


# (Removed redundant OPTIONS handler; CORS middleware handles preflight)
    
    
@app.post("/api/login")
async def login(credentials: dict):
    username = credentials.get("username")
    password = credentials.get("password")
    
    return {
        "success": True,
        "user": {
            "username": username,
            "name": username.replace("_", " ").title()  
        },
        "token": "some_token_here"
    }

@app.get("/api/chat-history")
async def get_chat_history():
    """Get all chat histories from SQLite"""
    try:
        memory = get_chat_memory()
        return memory.get_all_chats_with_messages()
    except Exception as e:
        logger.error(f"Error loading chat history: {e}")
        return []

@app.post("/api/chat-history")
async def save_chat_history(chat: ChatHistory):
    """Save or update a chat history in SQLite"""
    try:
        memory = get_chat_memory()
        memory.ensure_chat(chat.id, chat.title)
        memory.update_chat_title(chat.id, chat.title)
        memory.save_messages_bulk(chat.id, [m.dict() for m in chat.messages])
        return {"success": True, "id": chat.id}
    except Exception as e:
        logger.error(f"Error saving chat history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/chat-history/{chat_id}")
async def delete_chat_history(chat_id: str):
    """Delete a chat history from SQLite"""
    try:
        memory = get_chat_memory()
        memory.delete_chat(chat_id)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error deleting chat history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Shared Prompt Builder — model-aware formatting
# ============================================================================

def get_prompt_format() -> str:
    """Get the prompt format for the currently loaded model."""
    try:
        from inference_factory import get_inference_instance
        inference = get_inference_instance()
        return getattr(inference, 'model_config', {}).get('prompt_format', 'llama')
    except Exception:
        return 'llama'

def build_prompt(system_prompt: str, user_prompt: str, prompt_format: str = None) -> str:
    """Build a single-turn prompt in the correct format for the loaded model."""
    if prompt_format is None:
        prompt_format = get_prompt_format()
    if prompt_format == "chatml":
        return (
            f"<|im_start|>system\n{system_prompt}<|im_end|>\n"
            f"<|im_start|>user\n{user_prompt}<|im_end|>\n"
            f"<|im_start|>assistant\n"
        )
    else:
        return (
            f"<|begin_of_text|>"
            f"<|start_header_id|>system<|end_header_id|>\n\n{system_prompt}<|eot_id|>"
            f"<|start_header_id|>user<|end_header_id|>\n\n{user_prompt}<|eot_id|>"
            f"<|start_header_id|>assistant<|end_header_id|>\n\n"
        )

# ============================================================================
# Title Generation Helper Functions
# ============================================================================

def build_title_prompt(user_msg: str, assistant_msg: str) -> str:
    from tier1_prompts import get_tier1_system_prompt
    _tier_info = compute_effective_tier()

    if _tier_info["tier"] == 1:
        system_prompt = get_tier1_system_prompt("title-generation")
    else:
        system_prompt = """You are a title generation assistant. Create concise, descriptive titles for chat conversations.

Rules:
- Maximum 60 characters
- Use title case
- Be specific and descriptive
- Capture the main topic or question
- No special characters except hyphens and ampersands
- No quotes or punctuation at the end
- Focus on the key concept or action"""

    user_prompt = f"""Based on this conversation, create a concise title (max 60 characters):

User: {user_msg[:200]}
Assistant: {assistant_msg[:200]}

Generate only the title, nothing else."""

    return build_prompt(system_prompt, user_prompt)


def clean_title_text(raw_title: str) -> str:
    import re
    
    prefixes_to_remove = [
        r'^Title:\s*',
        r'^Here\'s a title:\s*',
        r'^Here is a title:\s*',
        r'^A title:\s*',
        r'^Chat title:\s*',
        r'^Conversation title:\s*',
    ]
    
    title = raw_title.strip()
    
    for prefix in prefixes_to_remove:
        title = re.sub(prefix, '', title, flags=re.IGNORECASE)
    
    title = title.strip('"\'""''')
    title = re.sub(r'[.,:;!?]+$', '', title)
    
    max_length = 60
    if len(title) > max_length:
        truncated = title[:max_length]
        last_space = truncated.rfind(' ')
        if last_space > 30:
            title = truncated[:last_space] + "..."
        else:
            title = truncated[:max_length-3] + "..."
    
    return title.strip()


def create_fallback_title(user_message: str) -> str:
    from datetime import datetime
    
    if user_message and len(user_message.strip()) > 0:
        title = user_message.strip()
        title = ' '.join(title.split())
        
        max_length = 60
        if len(title) > max_length:
            truncated = title[:max_length]
            last_space = truncated.rfind(' ')
            if last_space > 20:
                title = truncated[:last_space] + "..."
            else:
                title = truncated[:max_length-3] + "..."
        
        return title
    
    now = datetime.now()
    return f"Chat - {now.strftime('%I:%M %p')}"


def validate_title(title: str) -> bool:
    if not title or len(title.strip()) == 0:
        return False
    
    if len(title) > 63:  # 60 + "..."
        return False
    
    if len(title.strip()) < 3:
        return False
    
    return True

# ============================================================================
# Title Generation Endpoint
# ============================================================================

@app.post("/api/generate-title")
async def generate_title(request: TitleGenerateRequest):
    """Generate a concise title for a chat conversation"""
    start_time = time.time()
    
    try:
        prompt = build_title_prompt(request.user_message, request.assistant_message)
        
        from inference_factory import get_inference_instance, resolve_inference_for_task
        inference = resolve_inference_for_task("title-generation")
        result = await inference.generate(
            tool_name="title_generation",
            input_data=request.user_message,
            prompt_template=prompt,
            max_tokens=30,
            temperature=0.5
        )
        
        if result["metadata"]["status"] == "error":
            raise ValueError(result["metadata"].get("error_message", "Generation failed"))
        
        raw_title = result["result"]
        title = clean_title_text(raw_title)
        
        if validate_title(title):
            generation_time = time.time() - start_time
            return TitleGenerateResponse(
                title=title,
                generated=True,
                fallback=False,
                generationTime=generation_time
            )
        else:
            raise ValueError("Generated title failed validation")
        
    except Exception as e:
        logger.error(f"Title generation error: {e}")
        title = create_fallback_title(request.user_message)
        generation_time = time.time() - start_time
        
        return TitleGenerateResponse(
            title=title,
            generated=False,
            fallback=True,
            generationTime=generation_time
        )


@app.post("/api/autocomplete")
async def autocomplete(request: AutocompleteRequest):
    """Generate a short text completion for the auto-finish sentence feature."""
    try:
        text = request.text.strip()
        if not text or len(text) < 10:
            return {"completion": ""}

        prompt = f"Continue this text naturally with a few more words. Only output the continuation, nothing else:\n\n{text}"

        from inference_factory import get_inference_instance, resolve_inference_for_task
        from tier1_prompts import get_tier1_gen_params
        _tier_info = compute_effective_tier()
        _is_tier1 = _tier_info["tier"] == 1
        _t1_params = get_tier1_gen_params("autocomplete") if _is_tier1 else {}

        inference = resolve_inference_for_task("autocomplete")
        result = await inference.generate(
            tool_name="autocomplete",
            input_data=text,
            prompt_template=prompt,
            max_tokens=min(request.max_tokens, 30),
            temperature=_t1_params.get("temperature", 0.7)
        )

        if result["metadata"]["status"] == "error":
            return {"completion": ""}

        completion = result["result"].strip()
        # Remove any text that repeats the original input
        if completion.lower().startswith(text.lower()[-20:]):
            completion = completion[len(text[-20:]):]

        # Clean up: only return a short, clean continuation
        # Stop at sentence boundaries
        for stop_char in ['.', '!', '?', '\n']:
            idx = completion.find(stop_char)
            if idx >= 0:
                completion = completion[:idx + 1]
                break

        # Limit length
        words = completion.split()
        if len(words) > 15:
            completion = ' '.join(words[:15])

        return {"completion": completion}

    except Exception as e:
        logger.error(f"Autocomplete error: {e}")
        return {"completion": ""}


@app.post("/api/organize-note")
async def organize_note(request: OrganizeNoteRequest):
    """Use AI to organize and structure the content of a sticky note."""
    try:
        content = request.content.strip()
        if not content:
            return {"organized": content, "success": False}

        # Strip HTML tags to get plain text for the prompt
        import re
        plain = re.sub(r'<[^>]*>', ' ', content)
        plain = re.sub(r'\s+', ' ', plain).strip()

        prompt = (
            "You are organizing a sticky note. Rewrite the following messy notes into clean, "
            "well-structured HTML suitable for a rich-text sticky note. Group related ideas under "
            "short bold headings (<b>), use bullet lists (<ul><li>) for items, and keep it concise. "
            "Only output the organized HTML, no explanation.\n\n"
            f"Notes:\n{plain}"
        )

        from inference_factory import resolve_inference_for_task
        inference = resolve_inference_for_task("organize-note")
        result = await inference.generate(
            tool_name="organize_note",
            input_data=plain,
            prompt_template=prompt,
            max_tokens=500,
            temperature=0.3
        )

        if result["metadata"]["status"] == "error":
            return {"organized": content, "success": False}

        organized = result["result"].strip()
        return {"organized": organized, "success": True}

    except Exception as e:
        logger.error(f"Organize note error: {e}")
        return {"organized": content, "success": False}


@app.post("/api/smart-search")
async def smart_search(request: SmartSearchRequest):
    """AI-powered search that returns structured guidance for the command palette."""
    import json as _json
    import re as _re
    try:
        query = request.query.strip()
        if not query or len(query) < 3:
            return {"intent": "info", "summary": "", "steps": [], "confidence": 0}

        from inference_factory import resolve_inference_for_task
        from smart_search_prompt import build_smart_search_prompt, SMART_SEARCH_TIER1_PROMPT
        from tier1_prompts import get_tier1_gen_params

        _tier_info = compute_effective_tier()
        _is_tier1 = _tier_info["tier"] == 1

        system_prompt = SMART_SEARCH_TIER1_PROMPT if _is_tier1 else build_smart_search_prompt(query)
        user_prompt = f"Teacher asks: {query}"

        prompt = build_prompt(system_prompt, user_prompt)

        inference = resolve_inference_for_task("smart-search")
        _t1_params = get_tier1_gen_params("autocomplete") if _is_tier1 else {}

        result = await inference.generate(
            tool_name="smart-search",
            input_data=query,
            prompt_template=prompt,
            max_tokens=300,
            temperature=_t1_params.get("temperature", 0.3)
        )

        if result["metadata"]["status"] == "error":
            return {"intent": "info", "summary": "Sorry, I couldn't process that.", "steps": [], "confidence": 0}

        raw = result["result"].strip()
        logger.info(f"[SmartSearch] Raw response: {raw[:500]}")

        # Extract JSON from the response (model may wrap it in markdown code blocks)
        json_match = _re.search(r'\{[\s\S]*\}', raw)
        if json_match:
            json_str = json_match.group()
            # Try to fix common model errors before parsing
            # Some models output "step1": "..." instead of array
            try:
                data = _json.loads(json_str)
            except _json.JSONDecodeError:
                # Try fixing common issues: trailing commas, unescaped quotes
                cleaned = _re.sub(r',\s*}', '}', json_str)
                cleaned = _re.sub(r',\s*]', ']', cleaned)
                try:
                    data = _json.loads(cleaned)
                except _json.JSONDecodeError:
                    data = None

            if data and isinstance(data, dict):
                # Normalize steps — handle both array and dict formats
                raw_steps = data.get("steps", [])
                if isinstance(raw_steps, dict):
                    # Model returned {"step1": "...", "step2": "..."} instead of array
                    steps = [v for k, v in sorted(raw_steps.items()) if isinstance(v, str)]
                elif isinstance(raw_steps, list):
                    steps = [s for s in raw_steps if isinstance(s, str)]
                else:
                    steps = []

                # Normalize intent — model may output the template "navigation|generation|..."
                intent = data.get("intent", "info")
                if "|" in str(intent):
                    intent = "info"

                summary = data.get("summary", "")

                response = {
                    "intent": intent,
                    "summary": summary,
                    "steps": steps,
                    "confidence": float(data.get("confidence", 0.7)),
                }
                if "action" in data and isinstance(data["action"], dict):
                    action = data["action"]
                    # Clean up action — remove template values
                    if action.get("toolType") and "|" not in action["toolType"]:
                        response["action"] = action
                return response

        # Fallback: return the raw text as a summary
        return {
            "intent": "info",
            "summary": raw[:200] if raw else "I couldn't understand that query.",
            "steps": [],
            "confidence": 0.3,
        }

    except Exception as e:
        logger.error(f"Smart search error: {e}")
        return {"intent": "info", "summary": "An error occurred.", "steps": [], "confidence": 0}


@app.websocket("/ws/chat")
async def websocket_chat(websocket: WebSocket):
    await websocket.accept()
    from config import LLAMA_PARAMS
    import os

    def build_multi_turn_prompt(system_prompt, history, user_message, prompt_format="llama"):
        if prompt_format == "chatml":
            # ChatML format (Qwen, etc.)
            prompt = f"<|im_start|>system\n{system_prompt}<|im_end|>\n"
            for msg in history:
                prompt += f"<|im_start|>{msg['role']}\n{msg['content']}<|im_end|>\n"
            prompt += f"<|im_start|>user\n{user_message}<|im_end|>\n"
            prompt += "<|im_start|>assistant\n"
        elif prompt_format == "phi":
            # Phi-4 format
            prompt = f"<|system|>{system_prompt}<|end|>\n"
            for msg in history:
                prompt += f"<|{msg['role']}|>{msg['content']}<|end|>\n"
            prompt += f"<|user|>{user_message}<|end|>\n"
            prompt += "<|assistant|>"
        else:
            # Llama format
            prompt = "<|begin_of_text|>"
            prompt += f"<|start_header_id|>system<|end_header_id|>\n\n{system_prompt}<|eot_id|>"
            for msg in history:
                if msg["role"] == "user":
                    prompt += f"<|start_header_id|>user<|end_header_id|>\n\n{msg['content']}<|eot_id|>"
                elif msg["role"] == "assistant":
                    prompt += f"<|start_header_id|>assistant<|end_header_id|>\n\n{msg['content']}<|eot_id|>"
            prompt += f"<|start_header_id|>user<|end_header_id|>\n\n{user_message}<|eot_id|>"
            prompt += "<|start_header_id|>assistant<|end_header_id|>\n\n"
        return prompt

    try:
        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)
            user_message = message_data.get("message", "")
            chat_id = message_data.get("chat_id", None)
            # Support custom system prompt from AI assistant panel
            custom_system_prompt = message_data.get("system_prompt", None)
            # Support conversation history sent from frontend (for panels without chat_id)
            client_history = message_data.get("history", None)
            # Thinking mode toggle (for Qwen2.5/Qwen3 models)
            thinking_enabled = message_data.get("thinking_enabled", False)
            # Teacher profile context (grade/subject awareness)
            profile_context = message_data.get("profile_context", "")

            if not user_message:
                continue

            default_system_prompt = "You are a helpful AI assistant. Answer questions naturally and conversationally. Keep responses concise but informative. Adapt your detail level to what the user asks - brief for simple questions, detailed for complex topics."

            # Tier-1 awareness: use simpler prompt and tighter params for small models
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            if _is_tier1:
                default_system_prompt = get_tier1_system_prompt("chat")
                _t1_params = get_tier1_gen_params("chat")

            # Use custom system prompt if provided, otherwise use default
            base_system_prompt = custom_system_prompt if custom_system_prompt else default_system_prompt

            # Build context from SQLite (Tier 1 sliding window + Tier 2 summary)
            history = []
            summary_block = ""

            if chat_id:
                try:
                    memory = get_chat_memory()
                    N = LLAMA_PARAMS.get("conversation_history_length", 4)
                    summary_block, history = memory.build_context(chat_id, n_pairs=N)
                except Exception as e:
                    logger.error(f"Error loading chat context: {e}")
            elif client_history:
                # Use conversation history sent from the client (for AI assistant panel)
                history = client_history[-8:]  # Cap at last 8 messages to avoid context overflow

            # Inject summary into system prompt if available
            system_prompt = base_system_prompt + summary_block

            # Inject teacher profile context (grade/subject awareness)
            if profile_context:
                system_prompt += profile_context

            # Curriculum matching: find related curriculum based on user message + profile
            # Skip for short/casual messages (greetings, short phrases) to avoid false matches
            if curriculum_matcher and profile_context and len(user_message.split()) >= 4:
                try:
                    # Extract grade/subject from profile for filtering
                    chat_grade = ""
                    chat_subject = ""
                    profile_grades = message_data.get("profile_grades", [])
                    profile_subjects = message_data.get("profile_subjects", [])
                    if profile_grades:
                        chat_grade = profile_grades[0] if len(profile_grades) == 1 else ""
                    if profile_subjects:
                        chat_subject = profile_subjects[0] if len(profile_subjects) == 1 else ""

                    chat_matches = curriculum_matcher.find_matching_pages(
                        user_message, top_k=3,
                        grade=chat_grade, subject=chat_subject
                    )
                    # Only send refs with a meaningful match score
                    chat_refs = [
                        {
                            "id": m.get("id"),
                            "displayName": m.get("displayName"),
                            "grade": m.get("grade"),
                            "subject": m.get("subject"),
                            "route": m.get("route"),
                            "matchScore": m.get("matchScore"),
                        }
                        for m in chat_matches if m.get("matchScore", 0) >= 0.15
                    ]
                    if chat_refs:
                        await websocket.send_json({
                            "type": "curriculum_refs",
                            "references": chat_refs
                        })
                        # Add curriculum context to system prompt
                        context_blocks = []
                        for m in chat_matches:
                            if m.get("matchScore", 0) >= 0.15:
                                ctx = curriculum_matcher.get_curriculum_context(m.get("id"))
                                if ctx:
                                    context_blocks.append(ctx)
                        if context_blocks:
                            system_prompt += "\n\nRelated Curriculum:\n" + "\n\n".join(context_blocks)
                except Exception as e:
                    logger.error(f"Error matching curriculum in chat: {e}")

            # Separate image files from text files
            reference_files = message_data.get("reference_files", None)
            image_files = []
            text_files = []
            if reference_files and len(reference_files) > 0:
                for ref in reference_files[:5]:
                    if ref.get("is_image") and ref.get("base64"):
                        image_files.append(ref)
                    else:
                        text_files.append(ref)

            # Inject text reference files into the user message
            effective_user_message = user_message
            if text_files:
                file_context_parts = []
                for ref in text_files:
                    name = ref.get("name", "unknown")
                    content = ref.get("content", "")
                    if len(content) > 8000:
                        content = content[:8000] + "\n... [truncated]"
                    file_context_parts.append(f"--- File: {name} ---\n{content}\n--- End of {name} ---")
                file_context = "\n\n".join(file_context_parts)
                effective_user_message = f"The user has attached the following reference files:\n\n{file_context}\n\nUser's message: {user_message}"
                logger.info(f"Attached {len(text_files)} text file(s) to chat message")

            if image_files:
                logger.info(f"Attached {len(image_files)} image(s) to chat message — using vision model")

            from inference_factory import get_inference_instance, resolve_inference_for_task
            inference = resolve_inference_for_task("chat")

            # Detect prompt format and thinking support from the loaded model
            prompt_fmt = getattr(inference, 'model_config', {}).get('prompt_format', 'llama')
            model_supports_thinking = getattr(inference, 'model_config', {}).get('supports_thinking', False)

            # Inject /think or /no_think directive for thinking-capable models
            if model_supports_thinking:
                thinking_directive = "/think" if thinking_enabled else "/no_think"
                system_prompt = system_prompt + f"\n{thinking_directive}"

            prompt = build_multi_turn_prompt(system_prompt, history, effective_user_message, prompt_format=prompt_fmt)

            # Safety: truncate prompt if it's too long to prevent llama.cpp crash
            # Rough estimate: 1 token ≈ 4 chars. Leave room for max_tokens response.
            max_prompt_chars = (LLAMA_PARAMS.get("context_size", 16384) - LLAMA_PARAMS["max_tokens"]) * 4
            if len(prompt) > max_prompt_chars:
                logger.warning(f"Prompt too long ({len(prompt)} chars, max {max_prompt_chars}). Truncating file content.")
                if text_files:
                    file_context_parts = []
                    per_file_limit = min(2000, max_prompt_chars // (len(text_files) + 4))
                    for ref in text_files:
                        name = ref.get("name", "unknown")
                        content = ref.get("content", "")
                        if len(content) > per_file_limit:
                            content = content[:per_file_limit] + "\n... [truncated to fit context]"
                        file_context_parts.append(f"--- File: {name} ---\n{content}\n--- End of {name} ---")
                    file_context = "\n\n".join(file_context_parts)
                    effective_user_message = f"The user has attached the following reference files:\n\n{file_context}\n\nUser's message: {user_message}"
                    prompt = build_multi_turn_prompt(system_prompt, history, effective_user_message, prompt_format=prompt_fmt)

            logger.info(f"Context: chat_id={chat_id}, {len(history)} messages in window, summary={'yes' if summary_block else 'no'}, custom_prompt={'yes' if custom_system_prompt else 'no'}, prompt_len={len(prompt)}, images={len(image_files)}, format={prompt_fmt}")

            try:

                # Choose streaming method based on whether images are attached
                token_buffer = []
                full_response_tokens = []  # Accumulate full response for memory
                last_send = time.time()

                if image_files and hasattr(inference, 'generate_stream_vision') and getattr(inference, 'has_vision', False):
                    # Vision path: use create_chat_completion with images
                    vision_system = (
                        "You are a helpful AI assistant with vision capabilities. "
                        "When the user shares an image, describe what you actually see in concrete detail: "
                        "objects, text, colors, layout, people, and context. Be specific and literal — "
                        "say exactly what is in the image rather than guessing abstractly. "
                        "If there is text or numbers visible in the image, read them out. "
                        "Answer the user's question about the image directly."
                    )
                    vision_messages = []
                    vision_messages.append({"role": "system", "content": vision_system})
                    for msg in history:
                        vision_messages.append({"role": msg["role"], "content": msg["content"]})
                    vision_messages.append({"role": "user", "content": effective_user_message, "_is_current": True})

                    images_b64 = [img["base64"] for img in image_files[:2]]

                    # Try vision — if the projector fails at runtime, fall back to text-only
                    try:
                        # Test vision by creating the generator (errors surface on first use)
                        stream_gen = inference.generate_stream_vision(
                            messages=vision_messages,
                            images_base64=images_b64,
                            max_tokens=LLAMA_PARAMS["max_tokens"],
                            temperature=0.4,
                        )
                        # Peek at first chunk to detect errors early
                        first_chunk = None
                        async for first_chunk in stream_gen:
                            break
                        if first_chunk and first_chunk.get("error"):
                            raise RuntimeError(first_chunk["error"])

                        # Vision works — wrap the rest of the stream with the first chunk
                        async def _prepend_stream(first, rest):
                            if first and first.get("token"):
                                yield first
                            async for chunk in rest:
                                yield chunk
                        stream_gen = _prepend_stream(first_chunk, stream_gen)
                    except Exception as vision_err:
                        logger.warning(f"Vision failed, falling back to text-only: {vision_err}")
                        image_names = ", ".join(img.get("name", "image") for img in image_files)
                        effective_user_message = f"[The user attached image(s): {image_names}, but vision is not supported by this model. Respond to their text message only.]\n\nUser's message: {user_message}"
                        prompt = build_multi_turn_prompt(system_prompt, history, effective_user_message, prompt_format=prompt_fmt)
                        stream_gen = inference.generate_stream(
                            tool_name="chat",
                            input_data=user_message,
                            prompt_template=prompt,
                            max_tokens=_t1_params["max_tokens"] if _is_tier1 else LLAMA_PARAMS["max_tokens"],
                            temperature=_t1_params["temperature"] if _is_tier1 else LLAMA_PARAMS["temperature"],
                            repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else LLAMA_PARAMS.get("repeat_penalty", 1.1),
                        )
                elif image_files:
                    # Vision not available — tell the user
                    try:
                        await websocket.send_json({
                            "type": "token",
                            "content": "I can see you've attached an image, but vision support is not available with the current model configuration. I can only process text-based files (PDF, Word, Excel, etc.)."
                        })
                        await websocket.send_json({"type": "done"})
                    except:
                        pass
                    continue
                else:
                    # Text-only path: use raw prompt completion
                    # Double max_tokens when thinking is enabled (think block uses extra tokens)
                    _base_max = _t1_params["max_tokens"] if _is_tier1 else LLAMA_PARAMS["max_tokens"]
                    effective_max_tokens = _base_max * 2 if (thinking_enabled and model_supports_thinking) else _base_max
                    stream_gen = inference.generate_stream(
                        tool_name="chat",
                        input_data=user_message,
                        prompt_template=prompt,
                        max_tokens=effective_max_tokens,
                        temperature=_t1_params["temperature"] if _is_tier1 else LLAMA_PARAMS["temperature"],
                        repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else LLAMA_PARAMS.get("repeat_penalty", 1.1),
                    )

                async for chunk in stream_gen:
                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk["finished"]:
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")

                        # Trigger background summary update if needed (fire-and-forget)
                        if chat_id:
                            try:
                                memory = get_chat_memory()
                                asyncio.create_task(memory.maybe_update_summary(chat_id))
                            except Exception as e:
                                logger.warning(f"Summary trigger failed: {e}")

                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])
                        full_response_tokens.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Chat generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    # Connection already closed, just log
                    logger.error("Could not send error message - connection closed")

    except WebSocketDisconnect:
        logger.info("WebSocket disconnected")
    except Exception as e:
        logger.error(f"WebSocket error: {str(e)}")

class LessonPlanRequest(BaseModel):
    prompt: str


LESSON_PLAN_HISTORY_FILE = os.path.join(os.path.dirname(__file__), "lesson_plan_history.json")
LESSON_DRAFTS_FILE = os.path.join(os.path.dirname(__file__), "lesson_drafts.json")

class LessonPlanHistory(BaseModel):
    model_config = {"extra": "allow"}

    id: str
    title: str
    timestamp: str
    formData: dict
    generatedPlan: Optional[str] = None
    parsedLesson: Optional[dict] = None
    curriculumMatches: Optional[list] = None

@app.get("/api/lesson-plan-history")
async def get_lesson_plan_history():
    """Get all lesson plan histories"""
    try:
        if os.path.exists(LESSON_PLAN_HISTORY_FILE):
            with open(LESSON_PLAN_HISTORY_FILE, 'r', encoding='utf-8') as f:
                histories = json.load(f)
            histories.sort(key=lambda x: x['timestamp'], reverse=True)
            return histories
        return []
    except (json.JSONDecodeError, ValueError):
        return []
    except Exception as e:
        logger.error(f"Error loading lesson plan history: {e}")
        return []

@app.post("/api/lesson-plan-history")
async def save_lesson_plan_history(plan: LessonPlanHistory):
    """Save or update a lesson plan history"""
    try:
        histories = []
        if os.path.exists(LESSON_PLAN_HISTORY_FILE):
            try:
                with open(LESSON_PLAN_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    histories = json.load(f)
            except (json.JSONDecodeError, ValueError):
                histories = []
        
        existing_index = None
        for i, h in enumerate(histories):
            if h['id'] == plan.id:
                existing_index = i
                break
        
        plan_dict = plan.model_dump()
        
        if existing_index is not None:
            histories[existing_index] = plan_dict
        else:
            histories.append(plan_dict)
        
        tmp_file = LESSON_PLAN_HISTORY_FILE + ".tmp"
        with open(tmp_file, 'w', encoding='utf-8') as f:
            json.dump(histories, f, indent=2, ensure_ascii=False)
        os.replace(tmp_file, LESSON_PLAN_HISTORY_FILE)

        return {"success": True, "id": plan.id}
    except Exception as e:
        logger.error(f"Error saving lesson plan history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/lesson-plan-history/{plan_id}")
async def delete_lesson_plan_history(plan_id: str):
    """Delete a lesson plan history"""
    try:
        if os.path.exists(LESSON_PLAN_HISTORY_FILE):
            try:
                with open(LESSON_PLAN_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    histories = json.load(f)
            except (json.JSONDecodeError, ValueError):
                histories = []

            histories = [h for h in histories if h['id'] != plan_id]

            tmp_file = LESSON_PLAN_HISTORY_FILE + ".tmp"
            with open(tmp_file, 'w', encoding='utf-8') as f:
                json.dump(histories, f, indent=2, ensure_ascii=False)
            os.replace(tmp_file, LESSON_PLAN_HISTORY_FILE)
            
            return {"success": True}
        
        raise HTTPException(status_code=404, detail="Lesson plan history not found")
    except Exception as e:
        logger.error(f"Error deleting lesson plan history: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Lesson Plan Drafts ──────────────────────────────────────────────────

class LessonDraft(BaseModel):
    model_config = {"extra": "allow"}

    id: str
    title: str
    timestamp: str
    plannerType: str  # lesson, kindergarten, multigrade, cross-curricular
    formData: dict
    step: Optional[int] = 1
    curriculumMatches: Optional[list] = None

@app.get("/api/lesson-drafts")
async def get_lesson_drafts(plannerType: Optional[str] = None):
    """Get all lesson plan drafts, optionally filtered by planner type"""
    try:
        if os.path.exists(LESSON_DRAFTS_FILE):
            with open(LESSON_DRAFTS_FILE, 'r', encoding='utf-8') as f:
                drafts = json.load(f)
            if plannerType:
                drafts = [d for d in drafts if d.get('plannerType') == plannerType]
            drafts.sort(key=lambda x: x['timestamp'], reverse=True)
            return drafts
        return []
    except (json.JSONDecodeError, ValueError):
        return []
    except Exception as e:
        logger.error(f"Error loading lesson drafts: {e}")
        return []

@app.post("/api/lesson-drafts")
async def save_lesson_draft(draft: LessonDraft):
    """Save or update a lesson plan draft"""
    try:
        drafts = []
        if os.path.exists(LESSON_DRAFTS_FILE):
            try:
                with open(LESSON_DRAFTS_FILE, 'r', encoding='utf-8') as f:
                    drafts = json.load(f)
            except (json.JSONDecodeError, ValueError):
                drafts = []

        existing_index = None
        for i, d in enumerate(drafts):
            if d['id'] == draft.id:
                existing_index = i
                break

        draft_dict = draft.model_dump()

        if existing_index is not None:
            drafts[existing_index] = draft_dict
        else:
            drafts.append(draft_dict)

        tmp_file = LESSON_DRAFTS_FILE + ".tmp"
        with open(tmp_file, 'w', encoding='utf-8') as f:
            json.dump(drafts, f, indent=2, ensure_ascii=False)
        os.replace(tmp_file, LESSON_DRAFTS_FILE)

        return {"success": True, "id": draft.id}
    except Exception as e:
        logger.error(f"Error saving lesson draft: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/lesson-drafts/{draft_id}")
async def delete_lesson_draft(draft_id: str):
    """Delete a lesson plan draft"""
    try:
        if os.path.exists(LESSON_DRAFTS_FILE):
            try:
                with open(LESSON_DRAFTS_FILE, 'r', encoding='utf-8') as f:
                    drafts = json.load(f)
            except (json.JSONDecodeError, ValueError):
                drafts = []

            drafts = [d for d in drafts if d['id'] != draft_id]

            tmp_file = LESSON_DRAFTS_FILE + ".tmp"
            with open(tmp_file, 'w', encoding='utf-8') as f:
                json.dump(drafts, f, indent=2, ensure_ascii=False)
            os.replace(tmp_file, LESSON_DRAFTS_FILE)

            return {"success": True}

        raise HTTPException(status_code=404, detail="Draft not found")
    except Exception as e:
        logger.error(f"Error deleting lesson draft: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate-lesson-plan")
async def generate_lesson_plan(request: LessonPlanRequest):
    """Generate a lesson plan using the LLM (via process pool)"""
    try:
        prompt_text = build_prompt("You are an expert educational consultant and curriculum designer. Create detailed, engaging, and pedagogically sound lesson plans.", request.prompt)

        settings = {
            "model_path": get_model_path(),
            "n_ctx": MODEL_N_CTX,
            "max_tokens": 6000,
            "temperature": 0.7,
            "tool_name": "lesson_plan",
            "prompt_template": prompt_text,
        }
        future = submit_task(run_llama_inference, request.prompt, settings)
        result = await asyncio.wrap_future(future)

        if result["metadata"]["status"] == "error":
            raise HTTPException(status_code=500, detail=result["metadata"].get("error_message", "Generation failed"))

        return {"lessonPlan": result["result"]}

    except Exception as e:
        logger.error(f"Error generating lesson plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.websocket("/ws/lesson-plan")
async def websocket_lesson_plan(websocket: WebSocket):
    await websocket.accept()
    global curriculum_matcher
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            form_data = data.get("formData", {})
            job_id = data.get("jobId") or data.get("id") or "lesson-plan"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                continue

            # Use CurriculumMatcher to find relevant curriculum pages
            curriculum_refs = []
            curriculum_context = ""
            if curriculum_matcher:
                # Extract grade and subject for exact filtering
                grade_filter = ""
                subject_filter = ""
                query = ""
                if isinstance(form_data, dict) and form_data:
                    grade_filter = str(form_data.get("gradeLevel", "")).strip()
                    subject_filter = str(form_data.get("subject", "")).strip()
                    # Build query from topic-relevant fields only (not duration, class size, etc.)
                    topic_parts = [
                        str(form_data.get("topic", "")),
                        str(form_data.get("strand", "")),
                        str(form_data.get("subject", "")),
                    ]
                    query = " ".join(p for p in topic_parts if p and p != "None")
                if not query:
                    query = prompt
                matches = curriculum_matcher.find_matching_pages(
                    query, top_k=3,
                    grade=grade_filter, subject=subject_filter
                )
                curriculum_refs = [
                    {
                        "id": m.get("id"),
                        "displayName": m.get("displayName"),
                        "grade": m.get("grade"),
                        "subject": m.get("subject"),
                        "route": m.get("route"),
                        "matchScore": m.get("matchScore"),
                    }
                    for m in matches
                ]
                context_blocks = []
                for m in matches:
                    ctx = curriculum_matcher.get_curriculum_context(m.get("id"))
                    if ctx:
                        context_blocks.append(ctx)
                if context_blocks:
                    curriculum_context = "\n\n".join(context_blocks)

            # Tier-1 awareness: use template-based prompt for small models
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("lesson-plan") if _is_tier1 else {}

            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if _is_tier1:
                system_prompt = get_tier1_system_prompt("lesson-plan", _grade)
            else:
                system_prompt = "You are an expert educational consultant and curriculum designer. Create detailed, engaging, and pedagogically sound lesson plans that teachers can immediately implement. Focus on practical activities, clear assessment strategies, and alignment with curriculum standards."

            # Prefer user-selected ELO/SCOs from the form over generic curriculum context
            user_elo = ""
            user_scos = ""
            if isinstance(form_data, dict) and form_data:
                user_elo = str(form_data.get("essentialOutcomes", "")).strip()
                user_scos = str(form_data.get("specificOutcomes", "")).strip()

            if user_elo or user_scos:
                selected_context = "\n\nTeacher-Selected Curriculum Outcomes (align lesson to these):"
                if user_elo:
                    selected_context += f"\n\nEssential Learning Outcome:\n{user_elo}"
                if user_scos:
                    sco_list = [s.strip() for s in user_scos.split("\n") if s.strip()]
                    selected_context += "\n\nSpecific Curriculum Outcomes:"
                    for i, sco in enumerate(sco_list, 1):
                        selected_context += f"\n  {i}. {sco}"
                system_prompt += selected_context
            elif curriculum_context:
                system_prompt += "\n\nCurriculum Alignment Context:\n" + curriculum_context

            full_prompt = build_prompt(system_prompt, prompt)

            # Send curriculum references to frontend before generation
            if curriculum_refs:
                try:
                    await websocket.send_json({
                        "type": "curriculum_refs",
                        "references": curriculum_refs
                    })
                except Exception as e:
                    logger.error(f"Error sending curriculum references: {e}")

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                
                # ✅ FIX: Use inference factory instead of process pool
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("lesson-plan") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                # Stream tokens in real-time (works with both Gemma API and local models)
                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="lesson_plan",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Lesson plan generation error: {e}")
                import traceback
                logger.error(f"Full traceback: {traceback.format_exc()}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Lesson Plan WebSocket disconnected")
    except Exception as e:
        logger.error(f"Lesson Plan WebSocket error: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
 
 

@app.websocket("/ws/quiz")
async def quiz_websocket(websocket: WebSocket):
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            form_data = data.get("formData", {})
            job_id = data.get("jobId") or data.get("id") or "quiz"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                logger.error("Empty quiz prompt received")
                continue

            # Tier-1 awareness: use template-based prompt for small models
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("quiz") if _is_tier1 else {}

            # Use frontend-provided system prompt if available (contains all
            # instructional context: grade specs, curriculum, format templates).
            # Fall back to a basic system prompt for backward compatibility.
            system_prompt = data.get("systemPrompt", "").strip()
            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if not system_prompt:
                if _is_tier1:
                    system_prompt = get_tier1_system_prompt("quiz", _grade)
                else:
                    system_prompt = "You are an expert educational assessment designer. Create comprehensive, well-structured quizzes that accurately assess student learning."

                # Legacy path: add curriculum context for alignment
                if curriculum_matcher and isinstance(form_data, dict) and form_data:
                    grade_filter = str(form_data.get("gradeLevel", "")).strip()
                    subject_filter = str(form_data.get("subject", "")).strip()
                    topic_parts = [str(form_data.get("topic", "")), str(form_data.get("strand", "")), str(form_data.get("subject", ""))]
                    query = " ".join(p for p in topic_parts if p and p != "None")
                    if query:
                        matches = curriculum_matcher.find_matching_pages(query, top_k=2, grade=grade_filter, subject=subject_filter)
                        context_blocks = [curriculum_matcher.get_curriculum_context(m.get("id")) for m in matches]
                        context_blocks = [c for c in context_blocks if c]
                        if context_blocks:
                            system_prompt += "\n\nCurriculum Alignment Context:\n" + "\n\n".join(context_blocks)

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("quiz") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                # Stream tokens as they are generated
                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="quiz",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Quiz generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                # Always release the slot
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Quiz WebSocket disconnected")
    except Exception as e:
        logger.error(f"Quiz WebSocket error: {str(e)}")
        
        

@app.websocket("/ws/rubric")
async def rubric_websocket(websocket: WebSocket):
    await websocket.accept()
    logger.info("Rubric WebSocket connection accepted")
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            logger.info("Waiting for rubric request...")
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            logger.info(f"Received rubric request, prompt length: {len(data.get('prompt', ''))}")

            prompt = data.get("prompt", "")
            form_data = data.get("formData", {})
            job_id = data.get("jobId") or data.get("id") or "rubric"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                logger.error("Empty rubric prompt received")
                continue

            logger.info("Building rubric prompt...")
            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("rubric") if _is_tier1 else {}

            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if _is_tier1:
                system_prompt = get_tier1_system_prompt("rubric", _grade)
            else:
                system_prompt = "You are an expert educational assessment designer. Create detailed, fair, and comprehensive grading rubrics that clearly define performance criteria at each level."

            # Add curriculum context for alignment
            if curriculum_matcher and isinstance(form_data, dict) and form_data:
                grade_filter = str(form_data.get("gradeLevel", "")).strip()
                subject_filter = str(form_data.get("subject", "")).strip()
                topic_parts = [str(form_data.get("topic", "")), str(form_data.get("strand", "")), str(form_data.get("subject", ""))]
                query = " ".join(p for p in topic_parts if p and p != "None")
                if query:
                    matches = curriculum_matcher.find_matching_pages(query, top_k=2, grade=grade_filter, subject=subject_filter)
                    context_blocks = [curriculum_matcher.get_curriculum_context(m.get("id")) for m in matches]
                    context_blocks = [c for c in context_blocks if c]
                    if context_blocks:
                        system_prompt += "\n\nCurriculum Alignment Context:\n" + "\n\n".join(context_blocks)

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                logger.info("Getting LlamaInference instance...")
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("rubric") if generation_mode == "queued" else get_inference_instance(use_singleton=False)
                logger.info("Starting rubric generation...")

                # Use streaming method for real-time generation
                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="rubric",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        logger.info("Rubric generation complete")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Rubric generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Rubric WebSocket disconnected")
    except Exception as e:
        logger.error(f"Rubric WebSocket error: {str(e)}")



@app.websocket("/ws/kindergarten")
async def kindergarten_websocket(websocket: WebSocket):
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            job_id = data.get("jobId") or data.get("id") or "kindergarten"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                continue

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("kindergarten") if _is_tier1 else {}

            if _is_tier1:
                system_prompt = get_tier1_system_prompt("kindergarten", "K")
            else:
                system_prompt = "You are an expert early childhood educator specializing in kindergarten education. Create developmentally appropriate, engaging, and playful lesson plans that foster learning through exploration and hands-on activities."

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("kindergarten") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                # Use streaming method for real-time generation
                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="kindergarten",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Kindergarten generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Kindergarten WebSocket disconnected")
    except Exception as e:
        logger.error(f"Kindergarten WebSocket error: {str(e)}")

@app.websocket("/ws/multigrade")
async def multigrade_websocket(websocket: WebSocket):
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            job_id = data.get("jobId") or data.get("id") or "multigrade"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                continue

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("multigrade") if _is_tier1 else {}

            if _is_tier1:
                system_prompt = get_tier1_system_prompt("multigrade")
            else:
                system_prompt = "You are an expert educator specializing in multigrade and multi-age classroom instruction. Create comprehensive lesson plans that address multiple grade levels simultaneously with differentiated activities and flexible grouping strategies."

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("multigrade") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                # Use streaming method for real-time generation
                token_buffer = []
                last_send = time.time()
                chunk_count = 0
                async for chunk in inference.generate_stream(
                    tool_name="multigrade",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    chunk_count += 1
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        logger.info(f"Generation finished for job {job_id}, total chunks: {chunk_count}")
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Multigrade generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Multigrade WebSocket disconnected")
    except Exception as e:
        logger.error(f"Multigrade WebSocket error: {str(e)}")


@app.websocket("/ws/cross-curricular")
async def cross_curricular_websocket(websocket: WebSocket):
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()
            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            job_id = data.get("jobId") or data.get("id") or "cross-curricular"
            generation_mode = data.get("generationMode", "queued")

            if not prompt:
                continue

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("cross-curricular") if _is_tier1 else {}

            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if _is_tier1:
                system_prompt = get_tier1_system_prompt("cross-curricular", _grade)
            else:
                system_prompt = "You are an expert educational consultant specializing in integrated and cross-curricular lesson planning. Create comprehensive lesson plans that meaningfully connect multiple subject areas and demonstrate authentic interdisciplinary learning."

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("cross-curricular") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                # Use streaming method for real-time generation
                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="cross_curricular",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Cross-curricular generation error: {e}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Cross-curricular WebSocket disconnected")
    except Exception as e:
        logger.error(f"Cross-curricular WebSocket error: {str(e)}")


@app.websocket("/ws/worksheet")
async def worksheet_websocket(websocket: WebSocket):
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            logger.info("Worksheet WebSocket waiting for message...")
            data = await websocket.receive_json()
            logger.info(f"Worksheet WebSocket received data: {data.keys() if isinstance(data, dict) else 'non-dict'}")

            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            form_data = data.get("formData", {})
            job_id = data.get("jobId") or data.get("id") or "worksheet"
            generation_mode = data.get("generationMode", "queued")

            logger.info(f"Worksheet generation request - job_id: {job_id}, prompt length: {len(prompt)}")

            if not prompt:
                logger.warning("Empty prompt received, skipping")
                continue

            # Build worksheet-specific prompt
            subject = form_data.get("subject", "")
            grade_level = form_data.get("gradeLevel", "")
            strand = form_data.get("strand", "")
            topic = form_data.get("topic", "")
            question_type = form_data.get("questionType", "")
            question_count = form_data.get("questionCount", "")
            worksheet_title = form_data.get("worksheetTitle", "")

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("worksheet") if _is_tier1 else {}

            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if _is_tier1:
                system_prompt = get_tier1_system_prompt("worksheet", _grade)
            else:
                system_prompt = f"You are an expert educational worksheet designer. Create comprehensive, well-structured worksheets that accurately assess student learning and align with curriculum standards. Focus on clear instructions, appropriate difficulty level, and educational value."

            # Add curriculum context for alignment
            if curriculum_matcher and (subject or grade_level):
                query = " ".join(p for p in [topic, strand, subject] if p)
                if query:
                    matches = curriculum_matcher.find_matching_pages(query, top_k=2, grade=grade_level, subject=subject)
                    context_blocks = [curriculum_matcher.get_curriculum_context(m.get("id")) for m in matches]
                    context_blocks = [c for c in context_blocks if c]
                    if context_blocks:
                        system_prompt += "\n\nCurriculum Alignment Context:\n" + "\n\n".join(context_blocks)

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                logger.info(f"Acquiring generation slot for job {job_id}")
                # Acquire generation slot (queue or parallel)
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                logger.info(f"Generation slot acquired: {slot_mode}")

                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("worksheet") if generation_mode == "queued" else get_inference_instance(use_singleton=False)
                logger.info("Got inference instance, starting generation...")

                # Use streaming method for real-time generation
                token_buffer = []
                last_send = time.time()
                chunk_count = 0
                async for chunk in inference.generate_stream(
                    tool_name="worksheet",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 6000) if _is_tier1 else 6000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    chunk_count += 1
                    if chunk_count <= 5:  # Log first few chunks
                        logger.info(f"Received chunk {chunk_count}: {chunk}")
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({
                                "type": "error",
                                "message": chunk["error"]
                            })
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        # Send any remaining tokens
                        if token_buffer:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)  # Force flush
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    # Batch tokens before sending
                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])

                        # Send every 100ms or 10 tokens (whichever comes first)
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({
                                    "type": "token",
                                    "content": "".join(token_buffer)
                                })
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Worksheet generation error for job {job_id}: {e}")
                import traceback
                logger.error(f"Full traceback: {traceback.format_exc()}")
                try:
                    await websocket.send_json({
                        "type": "error",
                        "message": str(e)
                    })
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Worksheet WebSocket disconnected")
    except Exception as e:
        logger.error(f"Worksheet WebSocket error: {str(e)}")


@app.websocket("/ws/presentation")
async def presentation_websocket(websocket: WebSocket):
    """Generate presentation slide JSON from lesson content."""
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    try:
        while True:
            data = await websocket.receive_json()

            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            job_id = data.get("jobId") or data.get("id") or "presentation"
            generation_mode = data.get("generationMode", "queued")
            suggested_mode = data.get("suggestedMode", False)

            if not prompt:
                continue

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_prompt_key = "presentation_with_suggestions" if suggested_mode else "presentation"
            _t1_params = get_tier1_gen_params(_t1_prompt_key) if _is_tier1 else {}

            _grade = form_data.get("gradeLevel", "") if isinstance(form_data, dict) else ""
            if _is_tier1:
                system_prompt = get_tier1_system_prompt(_t1_prompt_key, _grade)
            else:
                system_prompt = (
                    "You are an expert presentation designer for educational content. "
                    "Convert lesson plans into concise, visually-oriented slide decks. "
                    "Return ONLY valid JSON with no markdown fences or explanation. "
                    "Each slide should have punchy headlines (max 7 words) and short bullet points (max 12 words each)."
                )

            full_prompt = build_prompt(system_prompt, prompt)

            slot_mode = None
            try:
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)

                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("presentation") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                token_buffer = []
                last_send = time.time()
                async for chunk in inference.generate_stream(
                    tool_name="presentation",
                    input_data=prompt,
                    prompt_template=full_prompt,
                    max_tokens=_t1_params.get("max_tokens", 4000) if _is_tier1 else 4000,
                    temperature=_t1_params.get("temperature", 0.7) if _is_tier1 else 0.7,
                    repeat_penalty=_t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1,
                ):
                    if job_id in cancelled_job_ids:
                        await websocket.send_json({"type": "cancelled", "jobId": job_id})
                        break

                    if chunk.get("error"):
                        try:
                            await websocket.send_json({"type": "error", "message": chunk["error"]})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send error message - connection closed")
                        break

                    if chunk.get("finished"):
                        if token_buffer:
                            try:
                                await websocket.send_json({"type": "token", "content": "".join(token_buffer)})
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending final tokens: {e}")
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send done message - connection closed")
                        break

                    if chunk.get("token"):
                        token_buffer.append(chunk["token"])
                        if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                            try:
                                await websocket.send_json({"type": "token", "content": "".join(token_buffer)})
                                token_buffer = []
                                last_send = time.time()
                                await asyncio.sleep(0)
                            except Exception as e:
                                logger.error(f"Error sending token: {e}")
                                break

            except Exception as e:
                logger.error(f"Presentation generation error: {e}")
                import traceback
                logger.error(f"Full traceback: {traceback.format_exc()}")
                try:
                    await websocket.send_json({"type": "error", "message": str(e)})
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Presentation WebSocket disconnected")
    except Exception as e:
        logger.error(f"Presentation WebSocket error: {str(e)}")


@app.websocket("/ws/storybook")
async def storybook_websocket(websocket: WebSocket):
    """Generate children's storybook JSON from teacher description."""
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    cancelled_job_ids = set()

    async def _run_generation(inference, prompt_text, full_prompt_tpl, max_tok, temp, job_id, *, stream_tokens=True, repeat_penalty=1.1, token_type="token"):
        """Run a generation pass. If stream_tokens is False, collect internally and return the text.
        token_type controls the WebSocket message type (e.g. 'narrative_token' for Pass 1 preview)."""
        collected = []
        token_buffer = []
        last_send = time.time()
        async for chunk in inference.generate_stream(
            tool_name="storybook",
            input_data=prompt_text,
            prompt_template=full_prompt_tpl,
            max_tokens=max_tok,
            temperature=temp,
            repeat_penalty=repeat_penalty,
        ):
            if job_id in cancelled_job_ids:
                await websocket.send_json({"type": "cancelled", "jobId": job_id})
                return None

            if chunk.get("error"):
                try:
                    await websocket.send_json({"type": "error", "message": chunk["error"]})
                    await asyncio.sleep(0)
                except:
                    logger.error("Could not send error message - connection closed")
                return None

            if chunk.get("finished"):
                if stream_tokens and token_buffer:
                    try:
                        await websocket.send_json({"type": token_type, "content": "".join(token_buffer)})
                        await asyncio.sleep(0)
                    except Exception as e:
                        logger.error(f"Error sending final tokens: {e}")
                return "".join(collected)

            if chunk.get("token"):
                collected.append(chunk["token"])
                if stream_tokens:
                    token_buffer.append(chunk["token"])
                    if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                        try:
                            await websocket.send_json({"type": token_type, "content": "".join(token_buffer)})
                            token_buffer = []
                            last_send = time.time()
                            await asyncio.sleep(0)
                        except Exception as e:
                            logger.error(f"Error sending token: {e}")
                            return None
        return "".join(collected)

    try:
        while True:
            data = await websocket.receive_json()

            # Handle cancellation message
            if isinstance(data, dict) and data.get("type") == "cancel" and "jobId" in data:
                cancelled_job_ids.add(data["jobId"])
                await websocket.send_json({"type": "cancelled", "jobId": data["jobId"]})
                continue

            prompt = data.get("prompt", "")
            job_id = data.get("jobId") or data.get("id") or "storybook"
            generation_mode = data.get("generationMode", "queued")
            grade = data.get("grade", "K")
            two_pass = data.get("twoPass", False)
            narrative_prompt = data.get("narrativePrompt", "")
            structure_prompt_template = data.get("structurePromptTemplate", "")

            if not prompt and not (two_pass and narrative_prompt):
                continue

            # Tier-1 awareness
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("storybook") if _is_tier1 else {}

            max_tokens = _t1_params.get("max_tokens", 4000) if _is_tier1 else 4000
            temperature = _t1_params.get("temperature", 0.7) if _is_tier1 else 0.7
            repeat_penalty = _t1_params.get("repeat_penalty", 1.1) if _is_tier1 else 1.1

            slot_mode = None
            try:
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)

                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("storybook") if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                if two_pass and narrative_prompt and structure_prompt_template:
                    # ── Two-pass generation ──────────────────────────────────
                    # Pass 1: Generate the story narrative (collect internally)
                    narrative_system = (
                        "You are a children's storybook author. Write engaging, age-appropriate stories "
                        "with clear characters and vivid descriptions. Write the story as plain text."
                    )
                    narrative_full = build_prompt(narrative_system, narrative_prompt)

                    await websocket.send_json({"type": "status", "status": "writing_story", "jobId": job_id})
                    narrative_text = await _run_generation(
                        inference, narrative_prompt, narrative_full,
                        max_tokens, temperature, job_id, stream_tokens=True,
                        repeat_penalty=repeat_penalty, token_type="narrative_token",
                    )
                    if narrative_text is None:
                        continue  # cancelled or errored

                    logger.info(f"[Storybook] Pass 1 complete — {len(narrative_text)} chars of narrative")

                    # Pass 2: Convert narrative to structured JSON (stream to client)
                    structure_prompt = structure_prompt_template.replace("{{NARRATIVE}}", narrative_text)
                    structure_system = (
                        "You are a JSON formatting assistant. Convert the provided story into the exact "
                        "JSON structure requested. Return ONLY valid JSON with no markdown fences or explanation."
                    )
                    structure_full = build_prompt(structure_system, structure_prompt)

                    await websocket.send_json({"type": "status", "status": "formatting_pages", "jobId": job_id})
                    result = await _run_generation(
                        inference, structure_prompt, structure_full,
                        max_tokens, temperature, job_id, stream_tokens=True,
                        repeat_penalty=repeat_penalty,
                    )
                    if result is not None:
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send done message - connection closed")
                else:
                    # ── Single-pass generation (original path) ───────────────
                    if _is_tier1:
                        system_prompt = get_tier1_system_prompt("storybook", grade)
                    else:
                        system_prompt = (
                            "You are a children's storybook author specializing in early childhood education (K-2). "
                            "Create engaging, age-appropriate stories with clear characters, simple vocabulary, "
                            "and vivid scene descriptions. Tag every text segment with its speaker. "
                            "Return ONLY valid JSON with no markdown fences or explanation."
                        )

                    full_prompt = build_prompt(system_prompt, prompt)

                    result = await _run_generation(
                        inference, prompt, full_prompt,
                        max_tokens, temperature, job_id, stream_tokens=True,
                        repeat_penalty=repeat_penalty,
                    )
                    if result is not None:
                        try:
                            await websocket.send_json({"type": "done"})
                            await asyncio.sleep(0)
                        except:
                            logger.error("Could not send done message - connection closed")

            except Exception as e:
                logger.error(f"Storybook generation error: {e}")
                import traceback
                logger.error(f"Full traceback: {traceback.format_exc()}")
                try:
                    await websocket.send_json({"type": "error", "message": str(e)})
                except:
                    logger.error("Could not send error message - connection closed")
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except Exception as e:
                    logger.error(f"Error releasing generation slot: {e}")

    except WebSocketDisconnect:
        logger.info("Storybook WebSocket disconnected")
    except Exception as e:
        logger.error(f"Storybook WebSocket error: {str(e)}")


# ─── Brain Dump: category keyword map & helpers ──────────────────────────────
BRAIN_DUMP_ACTION_DESCRIPTIONS = {
    "lesson-plan":           '- "lesson-plan": Create a lesson plan (details: subject, grade, topic, duration)',
    "quiz":                  '- "quiz": Create a quiz/assessment (details: subject, grade, topic, questionCount, difficulty)',
    "rubric":                '- "rubric": Create a grading rubric (details: subject, grade, topic, criteria)',
    "worksheet":             '- "worksheet": Create a worksheet (details: subject, grade, topic, questionType)',
    "calendar-task":         '- "calendar-task": Add a task/reminder to calendar (details: date, priority)',
    "kindergarten-plan":     '- "kindergarten-plan": Create an early childhood lesson plan (details: topic, theme, duration)',
    "multigrade-plan":       '- "multigrade-plan": Create a multi-level lesson plan (details: subject, grades, topic)',
    "cross-curricular-plan": '- "cross-curricular-plan": Create an integrated cross-subject lesson (details: subjects, grade, topic)',
    "image-studio":          '- "image-studio": Create a visual aid or image (details: description, style)',
    "storybook":             '- "storybook": Create an illustrated storybook for K-2 students (details: title, grade, description, pageCount)',
    "presentation":          '- "presentation": Create a slide deck or presentation (details: subject, grade, topic, slideCount)',
    "grade-quiz":            '- "grade-quiz": Grade or mark a quiz/test (details: subject, grade, topic)',
    "grade-worksheet":       '- "grade-worksheet": Grade or mark a worksheet (details: subject, grade, topic)',
    "curriculum-tracker":    '- "curriculum-tracker": Track or update curriculum progress/milestones (details: subject, grade, topic, status)',
    "curriculum-browse":     '- "curriculum-browse": Look up or browse curriculum standards (details: subject, grade)',
    "class-management":      '- "class-management": Manage students, add students, or manage a class (details: className, grade)',
    "attendance":            '- "attendance": Take or record attendance (details: className, grade, date)',
}

BRAIN_DUMP_CATEGORIES = {
    "assessment": {
        "types": ["quiz", "rubric", "grade-quiz", "grade-worksheet", "worksheet"],
        "keywords": ["test", "assess", "quiz", "grade", "mark", "evaluate", "rubric",
                     "worksheet", "score", "check", "exam", "question", "answer",
                     "multiple choice", "true or false", "fill in"],
    },
    "planning": {
        "types": ["lesson-plan", "kindergarten-plan", "multigrade-plan", "cross-curricular-plan"],
        "keywords": ["plan", "lesson", "teach", "unit", "week", "prepare", "objective",
                     "kindergarten", "early childhood", "preschool", "multigrade",
                     "multi-grade", "cross-curricular", "integrated", "activity",
                     "learning outcome", "scheme of work"],
    },
    "content": {
        "types": ["presentation", "image-studio"],
        "keywords": ["slide", "presentation", "visual", "image", "poster", "picture",
                     "powerpoint", "deck", "diagram", "chart", "illustration",
                     "graphic", "photo", "display"],
    },
    "organization": {
        "types": ["calendar-task", "curriculum-tracker", "curriculum-browse"],
        "keywords": ["schedule", "deadline", "date", "remind", "reminder", "track",
                     "progress", "curriculum", "standard", "calendar", "task",
                     "todo", "to-do", "due", "milestone", "syllabus", "pacing",
                     "timeline", "browse", "look up", "find standard"],
    },
    "management": {
        "types": ["class-management", "attendance"],
        "keywords": ["student", "attend", "attendance", "roster", "enroll", "absent",
                     "register", "class list", "pupil", "manage class", "add student",
                     "roll call", "present"],
    },
}

ALL_ACTION_TYPE_NAMES = list(BRAIN_DUMP_ACTION_DESCRIPTIONS.keys())

# Map each action type back to its category for auto-learning
_ACTION_TO_CATEGORY = {}
for _cat, _info in BRAIN_DUMP_CATEGORIES.items():
    for _at in _info["types"]:
        _ACTION_TO_CATEGORY[_at] = _cat

LEARNED_KEYWORDS_FILE = MODELS_DIR / ".brain-dump-learned-keywords.json"


def _load_learned_keywords() -> dict:
    """Load user-confirmed keyword→category mappings from disk."""
    try:
        if LEARNED_KEYWORDS_FILE.exists():
            with open(LEARNED_KEYWORDS_FILE, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _save_learned_keyword(word: str, category: str):
    """Persist a new keyword→category mapping learned from user confirmation."""
    try:
        data = _load_learned_keywords()
        if category not in data:
            data[category] = []
        if word not in data[category]:
            data[category].append(word)
            LEARNED_KEYWORDS_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(LEARNED_KEYWORDS_FILE, 'w') as f:
                json.dump(data, f, indent=2)
    except Exception as e:
        logger.warning(f"Failed to save learned keyword: {e}")


def _get_matching_categories(text: str) -> list[str]:
    """Return category names whose keywords appear in the text.
    Checks both hardcoded and user-learned keywords.
    If fewer than 2 categories match, return ALL categories (ambiguous input)."""
    text_lower = text.lower()
    matched = set()

    # Check hardcoded keywords
    for cat_name, cat_info in BRAIN_DUMP_CATEGORIES.items():
        if any(kw in text_lower for kw in cat_info["keywords"]):
            matched.add(cat_name)

    # Check learned keywords
    learned = _load_learned_keywords()
    for cat_name, words in learned.items():
        if cat_name in BRAIN_DUMP_CATEGORIES and any(w in text_lower for w in words):
            matched.add(cat_name)

    if len(matched) < 2:
        return list(BRAIN_DUMP_CATEGORIES.keys())
    return list(matched)


def _build_brain_dump_prompt(matched_categories: list[str]) -> str:
    """Assemble the brain dump system prompt with only the relevant action types."""
    action_lines = []
    for cat_name in matched_categories:
        for action_type in BRAIN_DUMP_CATEGORIES[cat_name]["types"]:
            if action_type in BRAIN_DUMP_ACTION_DESCRIPTIONS:
                action_lines.append(BRAIN_DUMP_ACTION_DESCRIPTIONS[action_type])

    return f"""You are an AI assistant for the OECS Learning Hub, a teacher productivity app. Your job is to analyze a teacher's free-form thoughts ("brain dump") and extract actionable items that map to features in the app.

Available action types:
{chr(10).join(action_lines)}

Return ONLY a valid JSON object with two keys:
- "actions": a JSON array where each item has:
  - "type": one of the action types above
  - "title": short descriptive title (max 60 chars)
  - "description": brief explanation of what to create (1-2 sentences)
  - "details": object with relevant fields for that action type
- "unmatched": an array of text snippets from the user's input that you could NOT confidently map to any of the available action types. If everything was matched, use an empty array.

If the text mentions dates, include them in details.date. If it mentions a subject/grade, include those.

IMPORTANT: Always try your best to match text to actions, even if you are not fully confident. Make your best guess — the teacher will confirm or reject it.
Only put text in "unmatched" if it truly does not relate to any available action type (e.g., greetings, off-topic remarks).
Some sentences may be context or elaboration for other sentences — include that context in the relevant action's description or details, do NOT put it in unmatched.
NEVER return both empty actions AND empty unmatched. If you genuinely cannot match anything, put ALL the text in "unmatched" so the teacher can clarify.
Do NOT include any text, markdown, or explanation — ONLY the JSON object."""


async def _stream_to_ws(websocket, inference, prompt, text, max_tokens, temperature,
                        token_type="token", done_type="done"):
    """Shared streaming helper for brain dump websocket responses."""
    token_buffer = []
    last_send = time.time()
    async for chunk in inference.generate_stream(
        tool_name="brain-dump",
        input_data=text,
        prompt_template=prompt,
        max_tokens=max_tokens,
        temperature=temperature,
    ):
        if chunk.get("error"):
            try:
                await websocket.send_json({"type": "error", "message": chunk["error"]})
                await asyncio.sleep(0)
            except:
                pass
            break

        if chunk.get("finished"):
            if token_buffer:
                try:
                    await websocket.send_json({"type": token_type, "content": "".join(token_buffer)})
                    await asyncio.sleep(0)
                except:
                    pass
            try:
                await websocket.send_json({"type": done_type})
                await asyncio.sleep(0)
            except:
                pass
            break

        if chunk.get("token"):
            token_buffer.append(chunk["token"])
            if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                try:
                    await websocket.send_json({"type": token_type, "content": "".join(token_buffer)})
                    token_buffer = []
                    last_send = time.time()
                    await asyncio.sleep(0)
                except:
                    break


@app.websocket("/ws/brain-dump")
async def brain_dump_websocket(websocket: WebSocket):
    """Analyze free-form teacher brain dump and map to app actions."""
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    try:
        while True:
            data = await websocket.receive_json()
            msg_type = data.get("type", "analyze")

            # ── Suggestion request: lightweight second pass ──────────────
            if msg_type == "suggest":
                unmatched_text = data.get("unmatchedText", "").strip()
                job_id = data.get("jobId") or "brain-dump-suggest"
                generation_mode = data.get("generationMode", "queued")
                if not unmatched_text:
                    await websocket.send_json({"type": "suggestions_done"})
                    continue

                type_names = ", ".join(ALL_ACTION_TYPE_NAMES)
                suggest_prompt_text = f"""The user wrote: "{unmatched_text}"

Available tools in the app: {type_names}

Return a JSON array of objects. Each object should have:
- "text": the relevant snippet from the user's input
- "suggestedTypes": array of 1-3 tool names from the list above that MIGHT apply
- "confidence": "low" or "medium"
If nothing applies, return an empty array: []
Do NOT include any text, markdown, or explanation — ONLY the JSON array."""

                suggest_full = "<|begin_of_text|>"
                suggest_full += f"<|start_header_id|>system<|end_header_id|>\n\nYou help match teacher notes to app features. Be generous — suggest plausible matches even if uncertain.<|eot_id|>"
                suggest_full += f"<|start_header_id|>user<|end_header_id|>\n\n{suggest_prompt_text}<|eot_id|>"
                suggest_full += "<|start_header_id|>assistant<|end_header_id|>\n\n"

                slot_mode = None
                try:
                    slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                    from inference_factory import get_inference_instance, resolve_inference_for_task
                    inference = resolve_inference_for_task("brain-dump") if generation_mode == "queued" else get_inference_instance(use_singleton=False)
                    await _stream_to_ws(websocket, inference, suggest_full, unmatched_text,
                                        max_tokens=500, temperature=0.3,
                                        token_type="suggestion_token", done_type="suggestions_done")
                except Exception as e:
                    logger.error(f"Brain dump suggestion error: {e}")
                    try:
                        await websocket.send_json({"type": "error", "message": str(e)})
                    except:
                        pass
                finally:
                    try:
                        release_generation_slot(slot_mode or generation_mode)
                    except:
                        pass
                continue

            # ── Generate single action from user-confirmed suggestion ────
            if msg_type == "generate-action":
                action_type = data.get("actionType", "").strip()
                action_text = data.get("text", "").strip()
                job_id = data.get("jobId") or "brain-dump-action"
                generation_mode = data.get("generationMode", "queued")
                if not action_type or not action_text:
                    await websocket.send_json({"type": "action_done"})
                    continue

                # Auto-learn: save keywords from this text → action's category
                category = _ACTION_TO_CATEGORY.get(action_type)
                if category:
                    words = [w.lower().strip() for w in action_text.split() if len(w.strip()) > 3]
                    # Filter out very common words
                    stopwords = {"the", "and", "for", "that", "this", "with", "from", "have", "will",
                                 "need", "want", "some", "about", "been", "also", "they", "their",
                                 "what", "when", "where", "which", "there", "would", "could", "should",
                                 "make", "like", "just", "them", "than", "into", "over", "very", "much"}
                    for word in words:
                        if word not in stopwords:
                            _save_learned_keyword(word, category)

                type_desc = BRAIN_DUMP_ACTION_DESCRIPTIONS.get(action_type, f'- "{action_type}": (unknown type)')
                gen_prompt_text = f"""Based on the teacher's note: "{action_text}"

Create a single action of this type:
{type_desc}

Return ONLY a valid JSON object with:
- "type": "{action_type}"
- "title": short descriptive title (max 60 chars)
- "description": brief explanation of what to create (1-2 sentences)
- "details": object with relevant fields for that action type
Do NOT include any text, markdown, or explanation — ONLY the JSON object."""

                gen_full = "<|begin_of_text|>"
                gen_full += f"<|start_header_id|>system<|end_header_id|>\n\nYou create structured actions for the OECS Learning Hub teacher app.<|eot_id|>"
                gen_full += f"<|start_header_id|>user<|end_header_id|>\n\n{gen_prompt_text}<|eot_id|>"
                gen_full += "<|start_header_id|>assistant<|end_header_id|>\n\n"

                slot_mode = None
                try:
                    slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                    from inference_factory import get_inference_instance, resolve_inference_for_task
                    inference = resolve_inference_for_task("brain-dump") if generation_mode == "queued" else get_inference_instance(use_singleton=False)
                    await _stream_to_ws(websocket, inference, gen_full, action_text,
                                        max_tokens=500, temperature=0.3,
                                        token_type="action_token", done_type="action_done")
                except Exception as e:
                    logger.error(f"Brain dump generate-action error: {e}")
                    try:
                        await websocket.send_json({"type": "error", "message": str(e)})
                    except:
                        pass
                finally:
                    try:
                        release_generation_slot(slot_mode or generation_mode)
                    except:
                        pass
                continue

            # ── Primary analysis (default) ───────────────────────────────
            text = data.get("text", "").strip()
            job_id = data.get("jobId") or "brain-dump"
            generation_mode = data.get("generationMode", "queued")

            if not text:
                continue

            matched_categories = _get_matching_categories(text)
            system_prompt = _build_brain_dump_prompt(matched_categories)

            # Tier-1 awareness: prepend extra JSON constraint for small models
            from tier1_prompts import get_tier1_system_prompt, get_tier1_gen_params
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1
            _t1_params = get_tier1_gen_params("brain-dump") if _is_tier1 else {}
            if _is_tier1:
                t1_prefix = get_tier1_system_prompt("brain-dump")
                # Simplify prompt for tier-1: reduce detail requirements to save tokens
                system_prompt = system_prompt.replace(
                    '  - "details": object with relevant fields for that action type',
                    '  - "details": object with only subject, grade, topic, and date if mentioned'
                )
                system_prompt = t1_prefix + "\n\n" + system_prompt

            full_prompt = build_prompt(system_prompt, text)

            slot_mode = None
            try:
                slot_mode = await acquire_generation_slot(websocket, generation_mode, job_id)
                from inference_factory import get_inference_instance, resolve_inference_for_task
                inference = resolve_inference_for_task("brain-dump") if generation_mode == "queued" else get_inference_instance(use_singleton=False)
                await _stream_to_ws(websocket, inference, full_prompt, text,
                                    max_tokens=_t1_params.get("max_tokens", 3000) if _is_tier1 else 3000,
                                    temperature=_t1_params.get("temperature", 0.3) if _is_tier1 else 0.3)
            except Exception as e:
                logger.error(f"Brain dump analysis error: {e}")
                try:
                    await websocket.send_json({"type": "error", "message": str(e)})
                except:
                    pass
            finally:
                try:
                    release_generation_slot(slot_mode or generation_mode)
                except:
                    pass

    except WebSocketDisconnect:
        logger.info("Brain Dump WebSocket disconnected")
    except Exception as e:
        logger.error(f"Brain Dump WebSocket error: {e}")


# ── Educator Insights — multi-pass LLM analysis ────────────────────────────

@app.websocket("/ws/educator-insights")
async def educator_insights_websocket(websocket: WebSocket):
    """Run multi-pass LLM analysis across teacher data and stream results progressively."""
    await websocket.accept()
    from generation_gate import acquire_generation_slot, release_generation_slot

    PASS_DEFINITIONS = [
        {"key": "curriculum", "name": "Curriculum Coverage", "task": "insights-curriculum"},
        {"key": "performance", "name": "Student Performance", "task": "insights-performance"},
        {"key": "content", "name": "Content Creation", "task": "insights-content"},
        {"key": "attendance", "name": "Attendance & Engagement", "task": "insights-attendance"},
        {"key": "achievements", "name": "Achievements & Engagement", "task": "insights-achievements"},
        {"key": "recommendations", "name": "Teaching Recommendations", "task": "insights-recommendations"},
        {"key": "synthesis", "name": "Final Synthesis", "task": "insights-synthesis"},
    ]

    try:
        while True:
            data = await websocket.receive_json()
            action = data.get("action", "")
            if action != "generate":
                continue

            generation_mode = data.get("generationMode", "queued")
            teacher_id = data.get("teacherId", "default_teacher")
            user_id = data.get("userId")
            registration_date = data.get("registrationDate")

            import insights_service

            # Get date context FIRST (needed for date-filtered aggregation)
            try:
                date_context = insights_service.get_report_date_context(teacher_id, user_id, registration_date)
            except Exception:
                date_context = {"is_first_report": True, "from_date": datetime.now().strftime("%Y-%m-%d"), "to_date": datetime.now().strftime("%Y-%m-%d"), "previous_report": None, "previous_report_id": None, "report_count": 0}

            # Aggregate all data with date range filtering
            try:
                all_data = insights_service.aggregate_all(teacher_id, user_id, date_context.get("from_date"), date_context.get("to_date"))
            except Exception as e:
                logger.error(f"Insights data aggregation error: {e}")
                await websocket.send_json({"type": "error", "message": f"Data aggregation failed: {e}"})
                continue

            # Debug: send aggregation summary to frontend console
            try:
                agg_summary = {}
                for _dk, _dv in all_data.items():
                    agg_summary[_dk] = {
                        "has_data": _dv.get("has_data"),
                        "llm_text_length": len(_dv.get("llm_text", "")),
                        "periodCompleted": _dv.get("periodCompleted"),
                        "periodAssessments": _dv.get("periodAssessments"),
                        "periodCount": _dv.get("periodCount"),
                        "periodRecords": _dv.get("periodRecords"),
                        "periodEarned": _dv.get("periodEarned"),
                    }
                await websocket.send_json({
                    "type": "debug",
                    "stage": "aggregation",
                    "dateContext": {
                        "is_first_report": date_context.get("is_first_report"),
                        "from_date": date_context.get("from_date"),
                        "to_date": date_context.get("to_date"),
                        "report_count": date_context.get("report_count"),
                    },
                    "summary": agg_summary,
                })
            except Exception:
                pass

            prev_report = date_context.get("previous_report")
            prev_pass_outputs = {}
            if prev_report:
                for p in prev_report.get("passes", []):
                    prev_pass_outputs[p["key"]] = p.get("output", "")

            # Determine which passes have data
            pass_outputs = {}
            from tier1_prompts import get_tier1_gen_params, TIER1_PROMPTS
            _tier_info = compute_effective_tier()
            _is_tier1 = _tier_info["tier"] == 1

            total_passes = len(PASS_DEFINITIONS)

            for pass_idx, pass_def in enumerate(PASS_DEFINITIONS):
                pass_num = pass_idx + 1
                pass_key = pass_def["key"]
                pass_name = pass_def["name"]
                task_type = pass_def["task"]

                # Send status
                try:
                    await websocket.send_json({
                        "type": "status",
                        "pass": pass_num,
                        "passName": pass_name,
                        "total": total_passes
                    })
                    await asyncio.sleep(0)
                except Exception:
                    break

                # Build the prompt data for this pass
                if pass_key in ("curriculum", "performance", "content", "attendance", "achievements"):
                    dimension_data = all_data.get(pass_key, {})
                    if not dimension_data.get("has_data"):
                        # Skip this pass — no data at all
                        no_data_msg = f"No {pass_name.lower()} data available yet."
                        pass_outputs[pass_key] = no_data_msg
                        try:
                            await websocket.send_json({
                                "type": "debug",
                                "stage": "pass_skipped",
                                "passKey": pass_key,
                                "passName": pass_name,
                                "reason": "has_data is False — no data exists for this dimension",
                            })
                            await websocket.send_json({
                                "type": "pass_complete",
                                "pass": pass_num,
                                "passName": pass_name,
                                "result": no_data_msg,
                                "skipped": True
                            })
                            await asyncio.sleep(0)
                        except Exception:
                            break
                        continue

                    # Check for "no change" — has_data is True but period activity is zero
                    period_field_map = {
                        "curriculum": "periodCompleted",
                        "performance": "periodAssessments",
                        "content": "periodCount",
                        "attendance": "periodRecords",
                        "achievements": "periodEarned",
                    }
                    period_field = period_field_map.get(pass_key)
                    period_count = dimension_data.get(period_field, None)
                    is_no_change = (
                        period_count is not None
                        and period_count == 0
                        and not date_context.get("is_first_report")
                    )

                    if is_no_change:
                        # Build a summary from all-time data instead of calling the LLM
                        from_date = date_context.get("from_date", "unknown")
                        summary_parts = {
                            "curriculum": lambda d: f"Total milestones: {d.get('totalMilestones', 0)} completed out of {d.get('totalMilestones', 0) + d.get('remainingMilestones', 0)}.",
                            "performance": lambda d: f"Total assessments: {d.get('totalAssessments', 0)}. Overall average: {d.get('overallAvg', 0):.1f}%.",
                            "content": lambda d: f"Total resources created: {d.get('totalResources', 0)}. Top type: {d.get('topType', 'N/A')}.",
                            "attendance": lambda d: f"Total records: {d.get('totalRecords', 0)}. Average rate: {d.get('avgRate', 0):.1f}%.",
                            "achievements": lambda d: f"Total earned: {d.get('totalEarned', 0)} of {d.get('totalAvailable', 0)}. Points: {d.get('totalPoints', 0)}.",
                        }
                        stats = summary_parts.get(pass_key, lambda d: "")(dimension_data)
                        no_change_msg = f"No new {pass_name.lower()} activity since the last report ({from_date}). {stats}"
                        pass_outputs[pass_key] = no_change_msg
                        try:
                            await websocket.send_json({
                                "type": "debug",
                                "stage": "no_change",
                                "passKey": pass_key,
                                "passName": pass_name,
                                "periodField": period_field,
                                "periodCount": period_count,
                                "message": no_change_msg,
                            })
                            await websocket.send_json({
                                "type": "pass_complete",
                                "pass": pass_num,
                                "passName": pass_name,
                                "result": no_change_msg,
                                "noChange": True,
                                "fromDate": from_date,
                                "stats": stats,
                            })
                            await asyncio.sleep(0)
                        except Exception:
                            break
                        continue

                    llm_input = dimension_data.get("llm_text", "")
                    system_prompt_template = TIER1_PROMPTS.get(task_type, "Analyze the data and provide bullet points.")
                    system_prompt = system_prompt_template.replace("{data}", llm_input)

                    # Inject progression context if we have a previous report
                    prev_output = prev_pass_outputs.get(pass_key, "")
                    if prev_output and not date_context.get("is_first_report"):
                        prev_date = date_context.get("from_date", "unknown")
                        system_prompt += (
                            f"\n\nPREVIOUS ANALYSIS (from {prev_date}):\n{prev_output[:300]}\n\n"
                            "Also note any changes, improvements, or regressions compared to the previous analysis period."
                        )

                elif pass_key == "recommendations":
                    # Combine outputs from passes 1-5, trimmed to save tokens
                    per_pass_limit = 300 if _is_tier1 else 600
                    combined = []
                    for key in ("curriculum", "performance", "content", "attendance", "achievements"):
                        output = pass_outputs.get(key, "")
                        if output:
                            trimmed = output[:per_pass_limit]
                            if len(output) > per_pass_limit:
                                trimmed = trimmed.rsplit("\n", 1)[0]  # cut at last full line
                            combined.append(f"[{key.upper()}]\n{trimmed}")
                    combined_text = "\n\n".join(combined) if combined else "No analysis data available."
                    system_prompt_template = TIER1_PROMPTS.get(task_type, "Provide teaching recommendations.")
                    system_prompt = system_prompt_template.replace("{data}", combined_text)

                    # Add previous recommendations for comparison
                    prev_recs = prev_pass_outputs.get("recommendations", "")
                    if prev_recs and not date_context.get("is_first_report"):
                        system_prompt += (
                            f"\n\nPREVIOUS RECOMMENDATIONS:\n{prev_recs[:300]}\n\n"
                            "Prioritize new recommendations that address recent changes. Note which prior recommendations have been addressed."
                        )

                elif pass_key == "synthesis":
                    # Combine all prior outputs, trimmed to save tokens
                    synth_limit = 400 if _is_tier1 else 800
                    system_prompt_template = TIER1_PROMPTS.get(task_type, "Write a summary report.")
                    system_prompt = system_prompt_template
                    for synth_key in ("curriculum", "performance", "content", "attendance", "achievements", "recommendations"):
                        raw = pass_outputs.get(synth_key, "No data")
                        if raw and raw != "No data" and len(raw) > synth_limit:
                            raw = raw[:synth_limit].rsplit("\n", 1)[0]  # cut at last full line
                        system_prompt = system_prompt.replace("{" + synth_key + "}", raw)

                    # Add previous synthesis for progression narrative
                    prev_synthesis = prev_pass_outputs.get("synthesis", "")
                    if prev_synthesis and not date_context.get("is_first_report"):
                        from_date = date_context.get("from_date", "unknown")
                        to_date = date_context.get("to_date", "today")
                        system_prompt += (
                            f"\n\nPREVIOUS EXECUTIVE SUMMARY (from {from_date}):\n{prev_synthesis[:400]}\n\n"
                            f"This report covers {from_date} to {to_date}. "
                            "Highlight what has changed since the previous report — improvements, regressions, and areas that stayed the same."
                        )
                else:
                    continue

                # Apply tone/depth progression based on how many reports have been generated
                report_count = date_context.get("report_count", 0)
                if report_count <= 2:
                    tone_prefix = "This is an early report for a new user. Use a warm, educational tone. Briefly explain what each metric means and why it matters. "
                elif report_count <= 7:
                    tone_prefix = "This teacher has a few reports under their belt. Be direct but provide some context. Focus on changes since last report. "
                else:
                    tone_prefix = "This is an experienced user. Be concise and data-driven. Skip explanations of metrics — focus on actionable deltas and trends. "
                system_prompt = tone_prefix + system_prompt

                user_prompt = "Analyze the data provided and generate your response."
                full_prompt = build_prompt(system_prompt, user_prompt)

                _t1_params = get_tier1_gen_params(task_type) if _is_tier1 else {}
                max_tokens = _t1_params.get("max_tokens", 1000) if _is_tier1 else 1500
                temperature = _t1_params.get("temperature", 0.5) if _is_tier1 else 0.5

                # Debug: send prompt being fed to the LLM
                try:
                    await websocket.send_json({
                        "type": "debug",
                        "stage": "llm_input",
                        "passKey": pass_key,
                        "passName": pass_name,
                        "promptLength": len(full_prompt),
                        "prompt": full_prompt,
                        "maxTokens": max_tokens,
                        "temperature": temperature,
                    })
                except Exception:
                    pass

                # Acquire slot and generate
                slot_mode = None
                accumulated_text = []
                try:
                    slot_mode = await acquire_generation_slot(websocket, generation_mode, f"insights-{pass_key}")
                    from inference_factory import get_inference_instance, resolve_inference_for_task
                    inference = resolve_inference_for_task(task_type) if generation_mode == "queued" else get_inference_instance(use_singleton=False)

                    token_buffer = []
                    last_send = time.time()
                    async for chunk in inference.generate_stream(
                        tool_name=f"insights_{pass_key}",
                        input_data=user_prompt,
                        prompt_template=full_prompt,
                        max_tokens=max_tokens,
                        temperature=temperature,
                    ):
                        if chunk.get("error"):
                            try:
                                await websocket.send_json({"type": "error", "message": chunk["error"]})
                                await asyncio.sleep(0)
                            except Exception:
                                pass
                            break

                        if chunk.get("finished"):
                            if token_buffer:
                                text_chunk = "".join(token_buffer)
                                accumulated_text.append(text_chunk)
                                try:
                                    await websocket.send_json({
                                        "type": "token",
                                        "pass": pass_num,
                                        "content": text_chunk
                                    })
                                    await asyncio.sleep(0)
                                except Exception:
                                    pass
                            break

                        if chunk.get("token"):
                            token_buffer.append(chunk["token"])
                            accumulated_text.append(chunk["token"])
                            if len(token_buffer) >= 10 or (time.time() - last_send) > 0.1:
                                try:
                                    await websocket.send_json({
                                        "type": "token",
                                        "pass": pass_num,
                                        "content": "".join(token_buffer)
                                    })
                                    token_buffer = []
                                    last_send = time.time()
                                    await asyncio.sleep(0)
                                except Exception:
                                    break

                except Exception as e:
                    logger.error(f"Insights pass {pass_name} error: {e}")
                    import traceback
                    logger.error(f"Full traceback: {traceback.format_exc()}")
                    try:
                        await websocket.send_json({"type": "error", "message": f"Pass '{pass_name}' failed: {e}"})
                    except Exception:
                        pass
                finally:
                    try:
                        release_generation_slot(slot_mode or generation_mode)
                    except Exception:
                        pass

                # Store pass output — strip <think>…</think> reasoning blocks
                full_output = "".join(accumulated_text)
                import re as _re
                full_output = _re.sub(r'<think>[\s\S]*?</think>\s*', '', full_output).strip()
                # Also strip unclosed <think> blocks (model didn't close the tag)
                full_output = _re.sub(r'<think>[\s\S]*$', '', full_output).strip()
                pass_outputs[pass_key] = full_output

                # Debug: send LLM output
                try:
                    await websocket.send_json({
                        "type": "debug",
                        "stage": "llm_output",
                        "passKey": pass_key,
                        "passName": pass_name,
                        "outputLength": len(full_output),
                        "output": full_output,
                    })
                except Exception:
                    pass

                # Send pass_complete
                try:
                    await websocket.send_json({
                        "type": "pass_complete",
                        "pass": pass_num,
                        "passName": pass_name,
                        "result": full_output
                    })
                    await asyncio.sleep(0)
                except Exception:
                    break

            # All passes done — build and save report
            report = {
                "id": str(uuid.uuid4()),
                "generated_at": datetime.now().isoformat(),
                "from_date": date_context.get("from_date"),
                "to_date": date_context.get("to_date", datetime.now().strftime("%Y-%m-%d")),
                "previous_report_id": date_context.get("previous_report_id"),
                "report_number": date_context.get("report_count", 0) + 1,
                "passes": [
                    {"key": pd["key"], "name": pd["name"], "output": pass_outputs.get(pd["key"], "")}
                    for pd in PASS_DEFINITIONS
                ],
                "synthesis": pass_outputs.get("synthesis", ""),
            }

            # ── Persistent issue detection & motivational reminders ──
            try:
                from routes.insights import _load_reports
                past_reports = _load_reports()
                # Include the current report for analysis
                all_reports = past_reports + [report]
                # Only look at the last 4 reports (current + 3 previous)
                recent = all_reports[-4:] if len(all_reports) >= 4 else all_reports

                reminders = []
                if len(recent) >= 3:
                    def _check_streak(reports_list, pass_key, check_fn, dimension_label, suggestion):
                        """Check if an issue persists across consecutive reports."""
                        streak = 0
                        for r in reports_list:
                            rp = next((p for p in r.get("passes", []) if p.get("key") == pass_key), None)
                            if rp and check_fn(rp.get("output", "")):
                                streak += 1
                            else:
                                streak = 0
                        if streak >= 3:
                            reminders.append({
                                "dimension": dimension_label,
                                "issue": f"No progress in {dimension_label.lower()} for {streak} consecutive reports.",
                                "streak_count": streak,
                                "suggestion": suggestion,
                            })

                    # Curriculum: no-change or gaps persisting
                    _check_streak(recent, "curriculum",
                        lambda o: o.startswith("No new ") or "no new" in o.lower(),
                        "Curriculum Coverage",
                        "Try updating a few milestones or marking some progress in your curriculum tracker to keep momentum!")

                    # Performance: no new assessments
                    _check_streak(recent, "performance",
                        lambda o: o.startswith("No new ") or "no new" in o.lower(),
                        "Student Performance",
                        "Consider adding a quick quiz or grading a recent assignment to keep performance data fresh.")

                    # Content: no new content created
                    _check_streak(recent, "content",
                        lambda o: o.startswith("No new ") or "no new" in o.lower(),
                        "Content Creation",
                        "Try creating a quick quiz or worksheet to keep your content library growing!")

                    # Attendance: no new records
                    _check_streak(recent, "attendance",
                        lambda o: o.startswith("No new ") or "no new" in o.lower(),
                        "Attendance & Engagement",
                        "Recording attendance regularly helps track student engagement trends over time.")

                    # Achievements: no new achievements
                    _check_streak(recent, "achievements",
                        lambda o: o.startswith("No new ") or "no new" in o.lower(),
                        "Achievements",
                        "Keep using the platform and you'll unlock new achievements naturally!")

                report["reminders"] = reminders
            except Exception as e:
                logger.error(f"Reminders detection error: {e}")
                report["reminders"] = []

            # Save to disk
            try:
                from routes.insights import save_report
                save_report(report)
            except Exception as e:
                logger.error(f"Failed to save insights report: {e}")

            # Send complete
            try:
                await websocket.send_json({
                    "type": "complete",
                    "report": report
                })
                await asyncio.sleep(0)
            except Exception:
                pass

    except WebSocketDisconnect:
        logger.info("Educator Insights WebSocket disconnected")
    except Exception as e:
        logger.error(f"Educator Insights WebSocket error: {e}")


@app.get("/api/quiz-history")
async def get_quiz_history():
    return load_json_data("quiz_history.json")

@app.post("/api/quiz-history")
async def save_quiz_history(data: dict):
    histories = load_json_data("quiz_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("quiz_history.json", histories[-20:])

    # Also save the answer key to DB for scan grading
    parsed_quiz = data.get('parsedQuiz')
    if parsed_quiz and data.get('id'):
        try:
            form_data = data.get('formData', {})
            student_service.save_quiz_answer_key(
                quiz_id=data['id'],
                quiz_title=parsed_quiz.get('metadata', {}).get('title', data.get('title', '')),
                subject=form_data.get('subject', ''),
                grade_level=form_data.get('gradeLevel', ''),
                questions=parsed_quiz.get('questions', []),
            )
            logger.info(f"Answer key saved for quiz {data['id']}")
        except Exception as e:
            logger.error(f"Failed to save answer key: {e}")

    return {"success": True}

@app.delete("/api/quiz-history/{quiz_id}")
async def delete_quiz_history(quiz_id: str):
    histories = load_json_data("quiz_history.json")
    histories = [h for h in histories if h.get("id") != quiz_id]
    save_json_data("quiz_history.json", histories)
    return {"success": True}


# ── Quiz Scan Grading (HunyuanOCR) ──────────────────────────────────────────

@app.get("/api/quiz/answer-keys")
async def list_answer_keys():
    """List all saved quiz answer keys."""
    return student_service.list_quiz_answer_keys()


@app.get("/api/quiz/answer-key/{quiz_id}")
async def get_answer_key(quiz_id: str):
    """Fetch answer key by quiz_id."""
    key = student_service.get_quiz_answer_key(quiz_id)
    if not key:
        raise HTTPException(status_code=404, detail="Quiz answer key not found")
    return key


@app.post("/api/quiz/extract-quiz-id")
async def extract_quiz_id_from_teacher(file: UploadFile = File(...)):
    """Extract quiz_id from a teacher version file (PDF, HTML, or image).

    For digital files (PDF/HTML): text extraction (fast, no OCR needed).
    For images (JPG/PNG): HunyuanOCR reads the quiz_id.
    """
    import re

    content = await file.read()
    fname = (file.filename or '').lower()
    quiz_id = None

    # Digital PDF — extract text directly
    if fname.endswith('.pdf'):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            text = ' '.join(page.extract_text() or '' for page in reader.pages[:2])
            match = re.search(r'quiz_\d+', text)
            if match:
                quiz_id = match.group()
        except Exception:
            pass

    # HTML file
    elif fname.endswith(('.html', '.htm')):
        text = content.decode('utf-8', errors='ignore')
        match = re.search(r'quiz_\d+', text)
        if match:
            quiz_id = match.group()

    # Image — use HunyuanOCR
    elif fname.endswith(('.jpg', '.jpeg', '.png', '.webp')):
        import ocr_service
        if not ocr_service.is_ocr_available():
            raise HTTPException(status_code=400, detail="OCR not available for image files")
        prompt = (
            "Find the Quiz ID on this page. It will be labeled 'Quiz ID:' followed by "
            "a code like 'quiz_1234567890'. Return ONLY the quiz ID value, nothing else."
        )
        loop = asyncio.get_event_loop()
        raw = await loop.run_in_executor(
            ocr_service._ocr_executor, ocr_service._run_ocr, content, prompt, 100
        )
        match = re.search(r'quiz_\d+', raw)
        quiz_id = match.group() if match else None

    if not quiz_id:
        raise HTTPException(status_code=400, detail="Could not find a quiz ID in this file")

    # Check if this quiz_id has an answer key
    key = student_service.get_quiz_answer_key(quiz_id)
    return {
        "quiz_id": quiz_id,
        "found": key is not None,
        "quiz_title": key.get('quiz_title') if key else None,
        "subject": key.get('subject') if key else None,
    }


def _grade_objective_question(question: dict, student_answer: str) -> dict:
    """Grade an objective question by direct comparison. No LLM needed."""
    qtype = question.get('type', '')
    correct = str(question.get('correctAnswer', '')).strip()
    answer = student_answer.strip()
    pts = question.get('points', 1) or 1

    if qtype == 'multiple-choice':
        # Compare letters (A, B, C, D) case-insensitive
        correct_letter = correct.upper()
        if isinstance(question.get('correctAnswer'), (int, float)):
            correct_letter = chr(65 + int(question['correctAnswer']))
        student_letter = answer.upper()[:1] if answer else ''
        is_correct = student_letter == correct_letter
        return {
            'answer': student_letter,
            'earned': pts if is_correct else 0,
            'max': pts,
        }

    elif qtype == 'true-false':
        # Normalize various forms: T/True/true, F/False/false
        student_norm = answer.lower().strip().rstrip('.')
        if student_norm in ('t', 'true', 'yes'):
            student_norm = 'true'
        elif student_norm in ('f', 'false', 'no'):
            student_norm = 'false'
        correct_norm = correct.lower().strip()
        is_correct = student_norm == correct_norm
        return {
            'answer': answer,
            'earned': pts if is_correct else 0,
            'max': pts,
        }

    elif qtype in ('fill-blank', 'word-bank'):
        # Case-insensitive, strip punctuation, allow minor spelling variance
        student_clean = answer.lower().strip().rstrip('.').strip()
        correct_clean = correct.lower().strip().rstrip('.').strip()
        # Accept if contained or exact match
        is_correct = (student_clean == correct_clean or
                      correct_clean in student_clean or
                      student_clean in correct_clean)
        return {
            'answer': answer,
            'earned': pts if is_correct else 0,
            'max': pts,
        }

    elif qtype == 'matching':
        # Matching: compare answer pairs
        return {
            'answer': answer,
            'earned': pts if answer.lower() == correct.lower() else 0,
            'max': pts,
        }

    # Unknown type — mark for LLM
    return None


@app.post("/api/quiz/grade-scans")
async def grade_quiz_scans(
    student_files: List[UploadFile] = File(...),
    quiz_id: str = Form(...),
):
    """Grade scanned student quizzes using HunyuanOCR + direct comparison + LLM.

    1. Fetch answer key from DB using quiz_id
    2. HunyuanOCR extracts student_id + answers from each scan
    3. Objective questions: direct comparison (fast)
    4. Subjective questions: extracted text → LLM for grading
    5. Auto-save grades to student DB
    """
    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR not available")

    # 1. Fetch answer key
    answer_key = student_service.get_quiz_answer_key(quiz_id)
    if not answer_key:
        raise HTTPException(status_code=404, detail=f"No answer key found for quiz ID: {quiz_id}")

    questions = answer_key['questions']
    quiz_title = answer_key['quiz_title']
    subject = answer_key['subject']

    # Build a description of the quiz for OCR extraction
    question_descriptions = []
    for i, q in enumerate(questions, 1):
        qtype = q.get('type', '')
        q_text = q.get('question', '')[:80]
        question_descriptions.append(f"Q{i} [{qtype}]: {q_text}")
    quiz_description = '\n'.join(question_descriptions)

    # Pre-read all files
    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    graded_results = []

    for file_content, filename in file_data:
        result = {
            'file_name': filename,
            'student_name': None,
            'student_id': None,
            'error': None,
            'score': 0,
            'total_points': 0,
            'percentage': 0,
            'letter_grade': 'F',
            'details': {},
            'unclear': [],
            'saved': False,
        }

        try:
            # 2. Preprocess and OCR extract
            processed = _preprocess_phone_image(file_content)

            ocr_prompt = f"""This is a student's completed quiz. Extract the following:
1. Student name (usually at the top)
2. Student ID (usually at the top)
3. The student's answer for each question

The quiz has these questions:
{quiz_description}

Return ONLY valid JSON:
{{
  "student_name": "name or null",
  "student_id": "ID or null",
  "answers": {{
    "1": "student's answer for Q1",
    "2": "student's answer for Q2"
  }}
}}"""
            extracted = await ocr_service.extract_text(processed)
            # Try structured extraction
            import re as _re
            # Run OCR with structured prompt
            loop = asyncio.get_event_loop()
            raw_json = await loop.run_in_executor(
                ocr_service._ocr_executor, ocr_service._run_ocr, processed, ocr_prompt, 2048
            )

            json_match = _re.search(r'\{[\s\S]*\}', raw_json)
            if not json_match:
                result['error'] = 'OCR could not extract answers from this scan.'
                graded_results.append(result)
                continue

            parsed = json.loads(json_match.group())
            student_name = parsed.get('student_name')
            student_id = parsed.get('student_id')
            student_answers = parsed.get('answers', {})

            result['student_name'] = student_name
            result['student_id'] = student_id

            # 3. Grade each question
            subjective_questions = []  # Collect for LLM batch
            details = {}

            for i, q in enumerate(questions, 1):
                q_key = str(i)
                student_answer = student_answers.get(q_key, '')

                if not student_answer or student_answer.lower() in ('[unclear]', 'unclear', ''):
                    details[q_key] = {
                        'answer': student_answer,
                        'earned': 0,
                        'max': q.get('points', 1) or 1,
                    }
                    result['unclear'].append(i)
                    continue

                qtype = q.get('type', '')
                if qtype in ('multiple-choice', 'true-false', 'fill-blank', 'word-bank', 'matching'):
                    # Direct comparison
                    grade_result = _grade_objective_question(q, student_answer)
                    if grade_result:
                        details[q_key] = grade_result
                    else:
                        subjective_questions.append((q_key, q, student_answer))
                else:
                    # Short-answer, comprehension, open-ended → LLM
                    subjective_questions.append((q_key, q, student_answer))

            # 4. LLM grades subjective questions (batched)
            if subjective_questions:
                from inference_factory import get_inference_instance
                inference = get_inference_instance()

                subj_lines = []
                for q_key, q, ans in subjective_questions:
                    pts = q.get('points', 1) or 1
                    ref = q.get('correctAnswer', '')
                    subj_lines.append(
                        f"Q{q_key} [{q.get('type')}, {pts}pt]: \"{q.get('question')}\"\n"
                        f"  Reference answer: {ref}\n"
                        f"  Student wrote: {ans}"
                    )

                llm_prompt = f"""Grade these subjective questions. Award partial credit (0 to max) based on accuracy and completeness.

{chr(10).join(subj_lines)}

Return ONLY valid JSON:
{{
  {', '.join(f'"{qk}": {{"earned": <points>, "max": {q.get("points",1) or 1}, "feedback": "brief reason"}}' for qk, q, _ in subjective_questions)}
}}"""
                response = await inference.generate(
                    tool_name="quiz_subjective_grading",
                    input_data="subjective grading",
                    prompt_template=llm_prompt,
                    max_tokens=500,
                    temperature=0.1,
                )
                raw_llm = response.get('result') or ''
                llm_match = _re.search(r'\{[\s\S]*\}', raw_llm)
                if llm_match:
                    llm_grades = json.loads(llm_match.group())
                    for q_key, q, ans in subjective_questions:
                        g = llm_grades.get(q_key, {})
                        details[q_key] = {
                            'answer': ans,
                            'earned': g.get('earned', 0),
                            'max': q.get('points', 1) or 1,
                            'feedback': g.get('feedback', ''),
                        }
                else:
                    # LLM failed — mark as 0
                    for q_key, q, ans in subjective_questions:
                        details[q_key] = {
                            'answer': ans,
                            'earned': 0,
                            'max': q.get('points', 1) or 1,
                            'feedback': 'Could not grade automatically',
                        }

            # 5. Calculate totals
            total_earned = sum(d.get('earned', 0) for d in details.values())
            total_max = sum(d.get('max', 1) for d in details.values())
            percentage = round((total_earned / total_max * 100), 1) if total_max > 0 else 0

            result.update({
                'score': total_earned,
                'total_points': total_max,
                'percentage': percentage,
                'letter_grade': student_service.get_letter_grade(percentage),
                'details': details,
            })

            # 6. Auto-save to student DB
            if student_id:
                try:
                    student_service.save_quiz_grade({
                        'student_id': student_id,
                        'quiz_title': quiz_title,
                        'subject': subject,
                        'score': total_earned,
                        'total_points': total_max,
                        'percentage': percentage,
                        'letter_grade': student_service.get_letter_grade(percentage),
                        'answers': details,
                    })
                    result['saved'] = True
                except Exception as e:
                    logger.error(f"Failed to save grade for {student_id}: {e}")

        except Exception as e:
            logger.error(f"Scan grading error for {filename}: {e}")
            result['error'] = f'Processing error: {str(e)}'

        graded_results.append(result)

    return graded_results


@app.post("/api/quiz/grade-scans-stream")
async def grade_quiz_scans_stream(
    request: Request,
    student_files: List[UploadFile] = File(...),
    quiz_id: str = Form(...),
):
    """SSE streaming version of quiz scan grading."""
    from starlette.responses import StreamingResponse

    # Validate up front
    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR not available")

    answer_key = student_service.get_quiz_answer_key(quiz_id)
    if not answer_key:
        raise HTTPException(status_code=404, detail=f"No answer key found for quiz ID: {quiz_id}")

    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    total = len(file_data)

    async def event_generator():
        # Re-use the batch endpoint logic per file
        for idx, (content, filename) in enumerate(file_data):
            # Grade this single file using the same logic
            # (Simplified: call the batch endpoint's inner logic)
            try:
                # Create a temporary UploadFile-like for the batch grader
                # Instead, just call grade_quiz_scans inline per-file
                result = await _grade_single_quiz_scan(
                    content, filename, answer_key
                )
            except Exception as e:
                result = {
                    'file_name': filename,
                    'error': str(e),
                    'score': 0, 'total_points': 0, 'percentage': 0,
                    'letter_grade': 'F', 'details': {}, 'unclear': [],
                    'saved': False,
                }

            event = json.dumps({
                'index': idx,
                'total': total,
                'result': result,
                'done': idx == total - 1,
            })
            yield f"data: {event}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


async def _grade_single_quiz_scan(file_content: bytes, filename: str, answer_key: dict) -> dict:
    """Grade a single student quiz scan. Extracted for streaming reuse."""
    import re as _re
    import ocr_service

    questions = answer_key['questions']
    quiz_title = answer_key['quiz_title']
    subject = answer_key['subject']

    question_descriptions = []
    for i, q in enumerate(questions, 1):
        q_text = q.get('question', '')[:80]
        question_descriptions.append(f"Q{i} [{q.get('type','')}]: {q_text}")
    quiz_description = '\n'.join(question_descriptions)

    result = {
        'file_name': filename,
        'student_name': None,
        'student_id': None,
        'error': None,
        'score': 0,
        'total_points': 0,
        'percentage': 0,
        'letter_grade': 'F',
        'details': {},
        'unclear': [],
        'saved': False,
    }

    try:
        processed = _preprocess_phone_image(file_content)

        ocr_prompt = f"""This is a student's completed quiz. Extract the following:
1. Student name (usually at the top)
2. Student ID (usually at the top)
3. The student's answer for each question

The quiz has these questions:
{quiz_description}

Return ONLY valid JSON:
{{
  "student_name": "name or null",
  "student_id": "ID or null",
  "answers": {{
    "1": "student's answer for Q1",
    "2": "student's answer for Q2"
  }}
}}"""
        loop = asyncio.get_event_loop()
        raw_json = await loop.run_in_executor(
            ocr_service._ocr_executor, ocr_service._run_ocr, processed, ocr_prompt, 2048
        )

        json_match = _re.search(r'\{[\s\S]*\}', raw_json)
        if not json_match:
            result['error'] = 'OCR could not extract answers from this scan.'
            return result

        parsed = json.loads(json_match.group())
        student_name = parsed.get('student_name')
        student_id = parsed.get('student_id')
        student_answers = parsed.get('answers', {})

        result['student_name'] = student_name
        result['student_id'] = student_id

        subjective_questions = []
        details = {}

        for i, q in enumerate(questions, 1):
            q_key = str(i)
            student_answer = student_answers.get(q_key, '')

            if not student_answer or student_answer.lower() in ('[unclear]', 'unclear', ''):
                details[q_key] = {
                    'answer': student_answer,
                    'earned': 0,
                    'max': q.get('points', 1) or 1,
                }
                result['unclear'].append(i)
                continue

            qtype = q.get('type', '')
            if qtype in ('multiple-choice', 'true-false', 'fill-blank', 'word-bank', 'matching'):
                grade_result = _grade_objective_question(q, student_answer)
                if grade_result:
                    details[q_key] = grade_result
                else:
                    subjective_questions.append((q_key, q, student_answer))
            else:
                subjective_questions.append((q_key, q, student_answer))

        if subjective_questions:
            from inference_factory import get_inference_instance
            inference = get_inference_instance()

            subj_lines = []
            for q_key, q, ans in subjective_questions:
                pts = q.get('points', 1) or 1
                ref = q.get('correctAnswer', '')
                subj_lines.append(
                    f"Q{q_key} [{q.get('type')}, {pts}pt]: \"{q.get('question')}\"\n"
                    f"  Reference answer: {ref}\n"
                    f"  Student wrote: {ans}"
                )

            llm_prompt = f"""Grade these subjective questions. Award partial credit (0 to max) based on accuracy and completeness.

{chr(10).join(subj_lines)}

Return ONLY valid JSON:
{{
  {', '.join(f'"{qk}": {{"earned": <points>, "max": {q.get("points",1) or 1}, "feedback": "brief reason"}}' for qk, q, _ in subjective_questions)}
}}"""
            response = await inference.generate(
                tool_name="quiz_subjective_grading",
                input_data="subjective grading",
                prompt_template=llm_prompt,
                max_tokens=500,
                temperature=0.1,
            )
            raw_llm = response.get('result') or ''
            llm_match = _re.search(r'\{[\s\S]*\}', raw_llm)
            if llm_match:
                llm_grades = json.loads(llm_match.group())
                for q_key, q, ans in subjective_questions:
                    g = llm_grades.get(q_key, {})
                    details[q_key] = {
                        'answer': ans,
                        'earned': g.get('earned', 0),
                        'max': q.get('points', 1) or 1,
                        'feedback': g.get('feedback', ''),
                    }
            else:
                for q_key, q, ans in subjective_questions:
                    details[q_key] = {
                        'answer': ans,
                        'earned': 0,
                        'max': q.get('points', 1) or 1,
                        'feedback': 'Could not grade automatically',
                    }

        total_earned = sum(d.get('earned', 0) for d in details.values())
        total_max = sum(d.get('max', 1) for d in details.values())
        percentage = round((total_earned / total_max * 100), 1) if total_max > 0 else 0

        result.update({
            'score': total_earned,
            'total_points': total_max,
            'percentage': percentage,
            'letter_grade': student_service.get_letter_grade(percentage),
            'details': details,
        })

        if student_id:
            try:
                student_service.save_quiz_grade({
                    'student_id': student_id,
                    'quiz_title': quiz_title,
                    'subject': subject,
                    'score': total_earned,
                    'total_points': total_max,
                    'percentage': percentage,
                    'letter_grade': student_service.get_letter_grade(percentage),
                    'answers': details,
                })
                result['saved'] = True
            except Exception as e:
                logger.error(f"Failed to save grade for {student_id}: {e}")

    except Exception as e:
        logger.error(f"Scan grading error for {filename}: {e}")
        result['error'] = f'Processing error: {str(e)}'

    return result


# ── Worksheet Scan Grading (HunyuanOCR) ─────────────────────────────────────

@app.get("/api/worksheet/answer-keys")
async def list_worksheet_answer_keys():
    """List all saved worksheet answer keys."""
    return student_service.list_worksheet_answer_keys()


@app.get("/api/worksheet/answer-key/{worksheet_id}")
async def get_worksheet_answer_key(worksheet_id: str):
    """Fetch worksheet answer key by worksheet_id."""
    key = student_service.get_worksheet_answer_key(worksheet_id)
    if not key:
        raise HTTPException(status_code=404, detail="Worksheet answer key not found")
    return key


@app.post("/api/worksheet/extract-worksheet-id")
async def extract_worksheet_id_from_teacher(file: UploadFile = File(...)):
    """Extract worksheet_id from a teacher version file."""
    import re

    content = await file.read()
    fname = (file.filename or '').lower()
    ws_id = None

    if fname.endswith('.pdf'):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            text = ' '.join(page.extract_text() or '' for page in reader.pages[:2])
            match = re.search(r'worksheet_\d+', text)
            if match:
                ws_id = match.group()
        except Exception:
            pass
    elif fname.endswith(('.html', '.htm')):
        text = content.decode('utf-8', errors='ignore')
        match = re.search(r'worksheet_\d+', text)
        if match:
            ws_id = match.group()
    elif fname.endswith(('.jpg', '.jpeg', '.png', '.webp')):
        import ocr_service
        if not ocr_service.is_ocr_available():
            raise HTTPException(status_code=400, detail="OCR not available for image files")
        prompt = (
            "Find the Worksheet ID on this page. It will be labeled 'Worksheet ID:' followed by "
            "a code like 'worksheet_1234567890'. Return ONLY the worksheet ID value, nothing else."
        )
        loop = asyncio.get_event_loop()
        raw = await loop.run_in_executor(
            ocr_service._ocr_executor, ocr_service._run_ocr, content, prompt, 100
        )
        match = re.search(r'worksheet_\d+', raw)
        ws_id = match.group() if match else None

    if not ws_id:
        raise HTTPException(status_code=400, detail="Could not find a worksheet ID in this file")

    key = student_service.get_worksheet_answer_key(ws_id)
    return {
        "worksheet_id": ws_id,
        "found": key is not None,
        "worksheet_title": key.get('worksheet_title') if key else None,
        "subject": key.get('subject') if key else None,
    }


@app.post("/api/worksheet/grade-scans")
async def grade_worksheet_scans(
    student_files: List[UploadFile] = File(...),
    worksheet_id: str = Form(...),
):
    """Grade scanned student worksheets using HunyuanOCR + direct comparison + LLM."""
    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR not available")

    answer_key = student_service.get_worksheet_answer_key(worksheet_id)
    if not answer_key:
        raise HTTPException(status_code=404, detail=f"No answer key found for worksheet ID: {worksheet_id}")

    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    graded_results = []
    for file_content, filename in file_data:
        result = await _grade_single_worksheet_scan(file_content, filename, answer_key)
        graded_results.append(result)

    return graded_results


@app.post("/api/worksheet/grade-scans-stream")
async def grade_worksheet_scans_stream(
    request: Request,
    student_files: List[UploadFile] = File(...),
    worksheet_id: str = Form(...),
):
    """SSE streaming version of worksheet scan grading."""
    from starlette.responses import StreamingResponse

    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR not available")

    answer_key = student_service.get_worksheet_answer_key(worksheet_id)
    if not answer_key:
        raise HTTPException(status_code=404, detail=f"No answer key found for worksheet ID: {worksheet_id}")

    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    total = len(file_data)

    async def event_generator():
        for idx, (content, filename) in enumerate(file_data):
            try:
                result = await _grade_single_worksheet_scan(content, filename, answer_key)
            except Exception as e:
                result = {
                    'file_name': filename, 'error': str(e),
                    'score': 0, 'total_points': 0, 'percentage': 0,
                    'letter_grade': 'F', 'details': {}, 'unclear': [], 'saved': False,
                }
            event = json.dumps({'index': idx, 'total': total, 'result': result, 'done': idx == total - 1})
            yield f"data: {event}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


async def _grade_single_worksheet_scan(file_content: bytes, filename: str, answer_key: dict) -> dict:
    """Grade a single student worksheet scan using OCR + direct comparison + LLM."""
    import re as _re
    import ocr_service

    questions = answer_key['questions']
    ws_title = answer_key['worksheet_title']
    subject = answer_key['subject']
    passage = answer_key.get('passage', '')
    matching_items = answer_key.get('matching_items', {})
    word_bank = answer_key.get('word_bank', [])

    question_descriptions = []
    for i, q in enumerate(questions, 1):
        q_text = q.get('question', '')[:80]
        question_descriptions.append(f"Q{i} [{q.get('type', '')}]: {q_text}")
    ws_description = '\n'.join(question_descriptions)

    result = {
        'file_name': filename, 'student_name': None, 'student_id': None,
        'error': None, 'score': 0, 'total_points': 0, 'percentage': 0,
        'letter_grade': 'F', 'details': {}, 'unclear': [], 'saved': False,
    }

    try:
        processed = _preprocess_phone_image(file_content)

        ocr_prompt = f"""This is a student's completed worksheet. Extract the following:
1. Student name (usually at the top)
2. Student ID (usually at the top)
3. The student's answer for each question

The worksheet has these questions:
{ws_description}

Return ONLY valid JSON:
{{
  "student_name": "name or null",
  "student_id": "ID or null",
  "answers": {{
    "1": "student's answer for Q1",
    "2": "student's answer for Q2"
  }}
}}"""
        loop = asyncio.get_event_loop()
        raw_json = await loop.run_in_executor(
            ocr_service._ocr_executor, ocr_service._run_ocr, processed, ocr_prompt, 2048
        )

        json_match = _re.search(r'\{[\s\S]*\}', raw_json)
        if not json_match:
            result['error'] = 'OCR could not extract answers from this scan.'
            return result

        parsed = json.loads(json_match.group())
        student_name = parsed.get('student_name')
        student_id = parsed.get('student_id')
        student_answers = parsed.get('answers', {})

        result['student_name'] = student_name
        result['student_id'] = student_id

        subjective_questions = []
        details = {}

        for i, q in enumerate(questions, 1):
            q_key = str(i)
            student_answer = student_answers.get(q_key, '')

            if not student_answer or student_answer.lower() in ('[unclear]', 'unclear', ''):
                details[q_key] = {'answer': student_answer, 'earned': 0, 'max': q.get('points', 1) or 1}
                result['unclear'].append(i)
                continue

            qtype = q.get('type', '')
            if qtype in ('multiple-choice', 'true-false', 'fill-blank', 'word-bank', 'matching'):
                grade_result = _grade_objective_question(q, student_answer)
                if grade_result:
                    details[q_key] = grade_result
                else:
                    subjective_questions.append((q_key, q, student_answer))
            else:
                subjective_questions.append((q_key, q, student_answer))

        # LLM grades subjective questions
        if subjective_questions:
            from inference_factory import get_inference_instance
            inference = get_inference_instance()

            subj_lines = []
            for q_key, q, ans in subjective_questions:
                pts = q.get('points', 1) or 1
                ref = q.get('correctAnswer', '')
                subj_lines.append(
                    f"Q{q_key} [{q.get('type')}, {pts}pt]: \"{q.get('question')}\"\n"
                    f"  Reference answer: {ref}\n"
                    f"  Student wrote: {ans}"
                )

            passage_ctx = f'\n\nPassage context:\n"""{passage[:1500]}"""' if passage else ''

            llm_prompt = f"""Grade these subjective questions. Award partial credit (0 to max) based on accuracy and completeness.{passage_ctx}

{chr(10).join(subj_lines)}

Return ONLY valid JSON:
{{
  {', '.join(f'"{qk}": {{"earned": <points>, "max": {q.get("points",1) or 1}, "feedback": "brief reason"}}' for qk, q, _ in subjective_questions)}
}}"""
            response = await inference.generate(
                tool_name="worksheet_subjective_grading",
                input_data="subjective grading",
                prompt_template=llm_prompt,
                max_tokens=500,
                temperature=0.1,
            )
            raw_llm = response.get('result') or ''
            llm_match = _re.search(r'\{[\s\S]*\}', raw_llm)
            if llm_match:
                llm_grades = json.loads(llm_match.group())
                for q_key, q, ans in subjective_questions:
                    g = llm_grades.get(q_key, {})
                    details[q_key] = {
                        'answer': ans, 'earned': g.get('earned', 0),
                        'max': q.get('points', 1) or 1, 'feedback': g.get('feedback', ''),
                    }
            else:
                for q_key, q, ans in subjective_questions:
                    details[q_key] = {
                        'answer': ans, 'earned': 0,
                        'max': q.get('points', 1) or 1, 'feedback': 'Could not grade automatically',
                    }

        total_earned = sum(d.get('earned', 0) for d in details.values())
        total_max = sum(d.get('max', 1) for d in details.values())
        percentage = round((total_earned / total_max * 100), 1) if total_max > 0 else 0

        result.update({
            'score': total_earned, 'total_points': total_max,
            'percentage': percentage,
            'letter_grade': student_service.get_letter_grade(percentage),
            'details': details,
        })

        if student_id:
            try:
                student_service.save_worksheet_grade({
                    'student_id': student_id, 'worksheet_title': ws_title,
                    'subject': subject, 'score': total_earned,
                    'total_points': total_max, 'percentage': percentage,
                    'letter_grade': student_service.get_letter_grade(percentage),
                    'answers': details,
                })
                result['saved'] = True
            except Exception as e:
                logger.error(f"Failed to save worksheet grade for {student_id}: {e}")

    except Exception as e:
        logger.error(f"Worksheet scan grading error for {filename}: {e}")
        result['error'] = f'Processing error: {str(e)}'

    return result


# ── Student Management ────────────────────────────────────────────────────────

@app.get("/api/students")
async def get_students(class_name: Optional[str] = None):
    return student_service.list_students(class_name)

@app.post("/api/students")
async def add_student(student: StudentCreate):
    return student_service.create_student(student.model_dump())

@app.get("/api/students/sample-excel")
async def download_sample_excel():
    """Generate and return a sample Excel file with student columns and example data."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Class List"

    headers = ["Full Name", "Date of Birth", "Class", "Grade Level", "Gender", "Contact Info"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    sample_data = [
        ["Jane Smith", "2015-03-12", "A", "3", "Female", "jane.parent@email.com"],
        ["Marcus Johnson", "2014-11-05", "A", "3", "Male", "555-0123"],
        ["Aaliyah Charles", "2015-06-22", "A", "3", "Female", "555-0456"],
        ["Devon Williams", "2014-09-18", "B", "4", "Male", "devon.parent@email.com"],
        ["Keisha Brown", "2013-12-01", "B", "4", "Female", "555-0789"],
        ["Tyler Joseph", "2015-01-30", "A", "3", "Male", "tyler.mom@email.com"],
        ["Sophia Pierre", "2014-07-14", "C", "4", "Female", "555-0321"],
        ["Elijah Francis", "2013-04-09", "D", "5", "Male", "eli.dad@email.com"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from starlette.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sample_class_list.xlsx"}
    )


@app.post("/api/students/import-excel")
async def import_excel(file: UploadFile = File(...)):
    """Import students from an uploaded Excel (.xlsx/.xls) or CSV file."""
    fname = (file.filename or "").lower()
    content = await file.read()

    rows: list[dict] = []

    if fname.endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            raise HTTPException(status_code=400, detail="Excel file has no data rows.")
        headers = [str(h or '').strip() for h in all_rows[0]]
        for data_row in all_rows[1:]:
            row_dict = {}
            for h, v in zip(headers, data_row):
                row_dict[h] = str(v) if v is not None else ''
            rows.append(row_dict)
        wb.close()
    elif fname.endswith('.csv'):
        import csv, io
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .xlsx, .xls, or .csv")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file.")

    result = student_service.bulk_import_students(rows)
    return result

@app.get("/api/students/{student_id}")
async def get_student_profile(student_id: str):
    s = student_service.get_student(student_id)
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return s

@app.put("/api/students/{student_id}")
async def update_student_profile(student_id: str, student: StudentCreate):
    s = student_service.update_student(student_id, student.model_dump())
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return s

@app.delete("/api/students/{student_id}")
async def remove_student(student_id: str):
    student_service.delete_student(student_id)
    return {"success": True}

@app.get("/api/classes")
async def get_classes(grade_level: str | None = None):
    return student_service.list_classes(grade_level=grade_level)

@app.post("/api/quiz-grades")
async def add_quiz_grade(grade: QuizGradeCreate):
    return student_service.save_quiz_grade(grade.model_dump())

@app.get("/api/quiz-grades/student/{student_id}")
async def get_grades_for_student(student_id: str):
    return student_service.get_student_grades(student_id)

@app.post("/api/worksheet-grades")
async def add_worksheet_grade(grade: WorksheetGradeCreate):
    return student_service.save_worksheet_grade(grade.model_dump())

@app.get("/api/worksheet-grades/student/{student_id}")
async def get_worksheet_grades_for_student(student_id: str):
    return student_service.get_worksheet_grades(student_id)


# ── Worksheet Packages ───────────────────────────────────────────────────────

@app.post("/api/worksheet-packages")
async def save_worksheet_package(data: dict):
    return student_service.save_worksheet_package(data)

@app.get("/api/worksheet-packages")
async def list_worksheet_packages(class_name: Optional[str] = None):
    return student_service.list_worksheet_packages(class_name)

@app.get("/api/worksheet-packages/{package_id}")
async def get_worksheet_package(package_id: str):
    pkg = student_service.get_worksheet_package(package_id)
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")
    return pkg


@app.get("/api/attendance")
async def get_attendance(class_name: str, date: str):
    return student_service.get_attendance(class_name, date)

@app.post("/api/attendance")
async def save_attendance(body: AttendanceSave):
    return student_service.save_attendance([r.model_dump() for r in body.records])


# ── Bulk Auto-Grading ─────────────────────────────────────────────────────────

def _extract_text_from_file(content: bytes, filename: str, content_type: str) -> tuple[str, str | None]:
    """
    Extract plain text from an uploaded file.
    Returns (extracted_text, error_message).
    """
    fname_lower = filename.lower()

    # HTML
    if fname_lower.endswith(('.html', '.htm')) or 'html' in content_type:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(content.decode('utf-8', errors='replace'), 'html.parser')
        # Remove script/style noise
        for tag in soup(['script', 'style']):
            tag.decompose()
        return soup.get_text(separator='\n', strip=True), None

    # PDF
    if fname_lower.endswith('.pdf') or 'pdf' in content_type:
        try:
            import io as _io
            from pypdf import PdfReader
            reader = PdfReader(_io.BytesIO(content))
            pages = [page.extract_text() or '' for page in reader.pages]
            text = '\n'.join(pages).strip()
            if not text:
                return '', 'PDF has no extractable text (may be a scanned image).'
            return text, None
        except ImportError:
            return '', 'PDF support requires pypdf: pip install pypdf'
        except Exception as e:
            return '', f'PDF read error: {e}'

    # Plain text / markdown
    if fname_lower.endswith(('.txt', '.md')) or 'text' in content_type:
        return content.decode('utf-8', errors='replace'), None

    # Image — no reliable OCR without tesseract
    if fname_lower.endswith(('.jpg', '.jpeg', '.png', '.webp', '.bmp')):
        return '', 'Image files cannot be read automatically. Please use HTML or PDF exports from the app.'

    return content.decode('utf-8', errors='replace'), None


def _build_extraction_prompt(quiz_questions: list, text: str) -> str:
    """Build LLM prompt that extracts student answers from quiz text."""
    question_hints = []
    for i, q in enumerate(quiz_questions, 1):
        qtype = q.get('type', '')
        q_text = q.get('question', '')[:80]
        if qtype == 'multiple-choice':
            options = q.get('options', [])
            letters = ', '.join(chr(65 + j) for j in range(len(options)))
            question_hints.append(f"Q{i} (Multiple Choice – answer is one of: {letters}): {q_text}")
        elif qtype == 'true-false':
            question_hints.append(f"Q{i} (True/False – answer is exactly True or False): {q_text}")
        elif qtype == 'fill-blank':
            question_hints.append(f"Q{i} (Fill in the blank – answer is a word or phrase): {q_text}")
        # open-ended: skip

    hints_block = '\n'.join(question_hints)

    return f"""You are a quiz grading assistant. A student's completed quiz is shown below.

IMPORTANT: The student's quiz may have questions in a DIFFERENT ORDER than the teacher's version.
You must match each student question to the correct teacher question by its content/text, not by its number.

Teacher's question list (use these to match):
{hints_block}

Rules:
- For each question on the student's paper, find the matching teacher question by content.
- Use the TEACHER's question number (Q1, Q2, etc.) as the key in your response, NOT the student's question number.
- For Multiple Choice, return only the letter (A, B, C, or D).
- For True/False, return exactly "True" or "False".
- For Fill in the blank, return the word or phrase the student wrote.
- If a question has no visible answer, return null.
- If a student name or ID is visible on the paper, extract them; otherwise return null.

Student quiz text:
\"\"\"
{text[:3000]}
\"\"\"

Return ONLY valid JSON in this exact structure, no explanation:
{{
  "student_name": null,
  "student_id": null,
  "answers": {{
    "1": "A",
    "2": "True",
    "3": "photosynthesis"
  }}
}}"""


def _grade_extracted_answers(quiz_questions: list, extracted_answers: dict) -> dict:
    """Grade extracted student answers against the answer key."""
    results = []
    total = 0
    earned = 0

    for i, q in enumerate(quiz_questions, 1):
        qtype = q.get('type', '')
        if qtype == 'open-ended':
            continue  # skip

        pts = q.get('points', 1) or 1
        total += pts
        student_raw = extracted_answers.get(str(i))

        correct = False
        if student_raw is not None:
            correct_answer = q.get('correctAnswer')
            if qtype == 'multiple-choice':
                try:
                    letter_index = ord(str(student_raw).strip().upper()) - 65
                    correct = letter_index == int(correct_answer)
                except Exception:
                    correct = False
            elif qtype == 'true-false':
                correct = str(student_raw).strip().lower() == str(correct_answer).strip().lower()
            elif qtype == 'fill-blank':
                student_norm = str(student_raw).strip().lower()
                accepted = [a.strip().lower() for a in str(correct_answer or '').split(',')]
                correct = student_norm in accepted

        points_earned = pts if correct else 0
        earned += points_earned
        results.append({
            'question_number': i,
            'question': q.get('question', ''),
            'type': qtype,
            'student_answer': student_raw,
            'correct_answer': q.get('correctAnswer'),
            'correct': correct,
            'points_earned': points_earned,
            'points_possible': pts,
        })

    percentage = round((earned / total * 100), 1) if total > 0 else 0
    return {
        'results': results,
        'score': earned,
        'total_points': total,
        'percentage': percentage,
        'letter_grade': student_service.get_letter_grade(percentage),
    }


@app.post("/api/parse-teacher-quiz")
async def parse_teacher_quiz(teacher_file: UploadFile = File(...)):
    """
    Parse a teacher version quiz file (HTML or PDF) and extract the answer key
    as a ParsedQuiz-compatible JSON structure.
    """
    content = await teacher_file.read()
    filename = teacher_file.filename or ''
    content_type = teacher_file.content_type or ''

    text, extract_error = _extract_text_from_file(content, filename, content_type)
    if extract_error:
        raise HTTPException(status_code=422, detail=extract_error)
    if not text.strip():
        raise HTTPException(status_code=422, detail="No text could be extracted from this file.")

    prompt = f"""You are a quiz parser. Extract all questions and answers from this teacher quiz.

Return ONLY valid JSON in this exact format (no markdown, no explanation):
{{
  "metadata": {{
    "title": "Quiz Title",
    "subject": "Subject",
    "gradeLevel": "5",
    "totalQuestions": 5
  }},
  "questions": [
    {{
      "id": "q_1",
      "type": "multiple-choice",
      "question": "Question text here",
      "options": ["Option A text", "Option B text", "Option C text", "Option D text"],
      "correctAnswer": 0,
      "points": 1
    }},
    {{
      "id": "q_2",
      "type": "true-false",
      "question": "Statement here",
      "options": ["True", "False"],
      "correctAnswer": "true",
      "points": 1
    }},
    {{
      "id": "q_3",
      "type": "fill-blank",
      "question": "The ___ is the powerhouse of the cell.",
      "correctAnswer": "mitochondria",
      "points": 1
    }}
  ]
}}

Rules:
- For multiple-choice: correctAnswer is a 0-based index (0=A, 1=B, 2=C, 3=D)
- For true-false: correctAnswer is "true" or "false" (lowercase string)
- For fill-blank: correctAnswer is the expected word or phrase
- Skip open-ended questions entirely
- Extract the subject and grade level from the quiz header if present

Quiz text:
\"\"\"
{text[:4000]}
\"\"\"
"""

    from inference_factory import get_inference_instance
    inference = get_inference_instance()
    llm_result = await inference.generate(
        tool_name='parse_teacher_quiz',
        input_data=text[:200],
        prompt_template=prompt,
        max_tokens=2000,
        temperature=0.1,
        stop=['```'],
    )

    raw = llm_result.get('result') or ''

    import re as _re
    json_match = _re.search(r'\{[\s\S]*\}', raw)
    if not json_match:
        raise HTTPException(status_code=422, detail="Could not extract quiz structure. Make sure the file is a teacher version with answers marked.")

    try:
        parsed = json.loads(json_match.group())
        # Ensure required structure
        if 'questions' not in parsed or not isinstance(parsed['questions'], list):
            raise ValueError("Missing questions array")
        if 'metadata' not in parsed:
            parsed['metadata'] = {'title': 'Uploaded Quiz', 'subject': '', 'gradeLevel': '', 'totalQuestions': len(parsed['questions'])}
        return parsed
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Parsed output was not valid quiz JSON: {e}")


async def _grade_single_text_file(
    content: bytes,
    filename: str,
    content_type: str,
    questions: list,
    inference,
    max_tokens: int,
) -> dict:
    """Grade a single text-based student file (HTML/PDF/TXT). Shared by quiz and worksheet bulk grading."""
    import re as _re

    file_result = {
        'file_name': filename,
        'student_name': None,
        'error': None,
        'score': 0,
        'total_points': 0,
        'percentage': 0,
        'letter_grade': 'F',
        'details': [],
    }

    try:
        text, extract_error = _extract_text_from_file(content, filename, content_type)
        if extract_error:
            file_result['error'] = extract_error
            return file_result

        if not text.strip():
            file_result['error'] = 'No text could be extracted from this file.'
            return file_result

        prompt = _build_extraction_prompt(questions, text)
        llm_result = await inference.generate(
            tool_name='bulk_grade_extract',
            input_data=text[:500],
            prompt_template=prompt,
            max_tokens=max_tokens,
            temperature=0.1,
            stop=['```', '\n\n\n'],
        )

        raw_output = llm_result.get('result') or ''

        json_match = _re.search(r'\{[\s\S]*\}', raw_output)
        if not json_match:
            file_result['error'] = 'LLM could not extract answers from this file.'
            return file_result

        extracted = json.loads(json_match.group())
        student_name = extracted.get('student_name')
        extracted_answers = extracted.get('answers', {})

        grading = _grade_extracted_answers(questions, extracted_answers)

        file_result.update({
            'student_name': student_name,
            'score': grading['score'],
            'total_points': grading['total_points'],
            'percentage': grading['percentage'],
            'letter_grade': grading['letter_grade'],
            'details': grading['results'],
        })

    except Exception as e:
        file_result['error'] = f'Processing error: {str(e)}'

    return file_result


@app.post("/api/bulk-grade")
async def bulk_grade(
    quiz_json: str = Form(...),
    student_files: List[UploadFile] = File(...),
):
    """
    Accept a quiz answer key (as JSON) and multiple student files.
    Extract student answers from each file using LLM, then auto-grade.
    """
    try:
        quiz_data = json.loads(quiz_json)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid quiz_json")

    quiz_questions = quiz_data.get('questions', [])
    if not quiz_questions:
        raise HTTPException(status_code=400, detail="Quiz has no questions")

    from inference_factory import get_inference_instance
    inference = get_inference_instance()
    max_tokens = _compute_dynamic_max_tokens(quiz_questions)

    # Pre-read all files
    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown', upload.content_type or ''))

    graded_results = []
    for content, filename, ctype in file_data:
        result = await _grade_single_text_file(content, filename, ctype, quiz_questions, inference, max_tokens)
        graded_results.append(result)

    return graded_results


@app.post("/api/bulk-grade-worksheet")
async def bulk_grade_worksheet(
    worksheet_json: str = Form(...),
    student_files: List[UploadFile] = File(...),
):
    """
    Accept a worksheet answer key (as JSON) and multiple student files.
    Extract student answers from each file using LLM, then auto-grade.
    """
    try:
        ws_data = json.loads(worksheet_json)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid worksheet_json")

    ws_questions = ws_data.get('questions', [])
    if not ws_questions:
        raise HTTPException(status_code=400, detail="Worksheet has no questions")

    from inference_factory import get_inference_instance
    inference = get_inference_instance()
    max_tokens = _compute_dynamic_max_tokens(ws_questions)

    # Pre-read all files
    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown', upload.content_type or ''))

    graded_results = []
    for content, filename, ctype in file_data:
        result = await _grade_single_text_file(content, filename, ctype, ws_questions, inference, max_tokens)
        graded_results.append(result)

    return graded_results


# ── Image Preprocessing for Phone Photos ──────────────────────────────────────

def _preprocess_phone_image(image_bytes: bytes, max_width: int = 1024) -> bytes:
    """Preprocess a phone photo for optimal vision model grading.

    - Auto-rotates based on EXIF orientation
    - Resizes to max_width while keeping aspect ratio
    - Normalizes contrast/brightness for readability
    - Converts to JPEG for consistent smaller payloads
    """
    from PIL import Image, ImageOps, ImageEnhance
    import io as _io

    img = Image.open(_io.BytesIO(image_bytes))

    # 1. Auto-rotate based on EXIF orientation data (phone camera metadata)
    img = ImageOps.exif_transpose(img)

    # 2. Convert to RGB (handles RGBA, palette-mode, etc.)
    if img.mode != 'RGB':
        img = img.convert('RGB')

    # 3. Resize if wider than max_width (preserves aspect ratio)
    if img.width > max_width:
        ratio = max_width / img.width
        new_size = (max_width, int(img.height * ratio))
        img = img.resize(new_size, Image.LANCZOS)

    # 4. Auto-contrast to normalize exposure from phone cameras
    img = ImageOps.autocontrast(img, cutoff=1)

    # 5. Slight sharpening to help with handwriting legibility
    enhancer = ImageEnhance.Sharpness(img)
    img = enhancer.enhance(1.3)

    # 6. Encode as JPEG (much smaller than PNG for photos)
    buf = _io.BytesIO()
    img.save(buf, format='JPEG', quality=85, optimize=True)
    return buf.getvalue()


def _compute_dynamic_max_tokens(questions: list) -> int:
    """Calculate max_tokens based on question count and types.

    MC/TF need ~20 tokens each; fill-blank ~30; short-answer/comprehension ~60.
    Add a base of 80 for the JSON wrapper + student_name/id + unclear array.
    """
    base = 80
    per_question = 0
    for q in questions:
        qtype = q.get('type', '')
        if qtype in ('multiple-choice', 'true-false'):
            per_question += 20
        elif qtype in ('fill-blank', 'word-bank', 'matching'):
            per_question += 30
        else:  # short-answer, comprehension, etc.
            per_question += 60
    # Clamp between 200 and 2500
    return max(200, min(base + per_question, 2500))


# ── Scan-to-Grade (Vision) ────────────────────────────────────────────────────

def _build_vision_grading_prompt(base_worksheet: dict, student_versions: list) -> str:
    """Build a prompt for the vision model to read and grade a scanned worksheet."""
    questions = base_worksheet.get('questions', [])
    passage = base_worksheet.get('passage', '')
    matching = base_worksheet.get('matchingItems', {})
    word_bank = base_worksheet.get('wordBank', [])

    lines = []
    for i, q in enumerate(questions, 1):
        qtype = q.get('type', '')
        pts = q.get('points', 1) or 1
        q_text = q.get('question', '')
        correct = q.get('correctAnswer', '')

        if qtype == 'multiple-choice':
            options = q.get('options', [])
            opts_str = ', '.join(f"{chr(65+j)}) {o}" for j, o in enumerate(options))
            correct_letter = chr(65 + int(correct)) if isinstance(correct, (int, float)) else correct
            lines.append(f"Q{i} [Multiple Choice, {pts}pt]: \"{q_text}\" Options: {opts_str} → Correct: {correct_letter}")

        elif qtype == 'true-false':
            lines.append(f"Q{i} [True/False, {pts}pt]: \"{q_text}\" → Correct: {correct}")

        elif qtype == 'fill-blank':
            lines.append(f"Q{i} [Fill-in-the-blank, {pts}pt]: \"{q_text}\" → Correct: {correct}")

        elif qtype == 'word-bank':
            lines.append(f"Q{i} [Word Bank, {pts}pt]: \"{q_text}\" → Correct: {correct}")

        elif qtype == 'matching':
            lines.append(f"Q{i} [Matching, {pts}pt]: \"{q_text}\" → Correct: {correct}")

        elif qtype == 'short-answer':
            lines.append(f"Q{i} [Short Answer, {pts}pt]: \"{q_text}\" → Reference: {correct}")

        elif qtype == 'comprehension':
            lines.append(f"Q{i} [Comprehension, {pts}pt]: \"{q_text}\" → Reference: {correct}")

        else:
            lines.append(f"Q{i} [{qtype}, {pts}pt]: \"{q_text}\" → Correct: {correct}")

    answer_key_block = '\n'.join(lines)

    # Add matching columns if present
    matching_block = ''
    if matching:
        col_a = matching.get('columnA', [])
        col_b = matching.get('columnB', [])
        if col_a and col_b:
            matching_block = f"\n\nMatching Columns:\nColumn A: {', '.join(f'{j+1}. {a}' for j, a in enumerate(col_a))}\nColumn B: {', '.join(f'{chr(65+j)}. {b}' for j, b in enumerate(col_b))}"

    # Add word bank if present
    word_bank_block = ''
    if word_bank:
        word_bank_block = f"\n\nWord Bank: {', '.join(word_bank)}"

    # Add passage if present
    passage_block = ''
    if passage:
        passage_block = f'\n\nPassage used:\n"""{passage[:1500]}"""'

    # Student IDs for matching
    student_ids = [sv.get('student', {}).get('id', '') for sv in student_versions]
    id_block = f"\n\nKnown student IDs in this class: {', '.join(student_ids)}" if student_ids else ''

    return f"""You are grading a student's handwritten worksheet from a scanned image.
Read the image carefully and grade every question.

ANSWER KEY:
{answer_key_block}{matching_block}{word_bank_block}{passage_block}{id_block}

INSTRUCTIONS:
1. Find the student's name and student ID printed at the top of the page.
2. Questions may appear in a DIFFERENT ORDER on the student's paper. Match each question by its content, not its number. Use the Q numbers from the answer key above.
3. Read what the student wrote for each question.

GRADING RULES:
- Multiple Choice: correct only if the exact letter matches. Return the letter.
- True/False: correct only if it matches exactly. Return "True" or "False".
- Fill-in-the-blank: correct if the word/phrase matches (minor spelling errors OK for young students). Return the word(s).
- Word Bank: correct if it matches the expected word. Return the word.
- Matching: correct for each pair that matches. Return an object mapping item numbers to letters.
- Math: correct if the numerical answer matches. Return the number.
- Comprehension / Short Answer: award partial credit (0 to max points) based on accuracy and completeness. Return the student's full written response.

If you cannot read an answer, set earned to 0 and add the question number to "unclear".

Return ONLY valid JSON with no extra text:
{{
  "student_name": "name or null",
  "student_id": "ID or null",
  "results": {{
    "1": {{"answer": "B", "earned": 1, "max": 1}},
    "2": {{"answer": "True", "earned": 1, "max": 1}},
    "3": {{"answer": "the student wrote this...", "earned": 2, "max": 2, "feedback": "brief reason"}}
  }},
  "unclear": []
}}"""


def _build_text_grading_prompt(base_worksheet: dict, student_versions: list, extracted_text: str) -> str:
    """Build a prompt for the text LLM to grade OCR-extracted student answers."""
    questions = base_worksheet.get('questions', [])
    passage = base_worksheet.get('passage', '')
    matching = base_worksheet.get('matchingItems', {})
    word_bank = base_worksheet.get('wordBank', [])

    lines = []
    for i, q in enumerate(questions, 1):
        qtype = q.get('type', '')
        pts = q.get('points', 1) or 1
        q_text = q.get('question', '')
        correct = q.get('correctAnswer', '')

        if qtype == 'multiple-choice':
            options = q.get('options', [])
            opts_str = ', '.join(f"{chr(65+j)}) {o}" for j, o in enumerate(options))
            correct_letter = chr(65 + int(correct)) if isinstance(correct, (int, float)) else correct
            lines.append(f"Q{i} [Multiple Choice, {pts}pt]: \"{q_text}\" Options: {opts_str} → Correct: {correct_letter}")
        elif qtype == 'true-false':
            lines.append(f"Q{i} [True/False, {pts}pt]: \"{q_text}\" → Correct: {correct}")
        elif qtype == 'fill-blank':
            lines.append(f"Q{i} [Fill-in-the-blank, {pts}pt]: \"{q_text}\" → Correct: {correct}")
        elif qtype == 'word-bank':
            lines.append(f"Q{i} [Word Bank, {pts}pt]: \"{q_text}\" → Correct: {correct}")
        elif qtype == 'matching':
            lines.append(f"Q{i} [Matching, {pts}pt]: \"{q_text}\" → Correct: {correct}")
        elif qtype == 'short-answer':
            lines.append(f"Q{i} [Short Answer, {pts}pt]: \"{q_text}\" → Reference: {correct}")
        elif qtype == 'comprehension':
            lines.append(f"Q{i} [Comprehension, {pts}pt]: \"{q_text}\" → Reference: {correct}")
        else:
            lines.append(f"Q{i} [{qtype}, {pts}pt]: \"{q_text}\" → Correct: {correct}")

    answer_key_block = '\n'.join(lines)

    matching_block = ''
    if matching:
        col_a = matching.get('columnA', [])
        col_b = matching.get('columnB', [])
        if col_a and col_b:
            matching_block = f"\n\nMatching Columns:\nColumn A: {', '.join(f'{j+1}. {a}' for j, a in enumerate(col_a))}\nColumn B: {', '.join(f'{chr(65+j)}. {b}' for j, b in enumerate(col_b))}"

    word_bank_block = ''
    if word_bank:
        word_bank_block = f"\n\nWord Bank: {', '.join(word_bank)}"

    passage_block = ''
    if passage:
        passage_block = f'\n\nPassage used:\n"""{passage[:1500]}"""'

    student_ids = [sv.get('student', {}).get('id', '') for sv in student_versions]
    id_block = f"\n\nKnown student IDs in this class: {', '.join(student_ids)}" if student_ids else ''

    return f"""You are grading a student's worksheet. An OCR system has already extracted the text from the scanned image. Grade every question based on the extracted text below.

ANSWER KEY:
{answer_key_block}{matching_block}{word_bank_block}{passage_block}{id_block}

EXTRACTED STUDENT WORK (from OCR):
\"\"\"
{extracted_text}
\"\"\"

INSTRUCTIONS:
1. Find the student's name and student ID from the extracted text.
2. Questions may appear in a DIFFERENT ORDER. Match each question by its content, not its number. Use the Q numbers from the answer key above.
3. Compare the student's answers against the answer key.

GRADING RULES:
- Multiple Choice: correct only if the exact letter matches. Return the letter.
- True/False: correct only if it matches exactly. Return "True" or "False".
- Fill-in-the-blank: correct if the word/phrase matches (minor spelling errors OK for young students). Return the word(s).
- Word Bank: correct if it matches the expected word. Return the word.
- Matching: correct for each pair that matches. Return an object mapping item numbers to letters.
- Math: correct if the numerical answer matches. Return the number.
- Comprehension / Short Answer: award partial credit (0 to max points) based on accuracy and completeness. Return the student's full written response.

If the OCR marked something as [unclear], set earned to 0 and add the question number to "unclear".

Return ONLY valid JSON with no extra text:
{{
  "student_name": "name or null",
  "student_id": "ID or null",
  "results": {{
    "1": {{"answer": "B", "earned": 1, "max": 1}},
    "2": {{"answer": "True", "earned": 1, "max": 1}},
    "3": {{"answer": "the student wrote this...", "earned": 2, "max": 2, "feedback": "brief reason"}}
  }},
  "unclear": []
}}"""


async def _grade_single_scan(
    file_content: bytes,
    filename: str,
    prompt: str,
    max_tokens: int,
    inference,
    package: dict | None,
    student_versions: list,
    worksheet_title: str,
    worksheet_subject: str,
    use_ocr: bool = False,
    base_worksheet: dict | None = None,
) -> dict:
    """Grade a single scanned student worksheet image.

    Two modes:
    - use_ocr=True: HunyuanOCR extracts text → text LLM grades (two-stage, more accurate)
    - use_ocr=False: Vision LLM reads + grades in one shot (original behavior)
    """
    import re as _re

    file_result = {
        'file_name': filename,
        'student_name': None,
        'student_id': None,
        'error': None,
        'score': 0,
        'total_points': 0,
        'percentage': 0,
        'letter_grade': 'F',
        'details': {},
        'unclear': [],
        'saved': False,
        'ocr_used': use_ocr,
    }

    try:
        # Preprocess phone photo (resize, rotate, enhance)
        processed = _preprocess_phone_image(file_content)

        if use_ocr:
            # ── Two-stage: HunyuanOCR reads → LLM grades ──
            import ocr_service
            extracted_text = await ocr_service.extract_text_for_grading(processed)
            file_result['extracted_text'] = extracted_text

            # Build a text-only grading prompt with the OCR output
            text_prompt = _build_text_grading_prompt(
                base_worksheet, student_versions, extracted_text
            )

            # Use the text LLM (no vision needed) to grade
            response = await inference.generate(
                tool_name="ocr_grading",
                input_data=extracted_text[:100],
                prompt_template=text_prompt,
                max_tokens=max_tokens,
                temperature=0.1,
            )
            raw_output = response.get('result') or ''

        else:
            # ── Single-stage: Vision LLM reads + grades ──
            image_b64 = base64.b64encode(processed).decode('utf-8')
            response = await inference.analyze_image(
                image_base64=image_b64,
                prompt=prompt,
                max_tokens=max_tokens,
                temperature=0.1,
            )
            raw_output = response.get('result') or ''

        json_match = _re.search(r'\{[\s\S]*\}', raw_output)
        if not json_match:
            file_result['error'] = 'Could not extract grading results from model output.'
            return file_result

        parsed = json.loads(json_match.group())
        student_name = parsed.get('student_name')
        student_id = parsed.get('student_id')
        results = parsed.get('results', {})
        unclear = parsed.get('unclear', [])

        total_earned = sum(r.get('earned', 0) for r in results.values())
        total_max = sum(r.get('max', 1) for r in results.values())
        percentage = round((total_earned / total_max * 100), 1) if total_max > 0 else 0

        file_result.update({
            'student_name': student_name,
            'student_id': student_id,
            'score': total_earned,
            'total_points': total_max,
            'percentage': percentage,
            'letter_grade': student_service.get_letter_grade(percentage),
            'details': results,
            'unclear': unclear,
        })

        # Auto-save grade if student ID matched
        if student_id:
            title = package['worksheet_title'] if package else worksheet_title
            subject = package['subject'] if package else worksheet_subject

            should_save = True
            if student_versions:
                should_save = any(
                    sv.get('student', {}).get('id') == student_id
                    for sv in student_versions
                )

            if should_save:
                student_service.save_worksheet_grade({
                    'student_id': student_id,
                    'worksheet_title': title,
                    'subject': subject,
                    'score': total_earned,
                    'total_points': total_max,
                    'percentage': percentage,
                    'letter_grade': student_service.get_letter_grade(percentage),
                    'answers': results,
                })
                file_result['saved'] = True

    except Exception as e:
        logger.error(f"Scan grading error for {filename}: {e}")
        file_result['error'] = f'Processing error: {str(e)}'

    return file_result


def _load_grading_context(package_id, worksheet_json):
    """Load answer key and student versions from package or JSON. Returns (package, base_worksheet, student_versions)."""
    package = None
    base_worksheet = None
    student_versions = []

    if package_id:
        package = student_service.get_worksheet_package(package_id)
        if not package:
            raise HTTPException(status_code=404, detail="Worksheet package not found")
        base_worksheet = json.loads(package['base_worksheet']) if isinstance(package['base_worksheet'], str) else package['base_worksheet']
        student_versions = json.loads(package['student_versions']) if isinstance(package['student_versions'], str) else package['student_versions']
    elif worksheet_json:
        try:
            base_worksheet = json.loads(worksheet_json)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid worksheet JSON")
    else:
        raise HTTPException(status_code=400, detail="Provide either package_id or worksheet_json")

    return package, base_worksheet, student_versions


@app.post("/api/grade-scanned-worksheets")
async def grade_scanned_worksheets(
    student_files: List[UploadFile] = File(...),
    package_id: Optional[str] = Form(None),
    worksheet_json: Optional[str] = Form(None),
    worksheet_title: Optional[str] = Form("Worksheet"),
    worksheet_subject: Optional[str] = Form("General"),
):
    """Grade scanned student worksheets using OCR + LLM or vision model.

    Accepts either a package_id (for class-mode worksheets) or
    worksheet_json (a ParsedWorksheet JSON for standalone worksheets).
    """
    package, base_worksheet, student_versions = _load_grading_context(package_id, worksheet_json)

    prompt = _build_vision_grading_prompt(base_worksheet, student_versions)
    max_tokens = _compute_dynamic_max_tokens(base_worksheet.get('questions', []))

    from inference_factory import get_inference_instance
    from config import get_ocr_enabled
    inference = get_inference_instance()

    # Determine if we should use OCR pipeline
    use_ocr = False
    if get_ocr_enabled():
        try:
            import ocr_service
            if ocr_service.is_ocr_available():
                use_ocr = True
        except ImportError:
            pass

    if not use_ocr and not getattr(inference, 'has_vision', False):
        raise HTTPException(status_code=400, detail="No grading backend available. Enable OCR or load a multimodal vision model.")

    # Pre-read all files so we can process them
    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    # Grade all files sequentially
    graded_results = []
    for content, filename in file_data:
        result = await _grade_single_scan(
            content, filename, prompt, max_tokens, inference,
            package, student_versions, worksheet_title, worksheet_subject,
            use_ocr=use_ocr, base_worksheet=base_worksheet,
        )
        graded_results.append(result)

    # Mark package as graded if all saved
    if package_id and all(r.get('saved') for r in graded_results if not r.get('error')):
        student_service.mark_package_graded(package_id)

    return graded_results


@app.post("/api/grade-scanned-worksheets-stream")
async def grade_scanned_worksheets_stream(
    request: Request,
    student_files: List[UploadFile] = File(...),
    package_id: Optional[str] = Form(None),
    worksheet_json: Optional[str] = Form(None),
    worksheet_title: Optional[str] = Form("Worksheet"),
    worksheet_subject: Optional[str] = Form("General"),
):
    """SSE streaming version — sends each graded result as it completes.

    Each SSE event is a JSON object with:
      - index: 0-based position in the file list
      - total: total number of files
      - result: the graded result for this file
      - done: true on the final event
    """
    from starlette.responses import StreamingResponse

    package, base_worksheet, student_versions = _load_grading_context(package_id, worksheet_json)
    prompt = _build_vision_grading_prompt(base_worksheet, student_versions)
    max_tokens = _compute_dynamic_max_tokens(base_worksheet.get('questions', []))

    from inference_factory import get_inference_instance
    from config import get_ocr_enabled
    inference = get_inference_instance()

    # Determine if we should use OCR pipeline
    use_ocr = False
    if get_ocr_enabled():
        try:
            import ocr_service
            if ocr_service.is_ocr_available():
                use_ocr = True
        except ImportError:
            pass

    if not use_ocr and not getattr(inference, 'has_vision', False):
        raise HTTPException(status_code=400, detail="No grading backend available. Enable OCR or load a multimodal vision model.")

    # Pre-read all files
    file_data = []
    for upload in student_files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    total = len(file_data)

    async def event_generator():
        all_results = []
        for idx, (content, filename) in enumerate(file_data):
            result = await _grade_single_scan(
                content, filename, prompt, max_tokens, inference,
                package, student_versions, worksheet_title, worksheet_subject,
                use_ocr=use_ocr, base_worksheet=base_worksheet,
            )
            all_results.append(result)

            event = json.dumps({
                'index': idx,
                'total': total,
                'result': result,
                'done': idx == total - 1,
            })
            yield f"data: {event}\n\n"

        # Mark package as graded if all saved
        if package_id and all(r.get('saved') for r in all_results if not r.get('error')):
            student_service.mark_package_graded(package_id)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ── OCR Service Endpoints ────────────────────────────────────────────────────

@app.get("/api/ocr/status")
async def ocr_status():
    """Get HunyuanOCR service status."""
    import ocr_service
    status = ocr_service.get_ocr_status()
    from config import get_ocr_enabled
    status["enabled"] = get_ocr_enabled()
    return status


@app.post("/api/ocr/toggle")
async def ocr_toggle(data: dict):
    """Enable/disable OCR for scan grading."""
    from config import set_ocr_enabled
    enabled = data.get("enabled", True)
    set_ocr_enabled(enabled)
    return {"enabled": enabled}


@app.post("/api/ocr/load")
async def ocr_load():
    """Pre-load the OCR model into VRAM."""
    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR dependencies not installed (bitsandbytes, torch, transformers)")
    try:
        ocr_service.load_ocr_model()
        return {"status": "loaded"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ocr/unload")
async def ocr_unload():
    """Unload OCR model to free VRAM."""
    import ocr_service
    ocr_service.unload_ocr_model()
    return {"status": "unloaded"}


@app.post("/api/ocr/extract")
async def ocr_extract(file: UploadFile = File(...)):
    """Extract text from an uploaded image using HunyuanOCR."""
    import ocr_service
    if not ocr_service.is_ocr_available():
        raise HTTPException(status_code=400, detail="OCR dependencies not installed")
    content = await file.read()
    try:
        text = await ocr_service.extract_text(content)
        return {"text": text, "filename": file.filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR extraction failed: {str(e)}")


@app.get("/api/ocr-models")
async def get_available_ocr_models():
    """Get list of available OCR models in the models directory."""
    try:
        from config import scan_ocr_models
        models = scan_ocr_models()
        return {
            "success": True,
            "models": models,
            "count": len(models),
        }
    except Exception as e:
        logger.error(f"Error retrieving OCR models: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving OCR models: {str(e)}")


@app.post("/api/ocr-models/select")
async def select_ocr_model(request: Request):
    """Set the active OCR model."""
    try:
        from config import set_selected_ocr_model, get_selected_ocr_model
        data = await request.json()
        model_name = data.get("modelName", "")

        if not model_name:
            return JSONResponse(status_code=400, content={"error": "Model name is required"})

        old_model = get_selected_ocr_model()
        if set_selected_ocr_model(model_name):
            # Unload old model if it was loaded so the new one loads on next use
            if model_name != old_model:
                try:
                    import ocr_service
                    if ocr_service.is_ocr_loaded():
                        ocr_service.unload_ocr_model()
                        logger.info(f"Unloaded previous OCR model ({old_model}) for switch to {model_name}")
                except Exception as e:
                    logger.warning(f"Could not unload previous OCR model: {e}")

            return JSONResponse(content={
                "success": True,
                "message": f"OCR model set to {model_name}",
                "selectedModel": model_name,
            })
        else:
            return JSONResponse(status_code=500, content={"error": "Failed to save OCR model selection"})
    except Exception as e:
        logger.error(f"Error selecting OCR model: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/rubric-history")
async def get_rubric_history():
    return load_json_data("rubric_history.json")

@app.post("/api/rubric-history")
async def save_rubric_history(data: dict):
    histories = load_json_data("rubric_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("rubric_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/rubric-history/{rubric_id}")
async def delete_rubric_history(rubric_id: str):
    histories = load_json_data("rubric_history.json")
    histories = [h for h in histories if h.get("id") != rubric_id]
    save_json_data("rubric_history.json", histories)
    return {"success": True}


@app.get("/api/kindergarten-history")
async def get_kindergarten_history():
    return load_json_data("kindergarten_history.json")

@app.post("/api/kindergarten-history")
async def save_kindergarten_history(data: dict):
    histories = load_json_data("kindergarten_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("kindergarten_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/kindergarten-history/{plan_id}")
async def delete_kindergarten_history(plan_id: str):
    histories = load_json_data("kindergarten_history.json")
    histories = [h for h in histories if h.get("id") != plan_id]
    save_json_data("kindergarten_history.json", histories)
    return {"success": True}


@app.get("/api/multigrade-history")
async def get_multigrade_history():
    return load_json_data("multigrade_history.json")

@app.post("/api/multigrade-history")
async def save_multigrade_history(data: dict):
    histories = load_json_data("multigrade_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("multigrade_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/multigrade-history/{plan_id}")
async def delete_multigrade_history(plan_id: str):
    histories = load_json_data("multigrade_history.json")
    histories = [h for h in histories if h.get("id") != plan_id]
    save_json_data("multigrade_history.json", histories)
    return {"success": True}


@app.get("/api/cross-curricular-history")
async def get_cross_curricular_history():
    return load_json_data("cross_curricular_history.json")

@app.post("/api/cross-curricular-history")
async def save_cross_curricular_history(data: dict):
    histories = load_json_data("cross_curricular_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("cross_curricular_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/cross-curricular-history/{plan_id}")
async def delete_cross_curricular_history(plan_id: str):
    histories = load_json_data("cross_curricular_history.json")
    histories = [h for h in histories if h.get("id") != plan_id]
    save_json_data("cross_curricular_history.json", histories)
    return {"success": True}

@app.get("/api/worksheet-history")
async def get_worksheet_history():
    return load_json_data("worksheet_history.json")

@app.post("/api/worksheet-history")
async def save_worksheet_history(data: dict):
    histories = load_json_data("worksheet_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("worksheet_history.json", histories[-20:])

    # Also save the answer key to DB for scan grading
    parsed_ws = data.get('parsedWorksheet')
    if parsed_ws and data.get('id'):
        try:
            meta = parsed_ws.get('metadata', {})
            form_data = data.get('formData', {})
            student_service.save_worksheet_answer_key(
                worksheet_id=data['id'],
                worksheet_title=meta.get('title', data.get('title', '')),
                subject=form_data.get('subject', meta.get('subject', '')),
                grade_level=form_data.get('gradeLevel', meta.get('gradeLevel', '')),
                questions=parsed_ws.get('questions', []),
                passage=parsed_ws.get('passage', ''),
                matching_items=parsed_ws.get('matchingItems', {}),
                word_bank=parsed_ws.get('wordBank', []),
            )
            logger.info(f"Worksheet answer key saved for {data['id']}")
        except Exception as e:
            logger.error(f"Failed to save worksheet answer key: {e}")

    return {"success": True}

@app.delete("/api/worksheet-history/{worksheet_id}")
async def delete_worksheet_history(worksheet_id: str):
    histories = load_json_data("worksheet_history.json")
    histories = [h for h in histories if h.get("id") != worksheet_id]
    save_json_data("worksheet_history.json", histories)
    return {"success": True}

@app.get("/api/images-history")
async def get_images_history():
    return load_json_data("images_history.json")

@app.post("/api/images-history")
async def save_images_history(data: dict):
    histories = load_json_data("images_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("images_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/images-history/{image_id}")
async def delete_images_history(image_id: str):
    histories = load_json_data("images_history.json")
    histories = [h for h in histories if h.get("id") != image_id]
    save_json_data("images_history.json", histories)
    return {"success": True}


# ── Presentation Image Analysis ──

@app.post("/api/analyze-presentation-images")
async def analyze_presentation_images(request: Request):
    """Analyze uploaded images using the multimodal model for presentation context."""
    data = await request.json()
    images = data.get("images", [])  # list of { dataUri, filename }
    subject = data.get("subject", "General")
    grade = data.get("grade", "4")
    topic = data.get("topic", "Lesson")

    if not images:
        raise HTTPException(status_code=400, detail="No images provided.")

    from inference_factory import get_inference_instance
    inference = get_inference_instance()

    if not getattr(inference, 'has_vision', False):
        raise HTTPException(status_code=400, detail="Vision model not available. Load a multimodal model to analyze images.")

    analyses = []
    for i, img in enumerate(images):
        image_b64 = img.get("dataUri", "")
        if not image_b64:
            continue

        # Strip data URI prefix — analyze_image handles the re-addition
        raw_b64 = image_b64.split(",", 1)[1] if "," in image_b64 else image_b64

        prompt = (
            f"This image will be used in an educational presentation for Grade {grade} {subject} "
            f"about '{topic}'. Describe what this image shows in 2-3 sentences. "
            f"Focus on: what is depicted, the educational relevance, and which part of a lesson "
            f"(introduction/hook, instruction, activity, or assessment) this image would best support."
        )

        try:
            result = await inference.analyze_image(
                image_base64=raw_b64,
                prompt=prompt,
                max_tokens=300,
                temperature=0.4,
            )
            description = result.get("result", "Unable to analyze image")
        except Exception as e:
            logger.error(f"Image analysis failed for image {i}: {e}")
            description = "Unable to analyze image"

        analyses.append({
            "imageIndex": i,
            "filename": img.get("filename", f"image_{i}"),
            "description": description,
        })

    return {"success": True, "analyses": analyses}


# ── Presentation History ──

@app.get("/api/presentation-history")
async def get_presentation_history():
    return load_json_data("presentation_history.json")

@app.post("/api/presentation-history")
async def save_presentation_history(data: dict):
    histories = load_json_data("presentation_history.json")
    existing = next((h for h in histories if h.get("id") == data.get("id")), None)
    if existing:
        histories = [h for h in histories if h.get("id") != data.get("id")]
    histories.append(data)
    save_json_data("presentation_history.json", histories[-20:])
    return {"success": True}

@app.delete("/api/presentation-history/{presentation_id}")
async def delete_presentation_history(presentation_id: str):
    histories = load_json_data("presentation_history.json")
    histories = [h for h in histories if h.get("id") != presentation_id]
    save_json_data("presentation_history.json", histories)
    return {"success": True}


def scan_models_directory():
    """Scan the models directory for available AI model files"""
    models = []
    
    if not MODELS_DIR.exists():
        MODELS_DIR.mkdir(parents=True, exist_ok=True)
        return models
    
    model_extensions = ['.gguf', '.bin', '.ggml']
    
    # Vision projector files are auxiliary — hide them from the selector
    vision_keywords = ["vision", "mmproj", "clip-model"]

    # OCR models belong in the OCR section, not the main model dropdown
    tier_config = get_tier_config()
    ocr_model_names = [m.lower() for m in tier_config.get("ocr_models", [])]
    # Also catch any file with "ocr" in the name as a safety net
    ocr_keywords = ["paddleocr", "hunyuanocr"]

    try:
        for file_path in MODELS_DIR.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in model_extensions:
                # Skip vision projector files
                name_lower = file_path.name.lower()
                if any(kw in name_lower for kw in vision_keywords):
                    continue

                # Skip OCR models (by tier config list + keyword fallback)
                if any(ocr_name in name_lower for ocr_name in ocr_model_names):
                    continue
                if any(kw in name_lower for kw in ocr_keywords):
                    continue

                file_size_mb = file_path.stat().st_size / (1024 * 1024)

                family_config = LlamaInference.detect_model_family(str(file_path))
                model_info = {
                    "name": file_path.name,
                    "path": str(file_path),
                    "size_mb": round(file_size_mb, 2),
                    "extension": file_path.suffix,
                    "is_active": file_path.name == get_selected_model(),
                    "supports_thinking": family_config.get("supports_thinking", False),
                }
                models.append(model_info)

        models.sort(key=lambda x: x["name"])
        
    except Exception as e:
        logger.error(f"Error scanning models directory: {e}")
    
    return models


@app.get("/api/models")
async def get_available_models():
    """Get list of available AI models in the models directory"""
    try:
        models = scan_models_directory()
        return {
            "success": True,
            "models": models,
            "models_directory": str(MODELS_DIR),
            "count": len(models)
        }
    except Exception as e:
        logger.error(f"Error retrieving models: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving models: {str(e)}")


@app.post("/api/models/select")
async def select_model(request: Request):
    """Set the active model for generation"""
    try:
        data = await request.json()
        model_name = data.get('modelName')
        
        if not model_name:
            return JSONResponse(
                status_code=400,
                content={"error": "Model name is required"}
            )
        
        model_path = MODELS_DIR / model_name
        if not model_path.exists():
            return JSONResponse(
                status_code=404,
                content={"error": f"Model not found: {model_name}"}
            )
        
        if set_selected_model(model_name):
            # Reload singleton so it picks up the new model
            try:
                from inference_factory import reload_local_model
                reload_local_model()
            except Exception as reload_err:
                logger.warning(f"Could not reload model singleton: {reload_err}")

            return JSONResponse(content={
                "success": True,
                "message": f"Model set to {model_name}",
                "selectedModel": model_name
            })
        else:
            return JSONResponse(
                status_code=500,
                content={"error": "Failed to save model selection"}
            )
            
    except Exception as e:
        logger.error(f"Error selecting model: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.get("/api/models/active")
async def get_active_model():
    """Get the currently active model"""
    try:
        selected_model = get_selected_model()
        model_path = get_model_path()

        has_vision = False
        try:
            from inference_factory import get_inference_instance
            instance = get_inference_instance(use_singleton=True)
            has_vision = getattr(instance, "has_vision", False)
        except Exception:
            pass

        return JSONResponse(content={
            "modelName": selected_model,
            "modelPath": model_path,
            "exists": Path(model_path).exists(),
            "has_vision": has_vision,
        })
    except Exception as e:
        logger.error(f"Error getting active model: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


# ============================================================================
# TIER & CAPABILITIES ENDPOINTS
# ============================================================================

@app.get("/api/capabilities")
async def get_capabilities():
    """Get computed tier and capability flags based on current model selection."""
    try:
        return JSONResponse(content=compute_effective_tier())
    except Exception as e:
        logger.error(f"Error computing capabilities: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/tier-config")
async def get_tier_config_endpoint():
    """Get the raw tier configuration."""
    try:
        return JSONResponse(content=get_tier_config())
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/tier-config")
async def update_tier_config(request: Request):
    """Update the full tier configuration."""
    try:
        body = await request.json()
        if set_tier_config(body):
            return JSONResponse(content={"status": "ok", "config": body})
        return JSONResponse(status_code=500, content={"error": "Failed to save tier config"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/tier-config/assign")
async def assign_model_tier(request: Request):
    """Quick assign a model to a tier. Body: { model: str, tier: 1|2 }"""
    try:
        body = await request.json()
        model = body.get("model")
        tier = body.get("tier", 1)
        if not model:
            return JSONResponse(status_code=400, content={"error": "model is required"})

        config = get_tier_config()
        # Remove from both lists first
        for key in ("tier1_models", "tier2_models"):
            config[key] = [m for m in config.get(key, []) if m.lower() != model.lower()]
        # Add to target list
        target_key = "tier2_models" if tier == 2 else "tier1_models"
        config[target_key].append(model)

        if set_tier_config(config):
            return JSONResponse(content={"status": "ok", "model": model, "tier": tier})
        return JSONResponse(status_code=500, content={"error": "Failed to save"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/tier-config/dual-model")
async def update_dual_model_config(request: Request):
    """Update dual-model routing settings."""
    try:
        body = await request.json()
        config = get_tier_config()
        # Merge incoming dual_model settings
        existing_dual = config.get("dual_model", {})
        existing_dual.update(body)
        config["dual_model"] = existing_dual

        if set_tier_config(config):
            return JSONResponse(content={"status": "ok", "dual_model": existing_dual})
        return JSONResponse(status_code=500, content={"error": "Failed to save"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/model/preload")
async def preload_model():
    """Preload the LLM model in the background.

    This triggers lazy loading of the model when a user opens
    an AI-powered tab, so the model is ready by the time they
    click Generate.
    """
    import traceback
    from inference_factory import get_inference_instance

    def _load_model():
        """Run model loading in a thread so it doesn't block the event loop."""
        return get_inference_instance()

    last_error = None
    for attempt in range(2):
        try:
            if attempt > 0:
                logger.info(f"Retrying model preload (attempt {attempt + 1})...")
                await asyncio.sleep(2)
            else:
                logger.info("Preloading LLM model...")
            loop = asyncio.get_event_loop()
            inference = await loop.run_in_executor(None, _load_model)
            if inference.is_loaded:
                logger.info("LLM model preloaded successfully")
                return {"status": "loaded", "has_vision": getattr(inference, 'has_vision', False)}
            else:
                last_error = "Model loaded but is_loaded=False"
                logger.error(f"Failed to preload LLM model (attempt {attempt + 1})")
        except Exception as e:
            last_error = str(e)
            logger.error(f"Error preloading model (attempt {attempt + 1}): {e}\n{traceback.format_exc()}")

    return JSONResponse(status_code=500, content={"status": "error", "error": last_error or "Unknown error"})


@app.get("/api/vision/status")
async def vision_status():
    """Check if the current model has vision capabilities."""
    try:
        from inference_factory import get_inference_instance
        instance = get_inference_instance(use_singleton=True)
        has_vision = getattr(instance, "has_vision", False)
        return {"has_vision": has_vision, "model": get_selected_model()}
    except Exception as e:
        return {"has_vision": False, "error": str(e)}


@app.post("/api/vision/analyze")
async def vision_analyze(request: Request):
    """Analyze an image using the multimodal model's vision capabilities.

    Expects JSON: { "image": "<base64>", "prompt": "Describe this image." }
    """
    try:
        data = await request.json()
        image_b64 = data.get("image")
        prompt = data.get("prompt", "Describe this image in detail.")

        if not image_b64:
            return JSONResponse(status_code=400, content={"error": "image (base64) is required"})

        from inference_factory import get_inference_instance
        instance = get_inference_instance(use_singleton=True)

        if not getattr(instance, "has_vision", False):
            return JSONResponse(status_code=400, content={
                "error": "Vision not available. The current model does not have a vision projector loaded."
            })

        result = await instance.analyze_image(
            image_base64=image_b64,
            prompt=prompt,
            max_tokens=data.get("max_tokens", 1024),
            temperature=data.get("temperature", 0.4),
        )

        if result["metadata"]["status"] == "error":
            return JSONResponse(status_code=500, content={"error": result["metadata"]["error_message"]})

        return {
            "success": True,
            "result": result["result"],
            "metadata": result["metadata"],
        }

    except Exception as e:
        logger.error(f"Vision analyze error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/models/open-folder")
async def open_models_folder():
    """Open the models directory in Windows File Explorer"""
    try:
        models_path = str(MODELS_DIR.resolve())
        
        if not MODELS_DIR.exists():
            MODELS_DIR.mkdir(parents=True, exist_ok=True)
        
        if os.name == 'nt':  # Windows
            subprocess.run(['explorer', models_path], check=True)
            return {
                "success": True,
                "message": "Models folder opened",
                "path": models_path
            }
        else:
            return {
                "success": False,
                "message": "Opening folder is only supported on Windows",
                "path": models_path
            }
    except Exception as e:
        logger.error(f"Error opening models folder: {e}")
        raise HTTPException(status_code=500, detail=f"Error opening models folder: {str(e)}")


# ============================================================================
# HARDWARE-AWARE MODEL RECOMMENDATIONS
# ============================================================================

@app.get("/api/recommendations")
async def get_model_recommendations():
    """Return hardware-aware model recommendations based on system specs."""
    try:
        from metrics_service import _get_system_specs
        from model_recommender import get_hardware_profile, recommend_models

        specs = _get_system_specs()
        hw_profile = get_hardware_profile(specs)

        available_llms = scan_models_directory()
        available_diffusion = scan_diffusion_models()
        available_ocr = scan_ocr_models()

        result = recommend_models(hw_profile, available_llms, available_diffusion, available_ocr, specs)
        return result
    except Exception as e:
        logger.error(f"Error generating recommendations: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating recommendations: {str(e)}")


# ============================================================================
# DIFFUSION MODEL MANAGEMENT ENDPOINTS
# ============================================================================

@app.get("/api/diffusion-models")
async def get_available_diffusion_models():
    """Get list of available diffusion models in the image_generation directory"""
    try:
        models = scan_diffusion_models()
        return {
            "success": True,
            "models": models,
            "models_directory": str(IMAGE_MODELS_DIR),
            "count": len(models)
        }
    except Exception as e:
        logger.error(f"Error retrieving diffusion models: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving diffusion models: {str(e)}")


@app.post("/api/diffusion-models/select")
async def select_diffusion_model(request: Request):
    """Set the active diffusion model for image generation"""
    try:
        data = await request.json()
        model_name = data.get('modelName', '')

        # Allow empty string to disable diffusion
        if not model_name:
            set_selected_diffusion_model('')
            return JSONResponse(content={"status": "ok", "message": "Diffusion model disabled"})

        model_path = get_image_model_path(model_name)
        if not model_path.exists():
            return JSONResponse(
                status_code=404,
                content={"error": f"Diffusion model not found: {model_name}"}
            )

        if set_selected_diffusion_model(model_name):
            # Reset the image service singleton so it picks up the new model on next use
            try:
                from image_service import reset_image_service
                reset_image_service(model_name)
                logger.info("Image service reset for new diffusion model")
            except Exception as e:
                logger.warning(f"Could not reset image service: {e}")

            return JSONResponse(content={
                "success": True,
                "message": f"Diffusion model set to {model_name}",
                "selectedModel": model_name
            })
        else:
            return JSONResponse(
                status_code=500,
                content={"error": "Failed to save diffusion model selection"}
            )

    except Exception as e:
        logger.error(f"Error selecting diffusion model: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.get("/api/diffusion-models/active")
async def get_active_diffusion_model():
    """Get the currently active diffusion model"""
    try:
        selected_model = get_selected_diffusion_model()
        model_path = get_diffusion_model_path()
        model_info = get_image_model_info(selected_model)

        return JSONResponse(content={
            "modelName": selected_model,
            "modelPath": model_path,
            "exists": Path(model_path).exists(),
            "steps": model_info.get("steps", 4),
            "guidance": model_info.get("guidance", 0.0),
            "supports_negative_prompt": model_info.get("supports_negative_prompt", True),
            "supports_img2img": model_info.get("supports_img2img", True),
        })
    except Exception as e:
        logger.error(f"Error getting active diffusion model: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/diffusion-models/open-folder")
async def open_diffusion_models_folder():
    """Open the image generation models directory in Windows File Explorer"""
    try:
        models_path = str(IMAGE_MODELS_DIR.resolve())

        if not IMAGE_MODELS_DIR.exists():
            IMAGE_MODELS_DIR.mkdir(parents=True, exist_ok=True)

        if os.name == 'nt':
            subprocess.run(['explorer', models_path], check=True)
            return {
                "success": True,
                "message": "Diffusion models folder opened",
                "path": models_path
            }
        else:
            return {
                "success": False,
                "message": "Opening folder is only supported on Windows",
                "path": models_path
            }
    except Exception as e:
        logger.error(f"Error opening diffusion models folder: {e}")
        raise HTTPException(status_code=500, detail=f"Error opening diffusion models folder: {str(e)}")


@app.get("/api/health")
async def health():
    return {
        "status": "healthy",
        "model_found": os.path.exists(get_model_path()),
        "llama_cli_found": os.path.exists(LLAMA_CLI_PATH),
        "chat_memory_db": os.path.exists(get_chat_memory().db_path)
    }

@app.post("/api/restart")
async def restart():
    """Restart the backend by cleaning up processes and triggering uvicorn reload"""
    cleanup_all_processes()

    # Schedule a file touch to trigger uvicorn --reload after response is sent
    async def _trigger_reload():
        await asyncio.sleep(0.5)
        Path(__file__).touch()

    asyncio.create_task(_trigger_reload())

    return {"status": "restarting"}

@app.post("/api/shutdown")
async def shutdown():
    """Gracefully shutdown the backend and cleanup all processes"""
    cleanup_all_processes()

    # Schedule server shutdown after sending response
    import asyncio
    asyncio.create_task(_shutdown_server())

    return {"status": "shutting down"}

async def _shutdown_server():
    """Shutdown the server after a brief delay"""
    await asyncio.sleep(0.5)  # Allow response to be sent
    os._exit(0)  # Force terminate the process

  
# =========================
# Export Endpoint
# =========================

from fastapi import Body
from fastapi.responses import StreamingResponse
# NOTE: This file must be run as a module/package (e.g., `python -m backend.main` or `uvicorn backend.main:app`)
from export_utils import EXPORT_FORMATTERS

EXPORT_TYPE_FORMATS = {
    "plan": {"pdf", "docx"},
    "quiz": {"pdf", "docx"},
    "rubric": {"pdf", "docx"},
    "kindergarten": {"pdf", "docx"},
    "multigrade": {"pdf", "docx"},
    "cross-curricular": {"pdf", "docx"},
    "curriculum": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "chat_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "lesson_plan_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "quiz_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "rubric_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "kindergarten_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "multigrade_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "cross_curricular_history": {"pdf", "docx", "csv", "json", "md", "markdown"},
    "worksheet": {"pdf", "docx"},
    "report-card": {"pdf"},
}

CONTENT_TYPES = {
    "pdf": "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "csv": "text/csv",
    "json": "application/json",
    "md": "text/markdown",
    "markdown": "text/markdown",
}

# ══════════════════════════════════════════════════════════════
# File Explorer - Browse Endpoints (dev mode fallback for Electron APIs)
# ══════════════════════════════════════════════════════════════

ALLOWED_EXTENSIONS = {'.docx', '.pptx', '.pdf', '.txt', '.md', '.xlsx', '.csv',
                      '.png', '.jpg', '.jpeg', '.gif', '.webp', '.doc', '.ppt', '.xls'}

# Persist allowed folders in a JSON file next to main.py
_FOLDER_CONFIG_PATH = os.path.join(BASE_DIR, 'file-explorer-config.json')

def _load_allowed_folders() -> list:
    import json
    if os.path.exists(_FOLDER_CONFIG_PATH):
        try:
            with open(_FOLDER_CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f).get('allowedFolders', [])
        except Exception:
            pass
    # Defaults: user's Desktop and Downloads
    home = os.path.expanduser('~')
    return [os.path.join(home, 'Downloads'), os.path.join(home, 'Desktop')]

def _save_allowed_folders(folders: list):
    import json
    with open(_FOLDER_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump({'allowedFolders': folders}, f)

@app.get("/api/file-explorer/allowed-folders")
async def get_allowed_folders():
    return JSONResponse(content={"folders": _load_allowed_folders()})

@app.post("/api/file-explorer/allowed-folders")
async def save_allowed_folders(body: dict = Body(...)):
    folders = body.get('folders', [])
    _save_allowed_folders(folders)
    return JSONResponse(content={"ok": True})

@app.get("/api/file-explorer/browse")
async def browse_folder(folderPath: str):
    """List contents of a folder — returns files (filtered by allowed extensions) and subdirectories."""
    if not os.path.isdir(folderPath):
        return JSONResponse(content={"items": [], "error": "Not a directory"})
    items = []
    try:
        for entry in os.scandir(folderPath):
            try:
                stat = entry.stat()
                ext = os.path.splitext(entry.name)[1].lower()
                if entry.is_dir(follow_symlinks=False):
                    items.append({
                        "name": entry.name,
                        "path": entry.path.replace('\\', '/'),
                        "isDirectory": True,
                        "size": 0,
                        "modifiedTime": str(stat.st_mtime),
                        "extension": ""
                    })
                elif ext in ALLOWED_EXTENSIONS:
                    items.append({
                        "name": entry.name,
                        "path": entry.path.replace('\\', '/'),
                        "isDirectory": False,
                        "size": stat.st_size,
                        "modifiedTime": str(stat.st_mtime),
                        "extension": ext
                    })
            except (PermissionError, OSError):
                continue
    except (PermissionError, OSError) as e:
        return JSONResponse(content={"items": [], "error": str(e)})
    # Sort: folders first, then files, alphabetically
    items.sort(key=lambda x: (not x['isDirectory'], x['name'].lower()))
    return JSONResponse(content={"items": items})

@app.get("/api/file-explorer/search")
async def search_files(query: str):
    """Search for files across allowed folders."""
    folders = _load_allowed_folders()
    results = []
    query_lower = query.lower()
    for folder in folders:
        if not os.path.isdir(folder):
            continue
        for root, dirs, files in os.walk(folder):
            for fname in files:
                ext = os.path.splitext(fname)[1].lower()
                if ext not in ALLOWED_EXTENSIONS:
                    continue
                if query_lower in fname.lower():
                    fpath = os.path.join(root, fname)
                    try:
                        stat = os.stat(fpath)
                        results.append({
                            "name": fname,
                            "path": fpath.replace('\\', '/'),
                            "isDirectory": False,
                            "size": stat.st_size,
                            "modifiedTime": str(stat.st_mtime),
                            "extension": ext
                        })
                    except (PermissionError, OSError):
                        continue
            if len(results) >= 200:
                break
        if len(results) >= 200:
            break
    return JSONResponse(content={"items": results})

@app.get("/api/file-explorer/preview-by-path")
async def preview_by_path(filePath: str):
    """Parse a local file by path and return preview content."""
    if not os.path.isfile(filePath):
        return JSONResponse(content={"error": "File not found"}, status_code=404)
    max_size = 50 * 1024 * 1024
    if os.path.getsize(filePath) > max_size:
        return JSONResponse(content={"error": "File too large (max 50MB)"}, status_code=400)
    try:
        from file_parser import parse_file
        with open(filePath, 'rb') as f:
            content = f.read()
        result = parse_file(content, os.path.basename(filePath))
        return JSONResponse(content=result)
    except Exception as e:
        logging.error(f"Error previewing file by path: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.get("/api/file-explorer/read-file")
async def read_file_content(filePath: str):
    """Read a file's raw bytes, return as base64."""
    if not os.path.isfile(filePath):
        return JSONResponse(content={"error": "File not found"})
    max_size = 50 * 1024 * 1024
    if os.path.getsize(filePath) > max_size:
        return JSONResponse(content={"error": "File too large (max 50MB)"})
    try:
        with open(filePath, 'rb') as f:
            data = f.read()
        return JSONResponse(content={"base64": base64.b64encode(data).decode('ascii')})
    except Exception as e:
        return JSONResponse(content={"error": str(e)})

# ══════════════════════════════════════════════════════════════
# File Explorer Endpoints
# ══════════════════════════════════════════════════════════════

@app.post("/api/file-explorer/parse")
async def file_explorer_parse(file: UploadFile = File(...)):
    """Parse a file and extract its full text content."""
    try:
        from file_parser import parse_file
        content = await file.read()
        result = parse_file(content, file.filename or "unknown")
        return JSONResponse(content=result)
    except Exception as e:
        logging.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/file-explorer/preview")
async def file_explorer_preview(file: UploadFile = File(...)):
    """Get a truncated preview of file content (first ~500 words)."""
    try:
        from file_parser import get_preview
        content = await file.read()
        result = get_preview(content, file.filename or "unknown", max_words=500)
        return JSONResponse(content=result)
    except Exception as e:
        logging.error(f"Error previewing file: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/file-explorer/organize")
async def file_explorer_organize(body: dict = Body(...)):
    """
    Generate an AI organization plan for files in a folder.
    Accepts: { files: [{name, extension, size, modifiedTime}], folder_name: str, instruction?: str }
    Returns: { action, description, folders_to_create, moves: [{file, from, to}] }
    """
    try:
        files = body.get("files", [])
        folder_name = body.get("folder_name", "folder")
        instruction = body.get("instruction", "Organize these files into logical subfolders")

        if not files:
            return JSONResponse(content={"error": "No files provided"}, status_code=400)

        # Build a file listing for the LLM
        file_listing = "\n".join(
            f"- {f['name']} ({f.get('extension', '')}), {f.get('size', 0)} bytes, modified {f.get('modifiedTime', 'unknown')}"
            for f in files[:200]  # Cap at 200 files
        )

        prompt = f"""You are a file organization assistant. A teacher wants to organize their "{folder_name}" folder.

Their instruction: {instruction}

Here are the files currently in the folder:
{file_listing}

Respond with ONLY a valid JSON object (no markdown, no explanation) in this exact format:
{{
  "action": "organize",
  "description": "Brief description of the organization plan",
  "folders_to_create": ["FolderName1", "FolderName2"],
  "moves": [
    {{"file": "exact filename.ext", "from": "{folder_name}", "to": "FolderName1"}},
    {{"file": "another file.ext", "from": "{folder_name}", "to": "FolderName2"}}
  ]
}}

Rules:
- Group files by type/purpose (e.g., Documents, Spreadsheets, Images, Presentations)
- Use clear, descriptive folder names
- Every file must appear in the moves array
- The "from" field is always "{folder_name}"
- The "to" field is one of the folders_to_create
- Do NOT create folders with only 1 file unless the instruction specifically asks for it
- If files don't fit any category, put them in a "Miscellaneous" folder
"""

        from inference_factory import get_inference_instance
        inference = get_inference_instance()
        llm_result = await inference.generate(
            tool_name='file_organize',
            input_data=f"Organize {len(files)} files in {folder_name}",
            prompt_template=prompt,
            max_tokens=2000,
            temperature=0.3,
        )

        raw = llm_result.get('result') or ''

        import re as _re
        json_match = _re.search(r'\{[\s\S]*\}', raw)
        if not json_match:
            return JSONResponse(
                content={"error": "AI could not generate a valid organization plan. Please try again."},
                status_code=422
            )

        plan = json.loads(json_match.group())

        # Validate plan structure
        if not isinstance(plan.get("folders_to_create"), list) or not isinstance(plan.get("moves"), list):
            return JSONResponse(
                content={"error": "AI returned an invalid plan structure. Please try again."},
                status_code=422
            )

        return JSONResponse(content=plan)

    except json.JSONDecodeError:
        return JSONResponse(
            content={"error": "AI response was not valid JSON. Please try again."},
            status_code=422
        )
    except Exception as e:
        logging.error(f"Error generating organization plan: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/file-explorer/execute-organize")
async def execute_organize(request: Request):
    """Execute a file organization plan — create folders and move files."""
    import shutil
    try:
        body = await request.json()
        folder_path = body.get("folderPath", "")
        folders_to_create = body.get("folders_to_create", [])
        moves = body.get("moves", [])

        if not folder_path or not Path(folder_path).is_dir():
            return JSONResponse(status_code=400, content={"error": "Invalid folder path"})

        base = Path(folder_path).resolve()
        results = []
        undo_log = []

        # Create folders
        for folder_name in folders_to_create:
            target = base / folder_name
            # Safety: ensure target stays within base
            if not str(target.resolve()).startswith(str(base)):
                continue
            target.mkdir(parents=True, exist_ok=True)

        # Move files
        for move in moves:
            file_name = move.get("file", "")
            dest_folder = move.get("to", "")
            if not file_name or not dest_folder:
                results.append({"file": file_name, "success": False, "error": "Missing file or destination"})
                continue

            source = base / file_name
            dest_dir = base / dest_folder
            dest = dest_dir / file_name

            # Safety checks
            if not str(source.resolve()).startswith(str(base)):
                results.append({"file": file_name, "success": False, "error": "Path escape blocked"})
                continue
            if not str(dest.resolve()).startswith(str(base)):
                results.append({"file": file_name, "success": False, "error": "Path escape blocked"})
                continue
            if not source.exists():
                results.append({"file": file_name, "success": False, "error": "File not found"})
                continue

            try:
                dest_dir.mkdir(parents=True, exist_ok=True)
                shutil.move(str(source), str(dest))
                results.append({"file": file_name, "success": True})
                undo_log.append({"source": str(dest), "original": str(source)})
            except Exception as e:
                results.append({"file": file_name, "success": False, "error": str(e)})

        success_count = sum(1 for r in results if r["success"])
        fail_count = sum(1 for r in results if not r["success"])

        return JSONResponse(content={
            "results": results,
            "undoLog": undo_log,
            "summary": {"success": success_count, "failed": fail_count},
        })

    except Exception as e:
        logger.error(f"Error executing organize plan: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/file-explorer/undo-organize")
async def undo_organize(request: Request):
    """Undo a file organization — move files back to original locations."""
    import shutil
    try:
        body = await request.json()
        undo_log = body.get("undoLog", [])
        results = []

        for entry in undo_log:
            source = entry.get("source", "")
            original = entry.get("original", "")
            if not source or not original:
                continue
            try:
                Path(original).parent.mkdir(parents=True, exist_ok=True)
                shutil.move(source, original)
                results.append({"file": Path(original).name, "success": True})
            except Exception as e:
                results.append({"file": Path(original).name, "success": False, "error": str(e)})

        return JSONResponse(content={"results": results})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/export")
async def export_data(
    data_type: str = Body(..., embed=True),
    format: str = Body(..., embed=True),
    data: dict = Body(..., embed=True),
    title: str = Body("Export", embed=True)
):
    """
    Export endpoint for various data types and formats.
    """
    logger.info(f"Export request: data_type={data_type}, format={format}, title={title}")

    # Validate data_type
    if data_type not in EXPORT_TYPE_FORMATS:
        logger.error(f"Invalid data_type: {data_type}")
        return JSONResponse(
            status_code=400,
            content={"error": f"Invalid data_type '{data_type}'. Allowed: {list(EXPORT_TYPE_FORMATS.keys())}"}
        )
    # Validate format
    allowed_formats = EXPORT_TYPE_FORMATS[data_type]
    if format not in allowed_formats:
        logger.error(f"Format '{format}' not allowed for data_type '{data_type}'")
        return JSONResponse(
            status_code=400,
            content={"error": f"Format '{format}' not allowed for data_type '{data_type}'. Allowed: {sorted(allowed_formats)}"}
        )
    # Validate formatter
    formatter = EXPORT_FORMATTERS.get(format)
    if not formatter:
        logger.error(f"Export format '{format}' is not supported")
        return JSONResponse(
            status_code=400,
            content={"error": f"Export format '{format}' is not supported."}
        )
    # Export
    try:
        logger.info(f"Starting export with formatter: {format}")
        exported_bytes = formatter(data, title=title)
        logger.info(f"Export successful, size: {len(exported_bytes)} bytes")
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"error": f"Export failed: {str(e)}"}
        )
    # Prepare response
    ext = "md" if format == "markdown" else format
    filename = f"{data_type}_{title.replace(' ', '_')}.{ext}"
    content_type = CONTENT_TYPES.get(format, "application/octet-stream")
    logger.info(f"Sending response: filename={filename}, content_type={content_type}")
    return StreamingResponse(
        io.BytesIO(exported_bytes),
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

    


# =========================
# IMAGE GENERATION ENDPOINTS
# =========================
 
@app.post("/api/generate-image-prompt")
async def generate_image_prompt(request: Request):
    """
    Generate an optimized image prompt using LLaMA model
    """
    try:
        data = await request.json()
        context = data.get('context', {})
        
        # Extract context information
        subject = context.get('subject', '')
        grade = context.get('grade', '')
        topic = context.get('topic', '')
        question_type = context.get('questionType', '')
        additional = context.get('additionalContext', '')
        
        # Build LLaMA prompt for image generation
        llama_prompt = f"""You are an expert at creating image prompts for educational materials.

Generate a detailed, specific image prompt for SDXL image generation with these requirements:

CONTEXT:
- Subject: {subject}
- Grade Level: {grade}
- Topic: {topic}
- Question Type: {question_type}
- Additional Context: {additional}

REQUIREMENTS:
1. The image must be age-appropriate for Grade {grade}
2. Style: Colorful, cartoon/illustration style, educational
3. Content: Simple, clear, focused on the main concept
4. Avoid: Text, words, numbers, multiple subjects, complex scenes
5. Format: Single clear subject, suitable for worksheet inclusion

Generate ONLY the image prompt (no explanation, no preamble). The prompt should be 1-2 sentences, specific and detailed.

PROMPT:"""
        
        # Get inference instance
        from inference_factory import get_inference_instance
        inference = get_inference_instance()
        
        if not inference.is_loaded:
            return JSONResponse(
                status_code=503,
                content={"error": "AI model not loaded"}
            )
        
        # Call LLaMA using the correct API (matches your llama_inference.py)
        result = await inference.generate(
            tool_name="image_prompt_generator",
            input_data=llama_prompt,
            prompt_template=None,  # Use input_data as-is
            max_tokens=150,
            temperature=0.7,
            top_p=0.9
        )
        
        # Extract generated text from result
        if result.get("metadata", {}).get("status") != "success":
            error_msg = result.get("metadata", {}).get("error_message", "Generation failed")
            return JSONResponse(
                status_code=500,
                content={"error": error_msg}
            )
        
        generated_prompt = result.get("result", "").strip()
        
        if not generated_prompt:
            return JSONResponse(
                status_code=500,
                content={"error": "No prompt generated"}
            )
        
        # Generate negative prompt based on educational context
        negative_prompt = "text, words, letters, numbers, multiple subjects, people, faces, complex scene, dark, scary, violent, inappropriate, blurry, distorted"
        
        if grade and grade.lower() in ['k', '1', '2', '3']:
            negative_prompt += ", realistic, photographic, detailed background"
        
        logger.info(f"Generated image prompt: {generated_prompt[:100]}...")
        
        return JSONResponse(content={
            "success": True,
            "prompt": generated_prompt,
            "negativePrompt": negative_prompt
        })
        
    except Exception as e:
        logger.error(f"Error generating image prompt: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )

@app.post("/api/generate-comic-prompts")
async def generate_comic_prompts(request: Request):
    """
    Generate scene prompts for a comic page using LLaMA.
    Takes a short description and number of panels, returns per-panel prompts.
    """
    try:
        data = await request.json()
        description = data.get('description', '')
        num_panels = data.get('numPanels', 4)
        style = data.get('style', 'cartoon_3d')

        if not description:
            return JSONResponse(
                status_code=400,
                content={"error": "Description is required"}
            )

        llama_prompt = f"""<|begin_of_text|><|start_header_id|>system<|end_header_id|>

You are an expert comic book writer. Given a short story description, break it into exactly {num_panels} sequential comic panel scenes. Each scene should be a vivid, visual image prompt suitable for AI image generation. Focus on action, composition, and visual storytelling. Do NOT include dialogue or text in the prompts.

IMPORTANT: Return ONLY a JSON array of strings, one per panel. No explanation, no markdown, no extra text. Example format:
["scene 1 prompt", "scene 2 prompt", "scene 3 prompt", "scene 4 prompt"]<|eot_id|><|start_header_id|>user<|end_header_id|>

Story description: {description}
Number of panels: {num_panels}

Generate {num_panels} sequential comic panel image prompts as a JSON array:<|eot_id|><|start_header_id|>assistant<|end_header_id|>

["""

        settings = {
            "model_path": get_model_path(),
            "n_ctx": MODEL_N_CTX,
            "max_tokens": 600,
            "temperature": 0.7,
            "tool_name": "comic_prompts",
            "prompt_template": llama_prompt,
        }
        future = submit_task(run_llama_inference, description, settings)
        result = await asyncio.wrap_future(future)

        if result.get("metadata", {}).get("status") != "success":
            error_msg = result.get("metadata", {}).get("error_message", "Generation failed")
            return JSONResponse(
                status_code=500,
                content={"error": error_msg}
            )

        generated_text = result.get("result", "").strip()

        # Try to parse the JSON array from the response
        # The model might return it with or without the leading [
        text_to_parse = generated_text
        if not text_to_parse.startswith("["):
            text_to_parse = "[" + text_to_parse

        # Find the closing bracket
        bracket_idx = text_to_parse.rfind("]")
        if bracket_idx != -1:
            text_to_parse = text_to_parse[:bracket_idx + 1]
        else:
            text_to_parse = text_to_parse + "]"

        try:
            prompts = json.loads(text_to_parse)
            if not isinstance(prompts, list):
                raise ValueError("Not a list")
            # Ensure we have exactly num_panels prompts
            prompts = [str(p) for p in prompts[:num_panels]]
            while len(prompts) < num_panels:
                prompts.append(f"Scene {len(prompts) + 1} of: {description}")
        except (json.JSONDecodeError, ValueError):
            # Fallback: split by newlines or create generic prompts
            logger.warning(f"Failed to parse comic prompts JSON, using fallback. Raw: {generated_text[:200]}")
            prompts = []
            for i in range(num_panels):
                prompts.append(f"Scene {i + 1} of a comic story: {description}, panel {i + 1} of {num_panels}")

        logger.info(f"Generated {len(prompts)} comic panel prompts")

        return JSONResponse(content={
            "success": True,
            "prompts": prompts
        })

    except Exception as e:
        logger.error(f"Error generating comic prompts: {e}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


# ============================================================================
# TEXT-TO-SPEECH (Piper TTS — fully offline)
# ============================================================================

@app.post("/api/tts")
async def text_to_speech(request: Request):
    """
    Synthesize text to speech using Piper TTS (offline).
    Returns WAV audio bytes.
    """
    try:
        data = await request.json()
        text = data.get("text", "").strip()
        speed = float(data.get("speed", 1.0))
        voice = data.get("voice", "lessac")  # 'lessac' | 'ryan' | 'amy'

        if not text:
            return JSONResponse(status_code=400, content={"error": "Text is required"})

        from tts_service import get_tts_service
        tts = get_tts_service()

        # Run synthesis in thread pool to avoid blocking the event loop
        loop = asyncio.get_event_loop()
        wav_bytes = await loop.run_in_executor(
            None, lambda: tts.synthesize(text, speed=speed, voice=voice)
        )

        return Response(content=wav_bytes, media_type="audio/wav")

    except Exception as e:
        logger.error(f"TTS error: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/tts/preload")
async def preload_tts():
    """Preload the TTS voice model in the background.

    This triggers lazy loading of the voice model when a user opens
    a tab with speech options, so audio is ready by the time they
    click Read Aloud.
    """
    try:
        from tts_service import get_tts_service
        tts = get_tts_service()
        if tts.is_loaded:
            return {"status": "loaded"}
        logger.info("Preloading TTS default voice model...")
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, lambda: tts.synthesize("ready", voice="lessac"))
        logger.info("TTS default voice preloaded successfully")
        return {"status": "loaded"}
    except Exception as e:
        logger.error(f"Error preloading TTS model: {e}")
        return JSONResponse(status_code=500, content={"status": "error", "error": str(e)})


@app.get("/api/tts/status")
async def tts_status():
    """Check if TTS voice model is loaded and ready."""
    try:
        from tts_service import get_tts_service
        tts = get_tts_service()
        return JSONResponse(content=tts.get_voice_info())
    except Exception as e:
        return JSONResponse(content={"loaded": False, "error": str(e)})


@app.get("/api/tts/voices")
async def tts_voices():
    """Return available storybook voices with download status."""
    try:
        from tts_service import get_tts_service, VOICE_REGISTRY, is_voice_available
        tts = get_tts_service()
        voices = []
        display_names = {
            "lessac": "Lessac (Female)",
            "ryan":   "Ryan (Male)",
            "amy":    "Amy (Female)",
        }
        for key in VOICE_REGISTRY:
            voices.append({
                "key": key,
                "displayName": display_names.get(key, key),
                "available": is_voice_available(key),
            })
        return JSONResponse(content={"voices": voices})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/generate-image")
async def generate_image(request: Request):
    """
    Generate image using SDXL-Turbo
    
    Request body:
    {
        "prompt": "image description",
        "negativePrompt": "things to avoid",
        "width": 1024,
        "height": 512,
        "numInferenceSteps": 2
    }
    
    Returns:
        PNG image as bytes
    """
    try:
        data = await request.json()
        
        prompt = data.get('prompt', '')
        negative_prompt = data.get('negativePrompt', 'deformed, distorted, blurry, extra fingers, mutated hands, poorly drawn hands, bad anatomy, extra limbs, fused fingers, too many fingers, ugly, low quality, worst quality')
        width = data.get('width', 1024)
        height = data.get('height', 512)
        num_steps = data.get('numInferenceSteps', None)
        
        if not prompt:
            return JSONResponse(
                status_code=400,
                content={"error": "Prompt is required"}
            )
        
        # Validate dimensions
        if width < 256 or width > 2048 or height < 256 or height > 2048:
            return JSONResponse(
                status_code=400,
                content={"error": "Width and height must be between 256 and 2048"}
            )
        
        # Get image service
        image_service = get_image_service()
        
        # Generate image
        image_bytes = image_service.generate_image(
            prompt=prompt,
            negative_prompt=negative_prompt,
            width=width,
            height=height,
            num_inference_steps=num_steps
        )
        
        if image_bytes is None:
            return JSONResponse(
                status_code=500,
                content={"error": "Image generation failed"}
            )
        
        # Return image as PNG
        return Response(
            content=image_bytes,
            media_type="image/png"
        )
        
    except Exception as e:
        logger.error(f"Error generating image: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/generate-image-base64")
async def generate_image_base64(request: Request):
    """
    Generate image and return as base64 (easier for frontend embedding)

    Same request body as /api/generate-image

    Returns:
    {
        "success": true,
        "imageData": "data:image/png;base64,..."
    }
    """
    try:
        data = await request.json()

        prompt = data.get('prompt', '')
        negative_prompt = data.get('negativePrompt', 'deformed, distorted, blurry, extra fingers, mutated hands, poorly drawn hands, bad anatomy, extra limbs, fused fingers, too many fingers, ugly, low quality, worst quality')
        width = data.get('width', 1024)
        height = data.get('height', 512)
        num_steps = data.get('numInferenceSteps', None)
        init_image_b64 = data.get('initImage')
        strength = data.get('strength', 0.5)

        if not prompt:
            return JSONResponse(
                status_code=400,
                content={"error": "Prompt is required"}
            )

        # Decode init_image if provided (for img2img)
        init_image_bytes = None
        if init_image_b64:
            if init_image_b64.startswith('data:'):
                init_image_b64 = init_image_b64.split(',')[1]
            try:
                init_image_bytes = base64.b64decode(init_image_b64)
            except Exception as e:
                logger.error(f"Failed to decode init image: {e}")
                return JSONResponse(
                    status_code=400,
                    content={"error": "Invalid init image data"}
                )

        # Get image service
        image_service = get_image_service()

        # Generate image (text-to-image or image-to-image)
        image_bytes = image_service.generate_image(
            prompt=prompt,
            negative_prompt=negative_prompt,
            width=width,
            height=height,
            num_inference_steps=num_steps,
            init_image=init_image_bytes,
            strength=strength
        )

        if image_bytes is None:
            return JSONResponse(
                status_code=500,
                content={"error": "Image generation failed"}
            )

        # Convert to base64
        image_b64 = base64.b64encode(image_bytes).decode('utf-8')
        data_uri = f"data:image/png;base64,{image_b64}"

        return JSONResponse(content={
            "success": True,
            "imageData": data_uri
        })

    except Exception as e:
        logger.error(f"Error generating image: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/generate-batch-images-base64")
async def generate_batch_images_base64(request: Request):
    """
    Generate multiple images in batch and return as base64

    Request body:
    {
        "prompt": "image description",
        "negativePrompt": "things to avoid",
        "width": 1024,
        "height": 512,
        "numInferenceSteps": 2,
        "numImages": 4
    }

    Returns:
    {
        "success": true,
        "images": [
            {
                "imageData": "data:image/png;base64,...",
                "seed": 12345
            },
            ...
        ]
    }
    """
    try:
        data = await request.json()

        prompt = data.get('prompt', '')
        negative_prompt = data.get('negativePrompt', 'deformed, distorted, blurry, extra fingers, mutated hands, poorly drawn hands, bad anatomy, extra limbs, fused fingers, too many fingers, ugly, low quality, worst quality')
        width = data.get('width', 1024)
        height = data.get('height', 512)
        num_steps = data.get('numInferenceSteps', None)
        num_images = data.get('numImages', 1)

        if not prompt:
            return JSONResponse(
                status_code=400,
                content={"error": "Prompt is required"}
            )

        if num_images < 1 or num_images > 10:
            return JSONResponse(
                status_code=400,
                content={"error": "numImages must be between 1 and 10"}
            )

        # Get image service
        image_service = get_image_service()

        # Generate batch images
        batch_results = image_service.generate_batch_images(
            prompt=prompt,
            negative_prompt=negative_prompt,
            width=width,
            height=height,
            num_inference_steps=num_steps,
            num_images=num_images
        )

        if not batch_results:
            return JSONResponse(
                status_code=500,
                content={"error": "Batch image generation failed"}
            )

        # Convert to base64
        images = []
        for result in batch_results:
            image_b64 = base64.b64encode(result['image_data']).decode('utf-8')
            data_uri = f"data:image/png;base64,{image_b64}"
            images.append({
                "imageData": data_uri,
                "seed": result['seed']
            })

        return JSONResponse(content={
            "success": True,
            "images": images
        })

    except Exception as e:
        logger.error(f"Error generating batch images: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/generate-image-from-seed")
async def generate_image_from_seed(request: Request):
    """
    Generate image using a specific seed and init image (for AI Edit functionality)

    Request body:
    {
        "prompt": "modified image description",
        "negativePrompt": "things to avoid",
        "width": 1024,
        "height": 512,
        "numInferenceSteps": 2,
        "seed": 12345,
        "initImage": "data:image/png;base64,..."  // optional, for img2img
    }

    Returns:
    {
        "success": true,
        "imageData": "data:image/png;base64,...",
        "seed": 12345
    }
    """
    try:
        data = await request.json()

        prompt = data.get('prompt', '')
        negative_prompt = data.get('negativePrompt', 'deformed, distorted, blurry, extra fingers, mutated hands, poorly drawn hands, bad anatomy, extra limbs, fused fingers, too many fingers, ugly, low quality, worst quality')
        width = data.get('width', 1024)
        height = data.get('height', 512)
        num_steps = data.get('numInferenceSteps', None)
        seed = data.get('seed')
        init_image_b64 = data.get('initImage')
        strength = data.get('strength', 0.5)

        if not prompt:
            return JSONResponse(
                status_code=400,
                content={"error": "Prompt is required"}
            )

        if seed is None:
            return JSONResponse(
                status_code=400,
                content={"error": "Seed is required"}
            )

        # Decode init_image if provided
        init_image_bytes = None
        if init_image_b64:
            if init_image_b64.startswith('data:'):
                init_image_b64 = init_image_b64.split(',')[1]
            try:
                init_image_bytes = base64.b64decode(init_image_b64)
            except Exception as e:
                logger.error(f"Failed to decode init image: {e}")
                return JSONResponse(
                    status_code=400,
                    content={"error": "Invalid init image data"}
                )

        # Get image service
        image_service = get_image_service()

        # Generate image with seed and optional init image
        image_bytes = image_service.generate_image(
            prompt=prompt,
            negative_prompt=negative_prompt,
            width=width,
            height=height,
            num_inference_steps=num_steps,
            seed=seed,
            init_image=init_image_bytes,
            strength=strength
        )

        if image_bytes is None:
            return JSONResponse(
                status_code=500,
                content={"error": "Image generation failed"}
            )

        # Convert to base64
        image_b64 = base64.b64encode(image_bytes).decode('utf-8')
        data_uri = f"data:image/png;base64,{image_b64}"

        return JSONResponse(content={
            "success": True,
            "imageData": data_uri,
            "seed": seed
        })

    except Exception as e:
        logger.error(f"Error generating image from seed: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/inpaint")
async def inpaint_image(
    image: UploadFile = File(...),
    mask: UploadFile = File(...),
    seed: Optional[int] = None
):
    """
    Remove objects from image using IOPaint (LaMa model)
    
    Form data:
    - image: Original image file
    - mask: Mask image file (white = remove, black = keep)
    - seed: Optional random seed
    
    Returns:
        Inpainted image as PNG
    """
    if not LAMA_MODEL_PATH.exists():
        return JSONResponse(
            status_code=503,
            content={"error": "Object Remover is unavailable: big-lama.pt not found in models/image_generation/lama/"}
        )

    try:
        # Read uploaded files
        image_data = await image.read()
        mask_data = await mask.read()

        if not image_data or not mask_data:
            return JSONResponse(
                status_code=400,
                content={"error": "Both image and mask are required"}
            )

        # Get image service
        image_service = get_image_service()
        
        # Perform inpainting
        result_bytes = image_service.inpaint_image(
            image_data=image_data,
            mask_data=mask_data,
            seed=seed
        )
        
        if result_bytes is None:
            return JSONResponse(
                status_code=500,
                content={"error": "Inpainting failed"}
            )
        
        # Return result image
        return Response(
            content=result_bytes,
            media_type="image/png"
        )
        
    except Exception as e:
        logger.error(f"Error in inpainting: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/inpaint-base64")
async def inpaint_image_base64(request: Request):
    """
    Inpaint with base64 input/output (easier for frontend)

    Request body:
    {
        "image": "data:image/png;base64,...",
        "mask": "data:image/png;base64,...",
        "seed": 12345 (optional)
    }

    Returns:
    {
        "success": true,
        "imageData": "data:image/png;base64,..."
    }
    """
    if not LAMA_MODEL_PATH.exists():
        return JSONResponse(
            status_code=503,
            content={"success": False, "error": "Object Remover is unavailable: big-lama.pt not found in models/image_generation/lama/"}
        )

    try:
        logger.info("=== INPAINT-BASE64 REQUEST RECEIVED ===")
        data = await request.json()
        logger.info(f"Request data keys: {list(data.keys())}")

        image_b64 = data.get('image', '')
        mask_b64 = data.get('mask', '')
        seed = data.get('seed')

        logger.info(f"Image b64 length: {len(image_b64) if image_b64 else 0}")
        logger.info(f"Mask b64 length: {len(mask_b64) if mask_b64 else 0}")
        logger.info(f"Seed: {seed}")

        if not image_b64 or not mask_b64:
            logger.error("Missing image or mask data")
            return JSONResponse(
                status_code=400,
                content={"error": "Both image and mask are required"}
            )

        # Remove data URI prefix if present
        if image_b64.startswith('data:'):
            image_b64 = image_b64.split(',')[1]
            logger.info("Removed data URI prefix from image")
        if mask_b64.startswith('data:'):
            mask_b64 = mask_b64.split(',')[1]
            logger.info("Removed data URI prefix from mask")

        logger.info(f"Cleaned image b64 length: {len(image_b64)}")
        logger.info(f"Cleaned mask b64 length: {len(mask_b64)}")

        # Decode base64
        try:
            image_data = base64.b64decode(image_b64)
            logger.info(f"Decoded image data: {len(image_data)} bytes")
        except Exception as e:
            logger.error(f"Failed to decode image base64: {e}")
            return JSONResponse(
                status_code=400,
                content={"error": f"Invalid image base64: {str(e)}"}
            )

        try:
            mask_data = base64.b64decode(mask_b64)
            logger.info(f"Decoded mask data: {len(mask_data)} bytes")
        except Exception as e:
            logger.error(f"Failed to decode mask base64: {e}")
            return JSONResponse(
                status_code=400,
                content={"error": f"Invalid mask base64: {str(e)}"}
            )

        # Get image service
        logger.info("Getting image service...")
        image_service = get_image_service()
        logger.info("Image service obtained")

        # Perform inpainting
        logger.info("Calling inpaint_image...")
        result_bytes = image_service.inpaint_image(
            image_data=image_data,
            mask_data=mask_data,
            seed=seed
        )

        if result_bytes is None:
            logger.error("inpaint_image returned None")
            return JSONResponse(
                status_code=500,
                content={"error": "Inpainting failed"}
            )

        logger.info(f"Inpainting successful, result size: {len(result_bytes)} bytes")

        # Convert to base64
        result_b64 = base64.b64encode(result_bytes).decode('utf-8')
        data_uri = f"data:image/png;base64,{result_b64}"

        logger.info(f"Returning result, data URI length: {len(data_uri)}")
        return JSONResponse(content={
            "success": True,
            "imageData": data_uri
        })

    except Exception as e:
        logger.error(f"Error in inpainting: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/remove-background-base64")
async def remove_background_base64(request: Request):
    """Remove image background using rembg. Returns transparent PNG as base64."""
    try:
        try:
            from rembg import remove as rembg_remove
        except ImportError:
            return JSONResponse(
                status_code=503,
                content={"error": "rembg is not installed. Run: pip install rembg"}
            )

        import io
        from PIL import Image as PILImage

        data = await request.json()
        image_data_uri = data.get("image", "")
        strength = max(0, min(100, int(data.get("strength", 100))))

        image_b64 = image_data_uri.split(",", 1)[1] if "," in image_data_uri else image_data_uri
        img_bytes = base64.b64decode(image_b64)
        img = PILImage.open(io.BytesIO(img_bytes)).convert("RGBA")

        result = rembg_remove(img)

        # Blend with original based on strength (100 = full removal, 0 = no removal)
        if strength < 100:
            # Get the alpha channel from the result and blend it toward fully opaque
            r_r, r_g, r_b, r_a = result.split()
            import numpy as np
            alpha_arr = np.array(r_a, dtype=np.float32)
            # Blend alpha toward 255 (opaque) based on inverse strength
            blend_factor = strength / 100.0
            blended_alpha = alpha_arr * blend_factor + 255.0 * (1.0 - blend_factor)
            r_a = PILImage.fromarray(blended_alpha.astype(np.uint8))
            result = PILImage.merge("RGBA", (r_r, r_g, r_b, r_a))

        output = io.BytesIO()
        result.save(output, format="PNG")
        output.seek(0)

        result_b64 = base64.b64encode(output.read()).decode("utf-8")
        return JSONResponse(content={"success": True, "imageData": f"data:image/png;base64,{result_b64}"})

    except Exception as e:
        logger.error(f"Background removal error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/image-service/status")
async def get_image_service_status():
    """
    Check status of image generation services
    
    Returns:
    {
        "sdxl": {"initialized": true/false},
        "iopaint": {"running": true/false, "port": 8080}
    }
    """
    try:
        image_service = get_image_service()
        
        return JSONResponse(content={
            "sdxl": {
                "initialized": image_service.pipeline is not None,
                "model_key": getattr(image_service, 'model_key', None),
            },
            "iopaint": {
                "running": image_service.is_iopaint_running(),
                "port": image_service.iopaint_port
            }
        })
        
    except Exception as e:
        logger.error(f"Error checking image service status: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


@app.post("/api/image-service/preload")
async def preload_image_pipeline():
    """Preload the image generation pipeline in the background so first request is fast."""
    try:
        image_service = get_image_service()
        if image_service.pipeline is not None:
            return JSONResponse(content={"status": "already_loaded"})
        success = image_service.initialize_pipeline()
        return JSONResponse(content={
            "status": "loaded" if success else "failed"
        })
    except Exception as e:
        logger.error(f"Error preloading image pipeline: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/image-service/start-iopaint")
async def start_iopaint_service():
    """
    Manually start IOPaint service
    
    Returns:
    {
        "success": true/false,
        "message": "status message"
    }
    """
    try:
        image_service = get_image_service()
        
        if image_service.is_iopaint_running():
            return JSONResponse(content={
                "success": True,
                "message": "IOPaint already running"
            })
        
        success = image_service.start_iopaint()
        
        if success:
            return JSONResponse(content={
                "success": True,
                "message": "IOPaint started successfully"
            })
        else:
            return JSONResponse(
                status_code=500,
                content={
                    "success": False,
                    "message": "Failed to start IOPaint"
                }
            )
        
    except Exception as e:
        logger.error(f"Error starting IOPaint: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )


# ── Bulk Export / Import ──────────────────────────────────────────────────────

@app.get("/api/export-data")
async def export_data(categories: str = ""):
    """Export selected data categories.
    categories is a comma-separated string of: chats, lesson_plans, quizzes, rubrics,
    kindergarten, multigrade, cross_curricular, worksheets, images, students, settings
    """
    cats = [c.strip() for c in categories.split(",") if c.strip()]
    result: dict = {}

    if "chats" in cats:
        try:
            memory = get_chat_memory()
            result["chats"] = memory.get_all_chats_with_messages()
        except Exception:
            result["chats"] = []

    if "lesson_plans" in cats:
        try:
            plans = []
            if os.path.exists(LESSON_PLAN_HISTORY_FILE):
                with open(LESSON_PLAN_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    plans = json.load(f)
            result["lesson_plans"] = plans
        except Exception:
            result["lesson_plans"] = []

    if "quizzes" in cats:
        result["quizzes"] = load_json_data("quiz_history.json")

    if "rubrics" in cats:
        result["rubrics"] = load_json_data("rubric_history.json")

    if "kindergarten" in cats:
        result["kindergarten"] = load_json_data("kindergarten_history.json")

    if "multigrade" in cats:
        result["multigrade"] = load_json_data("multigrade_history.json")

    if "cross_curricular" in cats:
        result["cross_curricular"] = load_json_data("cross_curricular_history.json")

    if "worksheets" in cats:
        result["worksheets"] = load_json_data("worksheet_history.json")

    if "images" in cats:
        result["images"] = load_json_data("images_history.json")

    if "presentations" in cats:
        result["presentations"] = load_json_data("presentation_history.json")

    if "achievements" in cats:
        try:
            from routes.achievements import achievement_service
            teacher_id = "default"  # Will be expanded with teacher-specific export
            earned = achievement_service.get_earned_achievements(teacher_id)
            # Extract just the all_earned list for portability
            result["achievements"] = earned.get("all_earned", []) if isinstance(earned, dict) else []
        except Exception as e:
            logging.warning(f"Failed to export achievements: {e}")
            result["achievements"] = []

    if "milestones" in cats:
        try:
            from routes.milestones import get_milestone_db
            teacher_id = "default"  # Will be expanded with teacher-specific export
            db = get_milestone_db()
            milestones = db.get_milestones(teacher_id) if hasattr(db, 'get_milestones') else []
            result["milestones"] = milestones
        except Exception as e:
            logging.warning(f"Failed to export milestones: {e}")
            result["milestones"] = []

    if "students" in cats:
        try:
            result["students"] = student_service.list_students()
        except Exception:
            result["students"] = []

    if "calendar" in cats:
        try:
            result["calendar"] = {
                "reminders": student_service.list_test_reminders(),
            }
        except Exception as e:
            logging.warning(f"Failed to export calendar: {e}")
            result["calendar"] = {"reminders": []}

    return {
        "exportDate": datetime.now().isoformat(),
        "version": "1.0.0",
        "categories": cats,
        "data": result
    }


@app.post("/api/import-data")
async def import_data(payload: dict):
    """Import data from a previously exported bundle.
    payload.categories: list of category keys to import
    payload.data: dict mapping category key -> array of records
    """
    cats = payload.get("categories", [])
    data = payload.get("data", {})
    imported: dict = {}
    errors: list = []

    if "chats" in cats and "chats" in data:
        try:
            memory = get_chat_memory()
            count = 0
            for chat in data["chats"]:
                cid = chat.get("id") or str(uuid.uuid4())
                title = chat.get("title", "Imported Chat")
                messages = chat.get("messages", [])
                memory.ensure_chat(cid, title)
                memory.update_chat_title(cid, title)
                if messages:
                    memory.save_messages_bulk(cid, messages)
                count += 1
            imported["chats"] = count
        except Exception as e:
            errors.append(f"chats: {e}")

    if "lesson_plans" in cats and "lesson_plans" in data:
        try:
            existing = []
            if os.path.exists(LESSON_PLAN_HISTORY_FILE):
                with open(LESSON_PLAN_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    existing = json.load(f)
            existing_ids = {h.get("id") for h in existing}
            added = 0
            for plan in data["lesson_plans"]:
                if plan.get("id") not in existing_ids:
                    existing.append(plan)
                    added += 1
            with open(LESSON_PLAN_HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(existing, f, indent=2)
            imported["lesson_plans"] = added
        except Exception as e:
            errors.append(f"lesson_plans: {e}")

    # Generic JSON history imports
    json_map = {
        "quizzes": "quiz_history.json",
        "rubrics": "rubric_history.json",
        "kindergarten": "kindergarten_history.json",
        "multigrade": "multigrade_history.json",
        "cross_curricular": "cross_curricular_history.json",
        "worksheets": "worksheet_history.json",
        "images": "images_history.json",
        "presentations": "presentation_history.json",
    }
    for cat_key, filename in json_map.items():
        if cat_key in cats and cat_key in data:
            try:
                existing = load_json_data(filename)
                existing_ids = {h.get("id") for h in existing}
                added = 0
                for item in data[cat_key]:
                    if item.get("id") not in existing_ids:
                        existing.append(item)
                        added += 1
                save_json_data(filename, existing)
                imported[cat_key] = added
            except Exception as e:
                errors.append(f"{cat_key}: {e}")

    # Import achievements
    if "achievements" in cats and "achievements" in data:
        try:
            from achievement_service import _get_conn
            teacher_id = "default"  # Will be expanded with teacher-specific import
            count = 0
            for achievement in data["achievements"]:
                try:
                    conn = _get_conn()
                    conn.execute(
                        "INSERT OR IGNORE INTO earned_achievements (teacher_id, achievement_id, earned_at) VALUES (?, ?, ?)",
                        (teacher_id, achievement.get("achievement_id"), achievement.get("earned_at"))
                    )
                    conn.commit()
                    conn.close()
                    count += 1
                except Exception:
                    pass  # skip duplicates
            imported["achievements"] = count
        except Exception as e:
            errors.append(f"achievements: {e}")

    # Import milestones
    if "milestones" in cats and "milestones" in data:
        try:
            from routes.milestones import get_milestone_db
            teacher_id = "default"  # Will be expanded with teacher-specific import
            db = get_milestone_db()
            count = 0
            for milestone in data["milestones"]:
                try:
                    db.create_milestone(
                        teacher_id=teacher_id,
                        topic_id=milestone.get("topic_id"),
                        topic_title=milestone.get("topic_title", ""),
                        grade=milestone.get("grade", ""),
                        subject=milestone.get("subject", ""),
                        strand=milestone.get("strand", ""),
                        route=milestone.get("route", ""),
                        status=milestone.get("status", "not_started")
                    )
                    count += 1
                except Exception:
                    pass  # skip duplicates
            imported["milestones"] = count
        except Exception as e:
            errors.append(f"milestones: {e}")

    if "students" in cats and "students" in data:
        try:
            count = 0
            for student in data["students"]:
                try:
                    student_service.create_student(student)
                    count += 1
                except Exception:
                    pass  # skip duplicates
            imported["students"] = count
        except Exception as e:
            errors.append(f"students: {e}")

    if "calendar" in cats and "calendar" in data:
        try:
            cal_data = data["calendar"]
            count = 0
            for reminder in cal_data.get("reminders", []):
                try:
                    student_service.create_test_reminder(reminder)
                    count += 1
                except Exception:
                    pass  # skip duplicates
            imported["calendar"] = count
        except Exception as e:
            errors.append(f"calendar: {e}")

    return {
        "success": len(errors) == 0,
        "imported": imported,
        "errors": errors
    }


@app.post("/api/factory-reset")
async def factory_reset():
    """Wipe all app data: databases, history files, and cached state.
    After this, the app behaves as if opened for the first time."""
    errors = []
    deleted = []

    data_dir = get_data_directory()

    # 1. Delete all JSON history files in data directory
    json_files = [
        "quiz_history.json",
        "rubric_history.json",
        "kindergarten_history.json",
        "multigrade_history.json",
        "cross_curricular_history.json",
        "worksheet_history.json",
        "images_history.json",
    ]
    for fname in json_files:
        fpath = data_dir / fname
        if fpath.exists():
            try:
                os.remove(fpath)
                deleted.append(str(fpath))
            except Exception as e:
                errors.append(f"Failed to delete {fname}: {e}")

    # 2. Delete lesson_plan_history.json (stored in backend dir)
    if os.path.exists(LESSON_PLAN_HISTORY_FILE):
        try:
            os.remove(LESSON_PLAN_HISTORY_FILE)
            deleted.append(LESSON_PLAN_HISTORY_FILE)
        except Exception as e:
            errors.append(f"Failed to delete lesson_plan_history.json: {e}")

    # 2b. Delete lesson_drafts.json (stored in backend dir)
    if os.path.exists(LESSON_DRAFTS_FILE):
        try:
            os.remove(LESSON_DRAFTS_FILE)
            deleted.append(LESSON_DRAFTS_FILE)
        except Exception as e:
            errors.append(f"Failed to delete lesson_drafts.json: {e}")

    # 3. Delete chat_history.json (legacy, stored in backend dir)
    if os.path.exists(CHAT_HISTORY_FILE):
        try:
            os.remove(CHAT_HISTORY_FILE)
            deleted.append(CHAT_HISTORY_FILE)
        except Exception as e:
            errors.append(f"Failed to delete chat_history.json: {e}")

    # 4. Delete SQLite databases (close singletons first)
    import chat_memory as _cm
    from milestones.milestone_db import _milestone_db as _mdb_ref
    import milestones.milestone_db as _mdb_mod

    # Close and delete chat_memory.db
    try:
        if _cm._instance is not None:
            _cm._instance = None
        db_path = data_dir / "chat_memory.db"
        if db_path.exists():
            os.remove(db_path)
            deleted.append(str(db_path))
    except Exception as e:
        errors.append(f"Failed to delete chat_memory.db: {e}")

    # Close and delete milestones.db
    try:
        _mdb_mod._milestone_db = None
        db_path = data_dir / "milestones.db"
        if db_path.exists():
            os.remove(db_path)
            deleted.append(str(db_path))
    except Exception as e:
        errors.append(f"Failed to delete milestones.db: {e}")

    # Close and delete students.db
    try:
        db_path = data_dir / "students.db"
        if db_path.exists():
            os.remove(db_path)
            deleted.append(str(db_path))
    except Exception as e:
        errors.append(f"Failed to delete students.db: {e}")

    logger.info(f"Factory reset completed. Deleted: {deleted}. Errors: {errors}")
    return {
        "success": len(errors) == 0,
        "deleted": deleted,
        "errors": errors
    }


# ============================================================================
# PERFORMANCE METRICS ENDPOINTS
# ============================================================================

@app.get("/api/metrics/summary")
async def metrics_summary():
    """Get aggregated performance summary with system specs."""
    try:
        collector = get_metrics_collector()
        return collector.get_summary()
    except Exception as e:
        logger.error(f"Error getting metrics summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/metrics/history")
async def metrics_history(type: str = "text", limit: int = 100, task_type: Optional[str] = None):
    """Get recent metrics history. type=text|image"""
    try:
        collector = get_metrics_collector()
        if type == "image":
            return {"metrics": collector.get_image_history(limit=limit)}
        else:
            return {"metrics": collector.get_inference_history(limit=limit, task_type=task_type)}
    except Exception as e:
        logger.error(f"Error getting metrics history: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/metrics/export")
async def metrics_export():
    """Export full performance report for documentation."""
    try:
        collector = get_metrics_collector()
        return collector.export_report()
    except Exception as e:
        logger.error(f"Error exporting metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/metrics/system-specs")
async def metrics_system_specs():
    """Get system hardware specs."""
    try:
        collector = get_metrics_collector()
        return collector.get_system_specs()
    except Exception as e:
        logger.error(f"Error getting system specs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/metrics/live-stats")
async def metrics_live_stats():
    """Get real-time CPU and RAM usage (system-wide + app process)."""
    try:
        collector = get_metrics_collector()
        return collector.get_live_stats()
    except Exception as e:
        logger.error(f"Error getting live stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/metrics/clear")
async def metrics_clear():
    """Clear all metrics data."""
    try:
        collector = get_metrics_collector()
        collector.clear_metrics()
        return {"success": True}
    except Exception as e:
        logger.error(f"Error clearing metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/logs/recent")
async def logs_recent(limit: int = 100):
    """Return recent backend log entries for support ticket diagnostics."""
    try:
        entries = _log_ring.recent(min(limit, 200))
        return {"logs": entries}
    except Exception as e:
        logger.error(f"Error fetching logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Test Reminders ─────────────────────────────────────────────────────────────

@app.get("/api/reminders")
async def list_reminders():
    return student_service.list_test_reminders()


@app.post("/api/reminders")
async def create_reminder(request: Request):
    data = await request.json()
    if not data.get("test_date"):
        raise HTTPException(status_code=400, detail="test_date is required")
    return student_service.save_test_reminder(data)


@app.get("/api/reminders/{reminder_id}")
async def get_reminder(reminder_id: str):
    r = student_service.get_test_reminder(reminder_id)
    if not r:
        raise HTTPException(status_code=404, detail="Reminder not found")
    return r


@app.delete("/api/reminders/{reminder_id}")
async def delete_reminder(reminder_id: str):
    if not student_service.delete_test_reminder(reminder_id):
        raise HTTPException(status_code=404, detail="Reminder not found")
    return {"success": True}


@app.get("/api/calendar/export.ics")
async def export_calendar_ics(request: Request):
    """Export all tasks and test reminders as a single .ics calendar file."""
    from datetime import datetime as dt

    reminders = student_service.list_test_reminders()

    # Also load tasks from the request if provided, or from Electron storage
    # Tasks are stored client-side, so we accept them as a query param
    tasks_json = request.query_params.get("tasks", "[]")
    try:
        tasks = json.loads(tasks_json)
    except Exception:
        tasks = []

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//OECS Learning Hub//Test Reminders//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:OECS Learning Hub",
    ]

    def _ics_date(date_str: str, time_str: str | None = None) -> str:
        """Convert date + optional time to ICS format."""
        if time_str:
            try:
                d = dt.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                return d.strftime("%Y%m%dT%H%M%S")
            except Exception:
                pass
        try:
            d = dt.strptime(date_str, "%Y-%m-%d")
            return d.strftime("%Y%m%d")
        except Exception:
            return date_str.replace("-", "")

    def _escape(text: str) -> str:
        return text.replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")

    # Add test reminders
    for r in reminders:
        uid = r.get("id", str(uuid.uuid4()))
        dtstart = _ics_date(r["test_date"], r.get("test_time"))
        summary = _escape(r.get("title", "Test"))
        desc = _escape(r.get("description", ""))
        subject = r.get("subject", "")
        grade = r.get("grade_level", "")
        if subject or grade:
            desc = f"{subject} - Grade {grade}\\n{desc}".strip("\\n")

        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}@oecs-learning-hub")
        if "T" in dtstart:
            lines.append(f"DTSTART:{dtstart}")
            # 1 hour duration for timed events
            try:
                start = dt.strptime(dtstart, "%Y%m%dT%H%M%S")
                end = start.replace(hour=start.hour + 1) if start.hour < 23 else start.replace(hour=23, minute=59)
                lines.append(f"DTEND:{end.strftime('%Y%m%dT%H%M%S')}")
            except Exception:
                pass
        else:
            lines.append(f"DTSTART;VALUE=DATE:{dtstart}")
            lines.append(f"DTEND;VALUE=DATE:{dtstart}")
        lines.append(f"SUMMARY:{summary}")
        if desc:
            lines.append(f"DESCRIPTION:{desc}")
        rtype = r.get("type", "test")
        lines.append(f"CATEGORIES:{rtype.upper()}")
        # Add alarm 1 day before
        lines.append("BEGIN:VALARM")
        lines.append("TRIGGER:-P1D")
        lines.append("ACTION:DISPLAY")
        lines.append(f"DESCRIPTION:Reminder: {summary}")
        lines.append("END:VALARM")
        lines.append("END:VEVENT")

    # Add tasks
    for t in tasks:
        if not t.get("date"):
            continue
        uid = t.get("id", str(uuid.uuid4()))
        dtstart = _ics_date(t["date"])
        summary = _escape(t.get("title", "Task"))
        desc = _escape(t.get("description", ""))
        completed = t.get("completed", False)

        lines.append("BEGIN:VTODO")
        lines.append(f"UID:task-{uid}@oecs-learning-hub")
        lines.append(f"DTSTART;VALUE=DATE:{dtstart}")
        lines.append(f"DUE;VALUE=DATE:{dtstart}")
        lines.append(f"SUMMARY:{summary}")
        if desc:
            lines.append(f"DESCRIPTION:{desc}")
        priority_map = {"urgent": 1, "high": 3, "medium": 5, "low": 9}
        p = priority_map.get(t.get("priority", "medium"), 5)
        lines.append(f"PRIORITY:{p}")
        if completed:
            lines.append("STATUS:COMPLETED")
            lines.append("PERCENT-COMPLETE:100")
        else:
            lines.append("STATUS:NEEDS-ACTION")
        lines.append("END:VTODO")

    lines.append("END:VCALENDAR")

    ics_content = "\r\n".join(lines)
    return Response(
        content=ics_content,
        media_type="text/calendar",
        headers={"Content-Disposition": "attachment; filename=oecs-calendar.ics"}
    )


# ── Scan-Grading Endpoints ──────────────────────────────────────────────────

@app.post("/api/generate-qr")
async def generate_qr(request: Request):
    """Generate a QR code encoding both the document ID and student ID.

    Expects JSON body:
    {
      "doc_type": "quiz" | "worksheet",
      "doc_id": "quiz_xxx" | "worksheet_xxx",
      "student_id": "JD12345"
    }

    Returns JSON: { "qr_base64": "<base64 PNG>" }
    """
    from qr_service import generate_page_qr

    body = await request.json()
    doc_type_raw = body.get("doc_type", "quiz")
    doc_id = body.get("doc_id", "")
    student_id = body.get("student_id", "")

    if not doc_id or not student_id:
        return JSONResponse({"error": "doc_id and student_id are required"}, status_code=400)

    qr_type = "q" if doc_type_raw == "quiz" else "w"
    # version_hash "0000" for non-randomized individual exports
    qr_bytes = generate_page_qr(qr_type, doc_id, student_id, "0000")
    qr_b64 = base64.b64encode(qr_bytes).decode()

    return JSONResponse({"qr_base64": qr_b64})


@app.post("/api/export-class-pack")
async def export_class_pack(request: Request):
    """Export a complete class set with QR codes for scan-grading.

    Expects JSON body:
    {
      "doc_type": "quiz" | "worksheet",
      "doc_id": "quiz_xxx" | "worksheet_xxx",
      "student_versions": [
        {"student_id": "JD12345", "name": "John Doe", "question_order": [2,0,4,1,3]},
        ...
      ],
      "format": "pdf" | "docx",
      "raw_html_per_student": {"JD12345": "<html>...</html>", ...}  // pre-rendered HTML per student
    }

    Returns a ZIP containing individual student files + teacher answer key + answer_regions.json
    """
    import zipfile
    from qr_service import generate_page_qr, compute_version_hash
    from export_utils import export_to_pdf, export_to_docx, add_alignment_markers_to_html, inject_qr_into_html

    body = await request.json()
    doc_type = body["doc_type"]
    doc_id = body["doc_id"]
    student_versions = body["student_versions"]
    fmt = body.get("format", "pdf")
    raw_html_map = body.get("raw_html_per_student", {})

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sv in student_versions:
            sid = sv["student_id"]
            name = sv.get("name", sid)
            q_order = sv["question_order"]
            v_hash = compute_version_hash(sid, q_order)

            # Save instance to DB
            if doc_type == "quiz":
                student_service.save_quiz_instance(
                    quiz_id=doc_id, student_id=sid,
                    question_order=q_order, version_hash=v_hash,
                    class_name=sv.get("class_name", "")
                )
            else:
                student_service.save_worksheet_instance(
                    worksheet_id=doc_id, student_id=sid,
                    question_order=q_order, version_hash=v_hash,
                    option_maps=sv.get("option_maps"),
                    shuffled_column_b=sv.get("shuffled_column_b"),
                    shuffled_word_bank=sv.get("shuffled_word_bank"),
                    package_id=sv.get("package_id"),
                    class_name=sv.get("class_name", "")
                )

            # Generate QR code
            qr_type = "q" if doc_type == "quiz" else "w"
            qr_bytes = generate_page_qr(qr_type, doc_id, sid, v_hash)
            qr_b64 = base64.b64encode(qr_bytes).decode()

            # Get or build HTML
            html = raw_html_map.get(sid, "")
            if html:
                html = inject_qr_into_html(html, qr_b64)
                html = add_alignment_markers_to_html(html)

                # Export to requested format
                data_payload = {"rawHtml": html}
                if fmt == "pdf":
                    file_bytes = export_to_pdf(data_payload, title=f"{name}")
                    ext = "pdf"
                else:
                    file_bytes = export_to_docx(data_payload, title=f"{name}")
                    ext = "docx"

                safe_name = name.replace(" ", "_").replace("/", "_")
                zf.writestr(f"{safe_name}_{sid}.{ext}", file_bytes)

    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=class_pack_{doc_id}.zip"}
    )


@app.post("/api/scan-grade-auto")
async def scan_grade_auto(
    files: List[UploadFile] = File(...),
):
    """Auto-grade scanned worksheets/quizzes using QR codes.

    Pipeline:
    1. QR decode → identify doc + student + version
    2. Fetch answer key + instance
    3. If answer region template exists → pixel detection; else → OCR fallback
    4. Re-map answers if randomized
    5. Grade against answer key
    6. Save grade to DB
    """
    from qr_service import extract_qr_from_scan, remap_answers, remap_mc_option
    from image_alignment import align_scanned_page
    from bubble_detector import detect_answers_from_template

    # Pre-read files
    file_data = []
    for upload in files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    results = []

    for file_bytes, filename in file_data:
        result = {
            "file_name": filename,
            "qr_status": "not_found",
            "student_name": None,
            "student_id": None,
            "doc_type": None,
            "doc_id": None,
            "score": 0,
            "total_points": 0,
            "percentage": 0.0,
            "letter_grade": "F",
            "answers": {},
            "error": None
        }

        try:
            # Step 1: QR decode
            qr_data = extract_qr_from_scan(file_bytes)
            if not qr_data:
                result["qr_status"] = "not_found"
                result["error"] = "No QR code detected on this page"
                results.append(result)
                continue

            result["qr_status"] = "decoded"
            result["doc_type"] = qr_data["type"]
            result["doc_id"] = qr_data["doc_id"]
            result["student_id"] = qr_data["student_id"]

            # Step 2: Fetch answer key
            if qr_data["type"] == "quiz":
                answer_key = student_service.get_quiz_answer_key(qr_data["doc_id"])
            else:
                answer_key = student_service.get_worksheet_answer_key(qr_data["doc_id"])

            if not answer_key:
                result["error"] = f"Answer key not found for {qr_data['doc_id']}"
                results.append(result)
                continue

            # Step 3: Get instance (for randomized order)
            instance = student_service.get_instance_by_qr(
                qr_data["type"], qr_data["doc_id"],
                qr_data["student_id"], qr_data["version_hash"]
            )

            # Step 4: Get student info
            student = student_service.get_student(qr_data["student_id"])
            if student:
                result["student_name"] = student["full_name"]

            # Step 5: Check for answer region template
            template = student_service.get_answer_region_template(qr_data["doc_id"])

            questions = answer_key["questions"]
            detected_answers = {}

            if template:
                # Pixel-based detection using template coordinates
                aligned_img, align_info = align_scanned_page(file_bytes)
                if aligned_img is not None:
                    detections = detect_answers_from_template(aligned_img, template["regions"])
                    for det in detections:
                        q_idx = det["question_index"]
                        if det["status"] == "detected":
                            detected_answers[q_idx] = det["selected"]
                        elif det["status"] == "needs_ocr":
                            # Fallback to OCR for open-answer questions
                            try:
                                import ocr_service
                                if ocr_service.is_ocr_available() and det.get("cropped_image") is not None:
                                    import cv2
                                    _, img_encoded = cv2.imencode('.png', det["cropped_image"])
                                    ocr_text = ocr_service.extract_text(img_encoded.tobytes())
                                    detected_answers[q_idx] = ocr_text.strip()
                            except Exception:
                                detected_answers[q_idx] = ""
            else:
                # Fallback: full-page OCR (existing pipeline)
                try:
                    import ocr_service
                    if ocr_service.is_ocr_available():
                        ocr_text = ocr_service.extract_text_for_grading(file_bytes)
                        # Parse OCR output into detected answers (simple line-based)
                        lines = ocr_text.strip().split('\n')
                        for i, line in enumerate(lines):
                            detected_answers[i] = line.strip()
                except Exception as e:
                    result["error"] = f"OCR fallback failed: {str(e)}"
                    results.append(result)
                    continue

            # Step 6: Re-map if randomized
            if instance and instance.get("question_order"):
                q_order = instance["question_order"]
                answer_list = [detected_answers.get(i, "") for i in range(len(q_order))]
                remapped = remap_answers(answer_list, q_order)

                # Also remap MC options if shuffled
                option_maps = instance.get("option_maps")
                if option_maps:
                    for q_idx, answer in remapped.items():
                        q_map = option_maps.get(str(q_idx))
                        if q_map and answer and len(answer) == 1 and answer.upper() in 'ABCDE':
                            remapped[q_idx] = remap_mc_option(answer, q_map)

                detected_answers = remapped

            # Step 7: Grade against answer key
            score = 0
            total = len(questions)
            answer_details = {}

            for i, q in enumerate(questions):
                student_answer = detected_answers.get(i, "")
                correct = q.get("correctAnswer", "")
                q_type = q.get("type", "")

                is_correct = False
                if q_type in ("multiple-choice", "true-false"):
                    if isinstance(correct, int):
                        # correctAnswer is option index
                        correct_letter = chr(65 + correct)
                        is_correct = student_answer.upper() == correct_letter
                    elif isinstance(correct, str):
                        is_correct = student_answer.strip().lower() == correct.strip().lower()
                elif q_type in ("fill-blank", "short-answer"):
                    is_correct = student_answer.strip().lower() == str(correct).strip().lower()

                if is_correct:
                    score += 1

                answer_details[str(i)] = {
                    "student_answer": student_answer,
                    "correct_answer": correct if isinstance(correct, str) else chr(65 + correct) if isinstance(correct, int) else str(correct),
                    "is_correct": is_correct
                }

            percentage = round((score / total * 100) if total > 0 else 0, 1)
            letter_grade = student_service.get_letter_grade(percentage)

            result["score"] = score
            result["total_points"] = total
            result["percentage"] = percentage
            result["letter_grade"] = letter_grade
            result["answers"] = answer_details

            # Step 8: Save grade to DB
            if result["student_id"]:
                grade_data = {
                    "student_id": result["student_id"],
                    "score": score,
                    "total_points": total,
                    "percentage": percentage,
                    "letter_grade": letter_grade,
                    "answers": answer_details,
                }
                if qr_data["type"] == "quiz":
                    grade_data["quiz_title"] = answer_key.get("quiz_title", "")
                    grade_data["subject"] = answer_key.get("subject", "")
                    grade_data["quiz_id"] = qr_data["doc_id"]
                    grade_data["instance_id"] = instance["id"] if instance else None
                    student_service.save_quiz_grade(grade_data)
                else:
                    grade_data["worksheet_title"] = answer_key.get("worksheet_title", "")
                    grade_data["subject"] = answer_key.get("subject", "")
                    grade_data["worksheet_id"] = qr_data["doc_id"]
                    grade_data["instance_id"] = instance["id"] if instance else None
                    student_service.save_worksheet_grade(grade_data)

        except Exception as e:
            logger.error(f"Scan grade error for {filename}: {e}", exc_info=True)
            result["error"] = str(e)

        results.append(result)

    # Summary
    graded = [r for r in results if r["error"] is None and r["qr_status"] == "decoded"]
    avg = round(sum(r["percentage"] for r in graded) / len(graded), 1) if graded else 0.0

    return JSONResponse(content={
        "results": results,
        "summary": {
            "total": len(results),
            "graded": len(graded),
            "failed": len(results) - len(graded),
            "class_average": avg
        }
    })


@app.post("/api/scan-grade-auto-stream")
async def scan_grade_auto_stream(
    files: List[UploadFile] = File(...),
):
    """SSE streaming version of scan-grade-auto for real-time progress."""
    from qr_service import extract_qr_from_scan, remap_answers, remap_mc_option
    from image_alignment import align_scanned_page
    from bubble_detector import detect_answers_from_template

    file_data = []
    for upload in files:
        content = await upload.read()
        file_data.append((content, upload.filename or 'unknown'))

    async def event_stream():
        graded_count = 0
        failed_count = 0
        total_pct = 0.0

        for file_bytes, filename in file_data:
            yield f"data: {json.dumps({'event': 'processing', 'file_name': filename, 'step': 'qr_decode', 'message': 'Scanning QR code...'})}\n\n"

            try:
                qr_data = extract_qr_from_scan(file_bytes)
                if not qr_data:
                    yield f"data: {json.dumps({'event': 'error', 'file_name': filename, 'error': 'No QR code detected'})}\n\n"
                    failed_count += 1
                    continue

                student = student_service.get_student(qr_data["student_id"])
                student_name = student["full_name"] if student else qr_data["student_id"]

                yield f"data: {json.dumps({'event': 'qr_decoded', 'file_name': filename, 'student_name': student_name, 'doc_type': qr_data['type'], 'doc_id': qr_data['doc_id']})}\n\n"

                # Fetch answer key
                if qr_data["type"] == "quiz":
                    answer_key = student_service.get_quiz_answer_key(qr_data["doc_id"])
                else:
                    answer_key = student_service.get_worksheet_answer_key(qr_data["doc_id"])

                if not answer_key:
                    doc_id = qr_data["doc_id"]
                    yield f"data: {json.dumps({'event': 'error', 'file_name': filename, 'error': f'Answer key not found for {doc_id}'})}\n\n"
                    failed_count += 1
                    continue

                yield f"data: {json.dumps({'event': 'processing', 'file_name': filename, 'step': 'detecting', 'message': 'Detecting answers...'})}\n\n"

                instance = student_service.get_instance_by_qr(
                    qr_data["type"], qr_data["doc_id"],
                    qr_data["student_id"], qr_data["version_hash"]
                )

                template = student_service.get_answer_region_template(qr_data["doc_id"])
                questions = answer_key["questions"]
                detected_answers = {}

                if template:
                    aligned_img, _ = align_scanned_page(file_bytes)
                    if aligned_img is not None:
                        detections = detect_answers_from_template(aligned_img, template["regions"])
                        for det in detections:
                            if det["status"] == "detected":
                                detected_answers[det["question_index"]] = det["selected"]
                else:
                    try:
                        import ocr_service
                        if ocr_service.is_ocr_available():
                            ocr_text = ocr_service.extract_text_for_grading(file_bytes)
                            for i, line in enumerate(ocr_text.strip().split('\n')):
                                detected_answers[i] = line.strip()
                    except Exception:
                        pass

                # Re-map if randomized
                if instance and instance.get("question_order"):
                    q_order = instance["question_order"]
                    answer_list = [detected_answers.get(i, "") for i in range(len(q_order))]
                    remapped = remap_answers(answer_list, q_order)
                    option_maps = instance.get("option_maps")
                    if option_maps:
                        for q_idx, answer in remapped.items():
                            q_map = option_maps.get(str(q_idx))
                            if q_map and answer and len(answer) == 1 and answer.upper() in 'ABCDE':
                                remapped[q_idx] = remap_mc_option(answer, q_map)
                    detected_answers = remapped

                # Grade
                score = 0
                total = len(questions)
                for i, q in enumerate(questions):
                    student_answer = detected_answers.get(i, "")
                    correct = q.get("correctAnswer", "")
                    if isinstance(correct, int):
                        if student_answer.upper() == chr(65 + correct):
                            score += 1
                    elif isinstance(correct, str):
                        if student_answer.strip().lower() == correct.strip().lower():
                            score += 1

                percentage = round((score / total * 100) if total > 0 else 0, 1)
                letter_grade = student_service.get_letter_grade(percentage)

                # Save
                if qr_data["student_id"]:
                    grade_data = {
                        "student_id": qr_data["student_id"],
                        "score": score, "total_points": total,
                        "percentage": percentage, "letter_grade": letter_grade,
                        "answers": detected_answers,
                    }
                    if qr_data["type"] == "quiz":
                        grade_data["quiz_title"] = answer_key.get("quiz_title", "")
                        grade_data["subject"] = answer_key.get("subject", "")
                        grade_data["quiz_id"] = qr_data["doc_id"]
                        grade_data["instance_id"] = instance["id"] if instance else None
                        student_service.save_quiz_grade(grade_data)
                    else:
                        grade_data["worksheet_title"] = answer_key.get("worksheet_title", "")
                        grade_data["subject"] = answer_key.get("subject", "")
                        grade_data["worksheet_id"] = qr_data["doc_id"]
                        grade_data["instance_id"] = instance["id"] if instance else None
                        student_service.save_worksheet_grade(grade_data)

                yield f"data: {json.dumps({'event': 'graded', 'file_name': filename, 'student_name': student_name, 'score': score, 'total': total, 'percentage': percentage, 'letter_grade': letter_grade})}\n\n"
                graded_count += 1
                total_pct += percentage

            except Exception as e:
                yield f"data: {json.dumps({'event': 'error', 'file_name': filename, 'error': str(e)})}\n\n"
                failed_count += 1

        avg = round(total_pct / graded_count, 1) if graded_count > 0 else 0.0
        yield f"data: {json.dumps({'event': 'complete', 'total': len(file_data), 'graded': graded_count, 'failed': failed_count, 'class_average': avg})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/save-quiz-instances")
async def save_quiz_instances_endpoint(request: Request):
    """Save per-student quiz instances when generating class versions.

    Body: { quiz_id, class_name, students: [{student_id, name, question_order}] }
    """
    body = await request.json()
    quiz_id = body["quiz_id"]
    class_name = body.get("class_name", "")
    students = body["students"]

    saved = []
    for s in students:
        v_hash = student_service.compute_version_hash(s["student_id"], s["question_order"])
        result = student_service.save_quiz_instance(
            quiz_id=quiz_id, student_id=s["student_id"],
            question_order=s["question_order"], version_hash=v_hash,
            class_name=class_name
        )
        saved.append({**result, "student_id": s["student_id"], "name": s.get("name", "")})

    return JSONResponse(content={"saved": saved, "count": len(saved)})


@app.post("/api/save-worksheet-instances")
async def save_worksheet_instances_endpoint(request: Request):
    """Save per-student worksheet instances when generating class versions.

    Body: { worksheet_id, class_name, package_id?, students: [{student_id, name, question_order, ...}] }
    """
    body = await request.json()
    worksheet_id = body["worksheet_id"]
    class_name = body.get("class_name", "")
    package_id = body.get("package_id")
    students = body["students"]

    saved = []
    for s in students:
        v_hash = student_service.compute_version_hash(s["student_id"], s["question_order"])
        result = student_service.save_worksheet_instance(
            worksheet_id=worksheet_id, student_id=s["student_id"],
            question_order=s["question_order"], version_hash=v_hash,
            option_maps=s.get("option_maps"),
            shuffled_column_b=s.get("shuffled_column_b"),
            shuffled_word_bank=s.get("shuffled_word_bank"),
            package_id=package_id, class_name=class_name
        )
        saved.append({**result, "student_id": s["student_id"], "name": s.get("name", "")})

    return JSONResponse(content={"saved": saved, "count": len(saved)})










